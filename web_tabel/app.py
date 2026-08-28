from __future__ import annotations

import datetime as dt
import json
import sqlite3
import os
import io
import sys
from pathlib import Path
from threading import Lock
from urllib.parse import unquote

from fastapi import FastAPI, Request, Response, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import uvicorn
import hmac
import hashlib

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT.parent))
from rtps_common import connect_sqlite, module_role, resolve_user_access

SHARED_DATA_DIR = ROOT.parent / "data"
WEB_SECRET_FILE = SHARED_DATA_DIR / "web_secret.txt"
DB_FILE = ROOT.parent / "base" / "common_database.db"
WEB_USERS_DB = ROOT.parent / "base" / "web_users.db"
SESSION_COOKIE = "rtps_session"
APP_PREFIX = "/tabel"
APP_VERSION = "web-tabel-1.57"
DB_LOCK = Lock()
COMMON_DB_FILE = DB_FILE
MAIN_LOGIN_URL = os.environ.get("MAIN_LOGIN_URL", "http://yrtps.ru/login")

def load_web_secret() -> str:
    if WEB_SECRET_FILE.exists():
        return WEB_SECRET_FILE.read_text(encoding="utf-8").strip()
    return "opYbo6NB8pb7dChYQkmHEvUH6K4hAHjuzi2qEYOC024"

WEB_SECRET = load_web_secret()

def init_db():
    DB_FILE.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(DB_FILE) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("CREATE TABLE IF NOT EXISTS month_hints (y INT, m TEXT, hint TEXT, PRIMARY KEY(y,m))")

        # Add is_excluded column if not exists
        cur.execute("PRAGMA table_info('employees')")
        emp_cols = [c["name"] for c in cur.fetchall()]
        if "is_excluded" not in emp_cols:
            cur.execute("ALTER TABLE employees ADD COLUMN is_excluded INT DEFAULT 0")
            conn.commit()

        
        # Check if we need to migrate timesheet/vacations from r -> tab_num
        cur.execute("PRAGMA table_info('timesheet')")
        cols = [c["name"] for c in cur.fetchall()]
        if "tab_num" not in cols and "r" in cols:
            print("Auto-migrating timesheet and vacations using true tab_num from legacy data...")
            years = cur.execute("SELECT DISTINCT y FROM employees").fetchall()
            for y_row in years:
                year = y_row["y"]
                
                # 1. Timesheet migration
                ts_rows = cur.execute("SELECT m, r, c, v FROM timesheet WHERE y=?", (year,)).fetchall()
                # Find true tab_num mapping for timesheet (stored in c=3)
                ts_r_to_tab = {}
                for row in ts_rows:
                    if row["c"] == 3 and len(str(row["v"])) > 3:
                        ts_r_to_tab[(row["m"], row["r"])] = str(row["v"])
                
                cur.execute("CREATE TABLE IF NOT EXISTS timesheet_new (y INT, m TEXT, tab_num TEXT, c INT, v TEXT, PRIMARY KEY(y,m,tab_num,c))")
                for row in ts_rows:
                    m = row["m"]
                    r = row["r"]
                    c = row["c"]
                    v = row["v"]
                    if c >= 4:  # Only migrate actual days, skipping legacy employee metadata
                        tab_num = ts_r_to_tab.get((m, r))
                        if tab_num:
                            cur.execute("INSERT OR REPLACE INTO timesheet_new (y, m, tab_num, c, v) VALUES (?, ?, ?, ?, ?)", 
                                        (year, m, tab_num, c - 3, v))
                
                # 2. Vacations migration
                vac_rows = cur.execute("SELECT r, c, v FROM vacations WHERE y=?", (year,)).fetchall()
                # Find true tab_num mapping for vacations (stored in c=0)
                vac_r_to_tab = {}
                for row in vac_rows:
                    if row["c"] == 0 and len(str(row["v"])) > 3:
                        vac_r_to_tab[row["r"]] = str(row["v"])
                        
                cur.execute("CREATE TABLE IF NOT EXISTS vacations_new (y INT, tab_num TEXT, c INT, v TEXT, PRIMARY KEY(y,tab_num,c))")
                for row in vac_rows:
                    r = row["r"]
                    c = row["c"]
                    v = row["v"]
                    if c >= 1:  # Skip tab_num column
                        tab_num = vac_r_to_tab.get(r)
                        if tab_num:
                            cur.execute("INSERT OR REPLACE INTO vacations_new (y, tab_num, c, v) VALUES (?, ?, ?, ?)", 
                                        (year, tab_num, c - 1, v))
                                        
            cur.execute("DROP TABLE IF EXISTS timesheet")
            cur.execute("ALTER TABLE timesheet_new RENAME TO timesheet")
            cur.execute("DROP TABLE IF EXISTS vacations")
            cur.execute("ALTER TABLE vacations_new RENAME TO vacations")
            

            conn.commit()

try:
    init_db()
except Exception as e:
    import traceback
    with open(ROOT / "startup_error.log", "w", encoding="utf-8") as f:
        f.write(traceback.format_exc())
    print("STARTUP ERROR:", traceback.format_exc())

def connect() -> sqlite3.Connection:
    return connect_sqlite(DB_FILE)

def connect_common() -> sqlite3.Connection | None:
    if not COMMON_DB_FILE.exists():
        return None
    conn = sqlite3.connect(COMMON_DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def text(value) -> str:
    return "" if value is None else str(value)

def _verify_cookie_fastapi(value: str) -> tuple[str, str, str, str] | None:
    for sep in (":", "|"):
        try:
            parts = value.rsplit(sep, 5)
            if len(parts) == 6:
                user_id, role, modules, safe_name, expiry_text, sig = parts
                payload = f"{user_id}{sep}{role}{sep}{modules}{sep}{safe_name}{sep}{expiry_text}"
            elif len(parts) == 5:
                user_id, role, modules, safe_name, sig = parts
                payload = f"{user_id}{sep}{role}{sep}{modules}{sep}{safe_name}"
                expiry_text = "2000000000"
            else:
                continue
                
            secrets_to_try = [WEB_SECRET]
                
            matched = False
            for secret in secrets_to_try:
                expected = hmac.new(secret.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
                if hmac.compare_digest(expected, sig):
                    matched = True
                    break
                    
            if not matched:
                continue
            if float(expiry_text) < dt.datetime.now().timestamp():
                return None
                
            return unquote(user_id), unquote(role), unquote(modules), unquote(safe_name)
        except Exception:
            continue
    return None

def get_current_session_fastapi(request: Request):
    cookie = request.cookies.get(SESSION_COOKIE)
    if cookie:
        return _verify_cookie_fastapi(cookie)
    return None

def get_mod_role_fastapi(session: tuple[str, str, str, str] | None, module: str) -> str:
    if not session:
        return ""
    username = session[0]
    role = session[1]
    modules = session[2]

    resolved = resolve_user_access(WEB_USERS_DB, username, role, modules)
    if not resolved:
        return ""
    role, modules = resolved
    return module_role(role, modules, module) or ""

app = FastAPI(title="RTPS Tabel")

@app.middleware("http")
async def strip_prefix(request: Request, call_next):
    if request.scope["path"].startswith(APP_PREFIX + "/"):
        request.scope["path"] = request.scope["path"][len(APP_PREFIX):]
    return await call_next(request)

app.mount("/static", StaticFiles(directory=str(ROOT / "static")), name="static")

def json_response(data: dict | list, status_code: int = 200) -> JSONResponse:
    return JSONResponse(content=data, status_code=status_code)

@app.get("/tabel", response_class=HTMLResponse)
@app.get("/", response_class=HTMLResponse)
async def home_route(request: Request):
    session = get_current_session_fastapi(request)
    if not session:
        return RedirectResponse(f"{MAIN_LOGIN_URL}?next=/tabel", status_code=303)
        
    mod_role = get_mod_role_fastapi(session, "tabel")
    if not mod_role:
        return HTMLResponse(content="<meta charset='utf-8'>У вас нет прав для доступа к этому модулю. Обратитесь к администратору. <a href='/'>Вернуться в главное меню</a>", status_code=403)
        
    can_edit = mod_role in ("edit", "editor", "admin")
    
    with open(ROOT / "templates" / "index.html", "r", encoding="utf-8") as f:
        html = f.read()
        
    html = html.replace("{{CAN_EDIT}}", "true" if can_edit else "false")
    html = html.replace("{{APP_PREFIX}}", APP_PREFIX)
    html = html.replace("{{APP_VERSION}}", APP_VERSION.replace("web-tabel-", ""))
    
    response = HTMLResponse(content=html)
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return response
MONTH_NAMES = ["", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]


FIXED_HOLIDAYS = {
    (1, 1), (1, 2), (1, 3), (1, 4), (1, 5), (1, 6), (1, 7), (1, 8),
    (2, 23), (3, 8), (5, 1), (5, 9), (6, 12), (11, 4),
}

def load_system_dates(year: int) -> dict[str, list[tuple[int, int]]]:
    transfer_dates: set[tuple[int, int]] = set()
    holiday_dates: set[tuple[int, int]] = set(FIXED_HOLIDAYS)
    db_path = ROOT.parent / "base" / "common_database.db"
    if not db_path.exists():
        return {
            "transfer": sorted(transfer_dates),
            "holiday": sorted(holiday_dates),
        }

    try:
        import sqlite3
        with sqlite3.connect(db_path) as conn:
            cur = conn.cursor()

            rows = cur.execute(
                "SELECT c, v FROM ts_norms_data WHERE y=? AND c IN (6, 7)",
                (year,),
            ).fetchall()
        for col_idx, raw_text in rows:
            if not raw_text:
                continue
            text = str(raw_text).replace(";", "\n").replace(",", "\n")
            for line in text.splitlines():
                line = line.strip()
                if not line:
                    continue
                parts = line.split(".")
                if len(parts) >= 2:
                    try:
                        d, m = int(parts[0]), int(parts[1])
                        if col_idx == 6:
                            transfer_dates.add((m, d))
                        elif col_idx == 7:
                            holiday_dates.add((m, d))
                    except ValueError:
                        pass
    except Exception:
        pass

    return {
        "transfer": sorted(transfer_dates),
        "holiday": sorted(holiday_dates),
    }

def load_state(year: int, month: int) -> dict:
    with DB_LOCK, connect() as conn:
        cur = conn.cursor()

        # Автоматическая миграция: добавляем колонку exclude_date, если её нет на сервере
        try:
            cur.execute("SELECT exclude_date FROM employees LIMIT 1")
        except sqlite3.OperationalError:
            try:
                cur.execute("ALTER TABLE employees ADD COLUMN exclude_date TEXT DEFAULT ''")
                conn.commit()
            except Exception:
                pass
            try:
                cur.execute("ALTER TABLE employees ADD COLUMN hire_date TEXT DEFAULT ''")
                conn.commit()
            except Exception:
                pass
        
        employees = []
        try:
            emp_rows = cur.execute(
                "SELECT pos, name, tab_num, milk, milk_issue, full_name, milk_note, is_excluded, hire_date, exclude_date FROM employees WHERE y=? ORDER BY rowid", (year,)
            ).fetchall()
            for r in emp_rows:
                employees.append({
                    "pos": text(r["pos"]),
                    "name": text(r["name"]),
                    "tab_num": text(r["tab_num"]),
                    "milk": int(r["milk"] or 0),
                    "milk_issue": int(r["milk_issue"] or 0),
                    "full_name": text(r["full_name"]),
                    "milk_note": text(r["milk_note"]),
                    "is_excluded": int(dict(r).get("is_excluded", 0) or 0),
                    "hire_date": text(dict(r).get("hire_date", "")),
                    "exclude_date": text(dict(r).get("exclude_date", ""))
                })
        except Exception:
            pass

        timesheet = {}
        try:
            m_str = MONTH_NAMES[month] if 1 <= month <= 12 else str(month)
            ts_rows = cur.execute(
                "SELECT tab_num, c, v FROM timesheet WHERE y=? AND m=?", (year, m_str)
            ).fetchall()
            for r in ts_rows:
                timesheet.setdefault(text(r["tab_num"]), {})[int(r["c"])] = text(r["v"])
        except Exception:
            pass
            
        vacations = {}
        try:
            vac_rows = cur.execute("SELECT tab_num, c, v FROM vacations WHERE y=?", (year,)).fetchall()
            for r in vac_rows:
                vacations.setdefault(text(r["tab_num"]), {})[int(r["c"])] = text(r["v"])
        except Exception:
            pass

        ts_norms_data = {}
        try:
            norms_rows = cur.execute("SELECT r, c, v FROM ts_norms_data WHERE y=?", (year,)).fetchall()
            for r in norms_rows:
                ts_norms_data.setdefault(int(r["r"]), {})[int(r["c"])] = text(r["v"])
        except Exception:
            pass

        month_hint = ""
        try:
            hint_row = cur.execute("SELECT v FROM ts_settings WHERE k='month_hint'").fetchone()
            if hint_row:
                month_hint = str(hint_row["v"])
        except Exception:
            pass

    return {
        "system_dates": load_system_dates(year),
        "year": year,
        "month": month,
        "employees": employees,
        "timesheet": timesheet,
        "vacations": vacations,
        "ts_norms_data": ts_norms_data,
        "month_hint": month_hint,
    }

@app.get("/api/debug_startup")
async def debug_startup(request: Request):
    error_file = ROOT / "startup_error.log"
    if error_file.exists():
        return {"error": error_file.read_text(encoding="utf-8")}
    return {"error": "No startup error found."}


import io
import base64
import copy
from fastapi.responses import FileResponse, StreamingResponse
from urllib.parse import quote


from fastapi.responses import HTMLResponse
import calendar

@app.get("/api/export-summary", response_class=HTMLResponse)
async def export_summary(year: int, month: int, type: str):
    import urllib.parse
    type = urllib.parse.unquote(type)
    if ":" in type:
        codes = [type.split(":")[0].strip()]
        title = type.split(":", 1)[1].strip()
    else:
        code_map = {
            "Отпуска": ["О", "ОВ"],
            "Отпуска внеплановые": ["ОВ"],
            "Отпуск б/с": ["ДО"],
            "Учебный отпуск": ["У"],
            "Больничный": ["Б"]
        }
        codes = code_map.get(type, [type])
        title = type

    with DB_LOCK, connect() as conn:
        cur = conn.cursor()

        emp_rows = cur.execute("SELECT DISTINCT name FROM employees WHERE y=? AND name != ''", (year,)).fetchall()
        employees = sorted(list(set(row["name"] for row in emp_rows)))
        
        result = {emp: 0 for emp in employees}
        
        sys_dates = load_system_dates(year)
        holiday_set = set(sys_dates.get("holiday", []))
        transfer_set = set(sys_dates.get("transfer", []))
        
        monthsNames = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
        month_map = {name: i+1 for i, name in enumerate(monthsNames)}
        
        if codes == ["WORK_WEEKEND"]:
            ts_rows = cur.execute("""
                SELECT e.name, t.v, t.m, t.c
                FROM timesheet t
                JOIN employees e ON t.tab_num = e.tab_num AND t.y = e.y
                WHERE t.y=? AND t.v != ''
            """, (year,)).fetchall()
            
            from datetime import date
            ignore_codes = {"В", "B", "О", "ДО", "К", "У", "Б", "БН", "ОВ"}
            
            for row in ts_rows:
                name = row["name"]
                v = str(row["v"]).strip().upper()
                if not v or v in ignore_codes:
                    continue
                    
                m_str = row["m"]
                d = int(row["c"])
                m_int = month_map.get(m_str)
                if not m_int: continue
                
                is_holiday = (m_int, d) in holiday_set
                is_transfer = (m_int, d) in transfer_set
                try:
                    weekday = date(year, m_int, d).weekday()
                    is_weekend = weekday in (5, 6)
                except ValueError:
                    is_weekend = False
                    
                if is_weekend or is_holiday or is_transfer:
                    if name in result:
                        result[name] += 1
        else:
            placeholders = ','.join(['?']*len(codes))
            query = f"""
                SELECT e.name, t.v, t.m, t.c
                FROM timesheet t
                JOIN employees e ON t.tab_num = e.tab_num AND t.y = e.y
                WHERE t.y=? AND t.v IN ({placeholders})
            """
            ts_rows = cur.execute(query, [year] + codes).fetchall()
            
            vacation_days = {}
            if "О" in codes or "ОВ" in codes:
                vac_rows = cur.execute("SELECT e.name, v.c, v.v as val FROM vacations v JOIN employees e ON v.tab_num = e.tab_num AND v.y = e.y WHERE v.y=?", (year,)).fetchall()
                emp_vacs = {}
                for r in vac_rows:
                    emp_vacs.setdefault(r["name"], {})[int(r["c"])] = r["val"]
                
                import datetime as dt
                def parse_vacation_date(date_str, default_year):
                    if not date_str: return None
                    s = str(date_str).strip()
                    if "." in s:
                        p = s.split(".")
                        try:
                            if len(p) == 2: return dt.date(default_year, int(p[1]), int(p[0]))
                            elif len(p) == 3:
                                y = int(p[2])
                                if y < 100: y += 2000
                                return dt.date(y, int(p[1]), int(p[0]))
                        except Exception: pass
                    return None

                for name, vac_data in emp_vacs.items():
                    vacation_days[name] = set()
                    for pair in [(1,2), (5,6), (9,10)]:
                        s_str = vac_data.get(pair[0])
                        e_str = vac_data.get(pair[1])
                        try:
                            s_date = parse_vacation_date(s_str, year)
                            e_date = parse_vacation_date(e_str, year)
                            if s_date and e_date and e_date >= s_date:
                                curr = s_date
                                while curr <= e_date:
                                    if curr.year == year:
                                        vacation_days[name].add((curr.month, curr.day))
                                    curr += dt.timedelta(days=1)
                        except Exception:
                            pass
            
            for row in ts_rows:
                name = row["name"]
                v = row["v"]
                m_str = row["m"]
                d = row["c"]
                
                if v in ("О", "ОВ"):
                    m_int = month_map.get(m_str)
                    if m_int:
                        if name not in vacation_days:
                            vacation_days[name] = set()
                        vacation_days[name].add((m_int, int(d)))
                else:
                    if name in result:
                        result[name] += 1
                        
            if "О" in codes or "ОВ" in codes:
                for name, days_set in vacation_days.items():
                    valid_days = 0
                    for m_int, d_int in days_set:
                        if (m_int, d_int) not in holiday_set:
                            valid_days += 1
                    if name in result:
                        result[name] += valid_days
                    
    total = sum(result.values())
    
    html = f"""
    <html><head><meta charset="utf-8"><title>Сводка: {title}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        table {{ border-collapse: collapse; width: 600px; margin-top: 20px; }}
        th, td {{ border: 1px solid #ccc; padding: 8px; text-align: left; }}
        th {{ background-color: #f0f0f0; }}
        .center {{ text-align: center; }}
        .right {{ text-align: right; font-weight: bold; }}
    </style>
    </head><body>
    <h2 class="center">{title} за {year} год</h2>
    <div class="center" style="color: #666;">Коды: {', '.join(codes)}</div>
    <table>
        <tr><th style="width: 50px;">№</th><th>ФИО</th><th style="width: 100px;">Дней</th></tr>
    """
    for idx, emp in enumerate(employees, 1):
        html += f"<tr><td class='center'>{idx}</td><td>{emp}</td><td class='center'>{result[emp]}</td></tr>"
        
    html += f"<tr><td colspan='2' class='right'>Итого:</td><td class='center'><b>{total}</b></td></tr>"
    html += "</table></body></html>"
    return HTMLResponse(content=html)

@app.get("/api/export-milk")
async def export_milk(year: int, month: int, type: str):
    import urllib.parse
    import tempfile
    import os
    import openpyxl
    type = urllib.parse.unquote(type)
    
    templates = {
        "компенсация": "Молоко_комп_шаблон.xlsx",
        "план": "Молоко_план_шаблон.xlsx",
        "факт": "Молоко_факт_шаблон.xlsx"
    }
    
    template_name = templates.get(type)
    if not template_name:
        return {"error": "Invalid template type"}
        
    template_path = ROOT / "resources" / template_name
    if not template_path.exists():
        return {"error": f"Template {template_name} not found"}
        
    sys_dates = load_system_dates(year)
    transfer_dates = sys_dates["transfer"]
    holiday_dates = sys_dates["holiday"]
    
    m_str = MONTH_NAMES[month] if 1 <= month <= 12 else str(month)
    days_cnt = calendar.monthrange(year, month)[1]

    def employee_exclude_start(raw_date: str) -> int | None:
        raw = text(raw_date).strip()
        if not raw:
            return None
        parts = raw.split("-")
        if len(parts) != 3:
            return None
        try:
            ex_year = int(parts[0])
            ex_month = int(parts[1])
            ex_day = int(parts[2])
        except Exception:
            return None
        if year > ex_year:
            return 1
        if year == ex_year and month > ex_month:
            return 1
        if year == ex_year and month == ex_month:
            return ex_day
        return None

    def employee_hire_start(raw_date: str) -> int | None:
        raw = text(raw_date).strip()
        if not raw:
            return None
        parts = raw.split("-")
        if len(parts) != 3:
            return None
        try:
            h_year = int(parts[0])
            h_month = int(parts[1])
            h_day = int(parts[2])
        except Exception:
            return None
        if year < h_year:
            return 32
        if year == h_year and month < h_month:
            return 32
        if year == h_year and month == h_month:
            return h_day
        return None

    def is_workday(y, m, d):
        dt_obj = dt.date(y, m, d)
        is_we = dt_obj.weekday() >= 5
        is_hol = (m, d) in holiday_dates
        is_tr = (m, d) in transfer_dates
        if is_tr:
            return True
        if is_hol:
            return False
        return not is_we

    non_shift_codes = {"В", "B", "О", "ОВ", "А", "У", "Б", "БН"}
    numeric_shift_re = __import__("re").compile(r"^\d+(?:[.,]\d+)?$")
    spacing_re = __import__("re").compile(r"\s*,\s*")

    def is_shift_mark(value: str) -> bool:
        val = text(value).strip().upper()
        if not val:
            return False
        compact = spacing_re.sub(",", val).replace(" ", "")
        if "М" in compact:
            return True
        if compact in non_shift_codes:
            return False
        return bool(numeric_shift_re.match(compact))

    def build_report_rows() -> tuple[list[dict], int]:
        final_rows: list[dict] = []
        grand_total = 0

        with DB_LOCK, connect() as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()

            emp_rows = cur.execute(
                "SELECT pos, name, tab_num, milk, milk_issue, full_name, milk_note, hire_date, exclude_date FROM employees WHERE y=? AND name != '' ORDER BY rowid",
                (year,),
            ).fetchall()

            m_comp_set = set()
            m_issue_set = set()

            for r in emp_rows:
                name = str(r["name"]).upper()
                if r["milk"]:
                    m_comp_set.add(name)
                if r["milk_issue"]:
                    m_issue_set.add(name)

            ts_rows = cur.execute("SELECT tab_num, c, v FROM timesheet WHERE y=? AND m=?", (year, m_str)).fetchall()
            ts_data = {}
            for r in ts_rows:
                ts_data.setdefault(str(r["tab_num"]), {})[int(r["c"])] = str(r["v"]).upper()

            def allowed_employee(emp_row):
                name_up = str(emp_row["name"]).upper()
                if type == "компенсация":
                    return name_up in m_comp_set
                return name_up in m_issue_set

            for emp in emp_rows:
                if not allowed_employee(emp):
                    continue

                name = str(emp["name"])
                tab_num = str(emp["tab_num"])
                pos = str(emp["pos"])
                full_name = str(emp["full_name"])
                milk_note = str(emp["milk_note"])
                exclude_start = employee_exclude_start(emp["exclude_date"] if "exclude_date" in emp.keys() else "")
                hire_start = employee_hire_start(emp["hire_date"] if "hire_date" in emp.keys() else "")

                if exclude_start == 1 or hire_start == 32:
                    continue

                count = 0
                missed_days: list[str] = []
                emp_ts = ts_data.get(tab_num, {})

                for d in range(1, days_cnt + 1):
                    if exclude_start is not None and d >= exclude_start:
                        continue
                    if hire_start is not None and d < hire_start:
                        continue

                    val = emp_ts.get(d, "").strip().upper()
                    if type == "компенсация":
                        if is_shift_mark(val):
                            count += 1
                        elif val:
                            missed_days.append(f"{d:02d}.{month:02d} {val}")
                    else:
                        if not is_workday(year, month, d):
                            continue
                        if not val:
                            count += 1
                        elif val not in non_shift_codes and is_shift_mark(val):
                            count += 1
                        else:
                            missed_days.append(f"{d:02d}.{month:02d} {val or 'пусто'}")

                final_rows.append({
                    "fio": name,
                    "full_name": full_name,
                    "pos": pos,
                    "tab": tab_num,
                    "shifts": count,
                    "missed_note": "; ".join(missed_days),
                    "milk_note": milk_note,
                })
                grand_total += count

        return final_rows, grand_total

    with DB_LOCK, connect() as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        norm_row = cur.execute("SELECT v FROM ts_norms_data WHERE y=? AND r=? AND c=2", (year, month - 1)).fetchone()
        work_days_norm = str(norm_row["v"]) if norm_row else "0"

    final_list, grand_total = build_report_rows()
        
    # Load workbook
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    row_tpl = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell.value = cell.value.replace("[МЕСЯЦ]", m_str).replace("[ГОД]", str(year)).replace("[НОРМА_ДНЕЙ]", work_days_norm).replace("[ИТОГО]", str(grand_total))
                if row_tpl is None and any(tag in cell.value for tag in [
                    "[№]",
                    "[ФИО]",
                    "[ФИО_ПОЛНОЕ]",
                    "[ДОЛЖНОСТЬ]",
                    "[ТАБ]",
                    "[СМЕНЫ]",
                    "[МОЛОКО_ПРИМ]"
                ]):
                    row_tpl = cell.row

    if row_tpl is not None and final_list:
        tpl_vals = {c: ws.cell(row=row_tpl, column=c).value for c in range(1, ws.max_column + 1)}
        ws.insert_rows(row_tpl + 1, len(final_list) - 1)
        for i, data in enumerate(final_list):
            curr_r = row_tpl + i
            ws.row_dimensions[curr_r].height = None 
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=curr_r, column=c_idx)
                if i > 0: 
                    src = ws.cell(row=row_tpl, column=c_idx)
                    cell.value = src.value
                    if src.has_style:
                        cell.font = copy.copy(src.font)
                        cell.border = copy.copy(src.border)
                        cell.fill = copy.copy(src.fill)
                        cell.alignment = copy.copy(src.alignment)
                v = tpl_vals.get(c_idx)
                if v and isinstance(v, str):
                    v = v.replace("[№]", str(i+1)).replace("[ФИО]", data["fio"]).replace("[ФИО_ПОЛНОЕ]", data["full_name"]).replace("[ДОЛЖНОСТЬ]", data["pos"]).replace("[ТАБ]", data["tab"]).replace("[СМЕНЫ]", str(data["shifts"])).replace("[МОЛОКО_ПРИМ]", data["milk_note"])
                    cell.value = int(v) if str(v).isdigit() else v

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    wb.save(tmp.name)
    
    # Формируем имя файла с добавлением названия месяца и года
    filename_formatted = f"Отчет_Молоко_{type}_{m_str}_{year}.xlsx"
    return FileResponse(tmp.name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=filename_formatted)


@app.get("/api/export-milk-details", response_class=HTMLResponse)
async def export_milk_details(year: int, month: int, type: str):
    import urllib.parse
    type = urllib.parse.unquote(type)

    templates = {
        "компенсация": "Компенсация (План)",
        "план": "Выдача (План)",
        "факт": "Выдача (Факт)",
    }
    title = templates.get(type, type)

    # Reuse the same calculator as the main milk report.
    # We inline the computation in a lightweight form so the diagnostics stay self-contained.
    sys_dates = load_system_dates(year)
    transfer_dates = sys_dates["transfer"]
    holiday_dates = sys_dates["holiday"]
    m_str = MONTH_NAMES[month] if 1 <= month <= 12 else str(month)
    days_cnt = calendar.monthrange(year, month)[1]

    def employee_exclude_start(raw_date: str) -> int | None:
        raw = text(raw_date).strip()
        if not raw:
            return None
        parts = raw.split("-")
        if len(parts) != 3:
            return None
        try:
            ex_year = int(parts[0])
            ex_month = int(parts[1])
            ex_day = int(parts[2])
        except Exception:
            return None
        if year > ex_year:
            return 1
        if year == ex_year and month > ex_month:
            return 1
        if year == ex_year and month == ex_month:
            return ex_day
        return None

    def employee_hire_start(raw_date: str) -> int | None:
        raw = text(raw_date).strip()
        if not raw:
            return None
        parts = raw.split("-")
        if len(parts) != 3:
            return None
        try:
            h_year = int(parts[0])
            h_month = int(parts[1])
            h_day = int(parts[2])
        except Exception:
            return None
        if year < h_year:
            return 32
        if year == h_year and month < h_month:
            return 32
        if year == h_year and month == h_month:
            return h_day
        return None

    def is_workday(y, m, d):
        dt_obj = dt.date(y, m, d)
        is_we = dt_obj.weekday() >= 5
        is_hol = (m, d) in holiday_dates
        is_tr = (m, d) in transfer_dates
        if is_tr:
            return True
        if is_hol:
            return False
        return not is_we

    non_shift_codes = {"В", "B", "О", "ОВ", "А", "У", "Б", "БН"}
    numeric_shift_re = __import__("re").compile(r"^\d+(?:[.,]\d+)?$")
    spacing_re = __import__("re").compile(r"\s*,\s*")

    def is_shift_mark(value: str) -> bool:
        val = text(value).strip().upper()
        if not val:
            return False
        compact = spacing_re.sub(",", val).replace(" ", "")
        if "М" in compact:
            return True
        if compact in non_shift_codes:
            return False
        return bool(numeric_shift_re.match(compact))

    with DB_LOCK, connect() as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        emp_rows = cur.execute(
            "SELECT pos, name, tab_num, milk, milk_issue, full_name, milk_note, hire_date, exclude_date FROM employees WHERE y=? AND name != '' ORDER BY rowid",
            (year,),
        ).fetchall()

        m_comp_set = {str(r["name"]).upper() for r in emp_rows if r["milk"]}
        m_issue_set = {str(r["name"]).upper() for r in emp_rows if r["milk_issue"]}
        ts_rows = cur.execute("SELECT tab_num, c, v FROM timesheet WHERE y=? AND m=?", (year, m_str)).fetchall()
        ts_data = {}
        for r in ts_rows:
            ts_data.setdefault(str(r["tab_num"]), {})[int(r["c"])] = str(r["v"]).upper()

    rows = []
    for emp in emp_rows:
        name = str(emp["name"])
        name_up = name.upper()
        if type == "компенсация" and name_up not in m_comp_set:
            continue
        if type in ("план", "факт") and name_up not in m_issue_set:
            continue

        tab_num = str(emp["tab_num"])
        pos = str(emp["pos"])
        full_name = str(emp["full_name"])
        exclude_start = employee_exclude_start(emp["exclude_date"] if "exclude_date" in emp.keys() else "")
        hire_start = employee_hire_start(emp["hire_date"] if "hire_date" in emp.keys() else "")
        
        if exclude_start == 1 or hire_start == 32:
            continue
            
        emp_ts = ts_data.get(tab_num, {})
        count = 0
        missed_days: list[str] = []

        for d in range(1, days_cnt + 1):
            if exclude_start is not None and d >= exclude_start:
                continue
            if hire_start is not None and d < hire_start:
                continue
            val = emp_ts.get(d, "").strip().upper()
            if type == "компенсация":
                if is_shift_mark(val):
                    count += 1
                elif val:
                    missed_days.append(f"{d:02d}.{month:02d} — {val}")
            else:
                if not is_workday(year, month, d):
                    continue
                if not val:
                    count += 1
                elif val not in non_shift_codes and is_shift_mark(val):
                    count += 1
                else:
                    missed_days.append(f"{d:02d}.{month:02d} — {val or 'пусто'}")

        rows.append({
            "tab": tab_num,
            "fio": full_name or name,
            "pos": pos,
            "shifts": count,
            "missed": missed_days,
        })

    html = [f"<html><head><meta charset='utf-8'><title>{title} — {MONTH_NAMES[month]} {year}</title>"]
    html.append("""
    <style>
      body { font-family: Arial, sans-serif; margin: 20px; color: #1f2937; }
      h1 { margin: 0 0 8px; font-size: 22px; }
      .sub { color: #6b7280; margin-bottom: 16px; }
      table { border-collapse: collapse; width: 100%; }
      th, td { border: 1px solid #cbd5e1; padding: 8px 10px; vertical-align: top; }
      th { background: #eef4ff; text-align: left; }
      tr:nth-child(even) td { background: #fafcff; }
      .num { text-align: center; white-space: nowrap; }
      .missed { white-space: pre-wrap; color: #b45309; }
      .empty { color: #94a3b8; }
    </style>
    </head><body>""")
    html.append(f"<h1>{title} — {MONTH_NAMES[month]} {year}</h1>")
    html.append("<div class='sub'>Показываем только те дни, которые не были засчитаны как смена. Если строка пустая, для этого сотрудника не найдено неучтённых дней.</div>")
    html.append("<table><thead><tr><th style='width:50px'>№</th><th>Таб. №</th><th>ФИО</th><th>Профессия</th><th style='width:90px'>Смен</th><th>Не засчитано</th></tr></thead><tbody>")
    for idx, row in enumerate(rows, 1):
        missed = "<div class='empty'>нет</div>" if not row["missed"] else "<br>".join(row["missed"])
        html.append(
            f"<tr><td class='num'>{idx}</td><td class='num'>{row['tab']}</td><td>{row['fio']}</td><td>{row['pos']}</td><td class='num'>{row['shifts']}</td><td class='missed'>{missed}</td></tr>"
        )
    html.append("</tbody></table></body></html>")
    return HTMLResponse("".join(html))




@app.get("/api/export-sick-email")
async def export_sick_email(emp: str, type: str, start: str, end: str, email: str):
    import urllib.parse
    emp = urllib.parse.unquote(emp)
    type = urllib.parse.unquote(type)
    email = urllib.parse.unquote(email)

    with DB_LOCK, connect() as conn:
        cur = conn.cursor()

        emp_data = cur.execute("SELECT tab_num FROM employees WHERE name=?", (emp,)).fetchone()
        tab_num = str(emp_data["tab_num"]) if emp_data else ""
        
    html_body = f"""
    <html>
    <head>
    <style>
        table {{ border-collapse: collapse; width: 550px; font-family: Arial, sans-serif; border: 1px solid #000000; }}
        td {{ border: 1px solid #000000; padding: 6px; text-align: left; font-size: 13px; }}
    </style>
    </head>
    <body>
        <table>
            <tr>
                <td style="width: 40%;">Табельный номер</td>
                <td style="width: 60%;">{tab_num}</td>
            </tr>
            <tr>
                <td>ФИО сотрудника</td>
                <td>{emp}</td>
            </tr>
            <tr>
                <td>Тип операции</td>
                <td>{type}</td>
            </tr>
            <tr>
                <td style="padding-left: 20px;">Дата начала</td>
                <td>{start}</td>
            </tr>
            <tr>
                <td style="padding-left: 20px;">Дата окончания</td>
                <td>{end}</td>
            </tr>
            <tr>
                <td>Структурное подразделение</td>
                <td>3040 Рудник Таймырский</td>
            </tr>
        </table>
        <br>
    </body>
    </html>
    """

    encoded_subject = base64.b64encode("3040".encode('utf-8')).decode('utf-8')
    mime_subject = f"=?utf-8?B?{encoded_subject}?="
    
    eml_headers = [
        "MIME-Version: 1.0",
        f"To: {email}",
        f"Subject: {mime_subject}",
        "X-Unsent: 1",
        "Content-Type: text/html; charset=utf-8",
        "Content-Transfer-Encoding: 8bit",
        "",
        html_body
    ]
    
    eml_data = "\r\n".join(eml_headers)
    
    return StreamingResponse(
        io.BytesIO(eml_data.encode('utf-8')),
        media_type="message/rfc822",
        headers={"Content-Disposition": f"attachment; filename=SickLeaveDraft.eml"}
    )


@app.get("/api/state")
async def api_get_state(request: Request, year: int, month: int):
    session = get_current_session_fastapi(request)
    if not get_mod_role_fastapi(session, "tabel"):
        return json_response({"error": "Unauthorized"}, 401)
    return json_response(load_state(year, month))

@app.post("/api/state")
async def api_save_state(request: Request):
    try:
        session = get_current_session_fastapi(request)
        role = get_mod_role_fastapi(session, "tabel")
        if role not in ("edit", "editor", "admin"):
            return json_response({"error": "Forbidden"}, 403)
            
        try:
            payload = await request.json()
        except Exception as e:
            return json_response({"error": "Invalid JSON"}, 400)
            
        year = int(payload.get("year", dt.date.today().year))
        month = int(payload.get("month", dt.date.today().month))
        timesheet = payload.get("timesheet")
        employees = payload.get("employees")
        vacations = payload.get("vacations")
        ts_norms_data = payload.get("ts_norms_data")
        month_hint = payload.get("month_hint")
        
        with DB_LOCK, connect() as conn:
            cur = conn.cursor()

            cur.execute("BEGIN")
            
            if timesheet is not None:
                m_str = MONTH_NAMES[month] if 1 <= month <= 12 else str(month)
                cur.execute("DELETE FROM timesheet WHERE y=? AND m=?", (year, m_str))
                insert_ts = []
                if isinstance(timesheet, dict):
                    for tab_num, row_data in timesheet.items():
                        if not row_data: continue
                        for c, v in row_data.items():
                            if v: insert_ts.append((year, m_str, str(tab_num), int(c), str(v)))
                cur.executemany("INSERT INTO timesheet(y, m, tab_num, c, v) VALUES(?,?,?,?,?)", insert_ts)

            if employees is not None:
                cur.execute("DELETE FROM employees WHERE y=?", (year,))
                insert_emp = []
                for r, emp in enumerate(employees):
                    insert_emp.append((year, emp.get("pos",""), emp.get("name",""), emp.get("tab_num",""), 
                                       emp.get("milk",0), emp.get("milk_issue",0), emp.get("full_name",""), emp.get("milk_note",""), emp.get("hire_date",""), emp.get("exclude_date","")))
                cur.executemany("INSERT INTO employees(y, pos, name, tab_num, milk, milk_issue, full_name, milk_note, hire_date, exclude_date) VALUES(?,?,?,?,?,?,?,?,?,?)", insert_emp)

            if vacations is not None:
                cur.execute("DELETE FROM vacations WHERE y=?", (year,))
                insert_vac = []
                if isinstance(vacations, dict):
                    for tab_num, row_data in vacations.items():
                        if not row_data: continue
                        for c, v in row_data.items():
                            if v: insert_vac.append((year, str(tab_num), int(c), str(v)))
                cur.executemany("INSERT INTO vacations(y, tab_num, c, v) VALUES(?,?,?,?)", insert_vac)

            if ts_norms_data is not None:
                cur.execute("DELETE FROM ts_norms_data WHERE y=?", (year,))
                insert_norms = []
                if isinstance(ts_norms_data, dict) or isinstance(ts_norms_data, list):
                    for r_idx, row_data in (ts_norms_data.items() if isinstance(ts_norms_data, dict) else enumerate(ts_norms_data)):
                        if not row_data: continue
                        for c, v in row_data.items():
                            if v: insert_norms.append((year, int(r_idx), int(c), str(v)))
                cur.executemany("INSERT INTO ts_norms_data(y, r, c, v) VALUES(?,?,?,?)", insert_norms)


            if month_hint is not None:
                cur.execute("INSERT OR REPLACE INTO ts_settings (k, v) VALUES ('month_hint', ?)", (str(month_hint),))

            conn.commit()
            
        return json_response({"status": "ok"})
    except Exception as e:
        import traceback
        return json_response({"error": "Global error: " + str(e) + " - " + traceback.format_exc()}, 500)

if __name__ == "__main__":
    host = os.environ.get("WEB_HOST", "0.0.0.0")
    port = int(os.environ.get("WEB_PORT", "8005"))
    uvicorn.run("app:app", host=host, port=port, reload=False)
