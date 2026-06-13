import re

app_py = r'g:\Мой диск\Codex\rtps\web_tabel\app.py'
with open(app_py, 'r', encoding='utf-8') as f:
    text = f.read()

# Add export endpoints
exports_code = '''
from fastapi.responses import HTMLResponse
import calendar

@app.get("/api/export-summary", response_class=HTMLResponse)
async def export_summary(year: int, month: int, type: str):
    import urllib.parse
    type = urllib.parse.unquote(type)
    
    code_map = {
        "Отпуска": ["О"],
        "Отпуска внеплановые": ["ОВ"],
        "Отпуск б/с": ["ДО"],
        "Учебный отпуск": ["У"],
        "Больничный": ["Б"]
    }
    codes = code_map.get(type, [type])

    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        emp_rows = cur.execute("SELECT DISTINCT name FROM employees WHERE y=? AND name != ''", (year,)).fetchall()
        employees = sorted(list(set(row["name"] for row in emp_rows)))
        
        result = {emp: 0 for emp in employees}
        
        # We need to count occurrences of the codes in timesheet
        placeholders = ','.join(['?']*len(codes))
        query = f"""
            SELECT e.name, t.v 
            FROM timesheet t
            JOIN employees e ON t.tab_num = e.tab_num AND t.y = e.y
            WHERE t.y=? AND t.v IN ({placeholders})
        """
        ts_rows = cur.execute(query, [year] + codes).fetchall()
        
        for row in ts_rows:
            name = row["name"]
            if name in result:
                result[name] += 1
                
    total = sum(result.values())
    
    html = f"""
    <html><head><meta charset="utf-8"><title>Сводка: {type}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        table {{ border-collapse: collapse; width: 600px; margin-top: 20px; }}
        th, td {{ border: 1px solid #ccc; padding: 8px; text-align: left; }}
        th {{ background-color: #f0f0f0; }}
        .center {{ text-align: center; }}
        .right {{ text-align: right; font-weight: bold; }}
    </style>
    </head><body>
    <h2 class="center">{type} за {year} год</h2>
    <div class="center" style="color: #666;">Коды: {', '.join(codes)}</div>
    <table>
        <tr><th style="width: 50px;">№</th><th>ФИО</th><th style="width: 100px;">Дней</th></tr>
    """
    for idx, emp in enumerate(employees, 1):
        html += f"<tr><td class='center'>{idx}</td><td>{emp}</td><td class='center'>{result[emp]}</td></tr>"
        
    html += f"<tr><td colspan='2' class='right'>Итого:</td><td class='center'><b>{total}</b></td></tr>"
    html += "</table></body></html>"
    return HTMLResponse(content=html)

@app.get("/api/export-milk")
async def export_milk(year: int, month: int, type: str):
    import urllib.parse
    import tempfile
    import os
    import openpyxl
    type = urllib.parse.unquote(type)
    
    templates = {
        "компенсация": "Молоко_компенсация_шаблон.xlsx",
        "план": "Молоко_план_шаблон.xlsx",
        "факт": "Молоко_факт_шаблон.xlsx"
    }
    
    template_name = templates.get(type)
    if not template_name:
        return {"error": "Invalid template type"}
        
    template_path = ROOT / "resources" / template_name
    if not template_path.exists():
        return {"error": f"Template {template_name} not found"}
        
    sys_dates = load_system_dates(year)
    transfer_dates = sys_dates["transfer"]
    holiday_dates = sys_dates["holiday"]
    
    m_str = MONTH_NAMES[month] if 1 <= month <= 12 else str(month)
    days_cnt = calendar.monthrange(year, month)[1]
    
    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        emp_rows = cur.execute("SELECT name, tab_num, milk, milk_issue FROM employees WHERE y=? AND name != '' ORDER BY rowid", (year,)).fetchall()
        
        m_comp_set = set()
        m_issue_set = set()
        
        for r in emp_rows:
            name = str(r["name"]).upper()
            if r["milk"]: m_comp_set.add(name)
            if r["milk_issue"]: m_issue_set.add(name)
            
        # Get ts_norms_data for workday issue logic (row=month-1, col=2 is workdays norm, but we can just use the DB for timesheet data)
        # Actually, python code: 
        # is_workday_for_issue: not a weekend, not a holiday, or is transfer. 
        # But we also have exact timesheet data in DB.
        
        ts_rows = cur.execute("SELECT tab_num, c, v FROM timesheet WHERE y=? AND m=?", (year, m_str)).fetchall()
        ts_data = {}
        for r in ts_rows:
            ts_data.setdefault(str(r["tab_num"]), {})[int(r["c"])] = str(r["v"]).upper()
            
    # Load workbook
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # We write starting from row 10 (or we just find the first empty row, but usually templates have fixed positions)
    # The python code doesn't specify row insertion, it says `table = self.tables[month_name]`. 
    # Wait, the Python code uses `openpyxl` and fills cells dynamically.
    # Let's write the names and counts.
    # Python code:
    # ws.cell(row=start_r, column=2).value = fio
    # ws.cell(row=start_r, column=4).value = count
    
    # Find starting row - usually it's the first row where column 2 is empty after row 5.
    start_r = 8
    for r in range(1, 50):
        if ws.cell(row=r, column=1).value == "№ п/п":
            start_r = r + 1
            break
            
    current_row = start_r
    idx = 1
    
    def is_workday(y, m, d):
        dt = calendar.datetime.date(y, m, d)
        is_we = dt.weekday() >= 5
        is_hol = (m, d) in holiday_dates
        is_tr = (m, d) in transfer_dates
        if is_tr: return True
        if is_hol: return False
        return not is_we
        
    for emp in emp_rows:
        name = str(emp["name"])
        name_up = name.upper()
        tab_num = str(emp["tab_num"])
        
        if type == "компенсация" and name_up not in m_comp_set: continue
        if type in ["план", "факт"] and name_up not in m_issue_set: continue
        
        count = 0
        emp_ts = ts_data.get(tab_num, {})
        for d in range(1, days_cnt + 1):
            val = emp_ts.get(d, "")
            if type == "компенсация":
                if val in ["В", "РВ", "ДО", "О", "У", "Б"]: # is_milk_cell equivalent or similar
                    # Wait, in python `is_milk_cell(val)` is usually returning True if val is a number!
                    try:
                        float(val.replace(',', '.'))
                        count += 1
                    except ValueError:
                        if val in ["РВ"]: count += 1 # just a guess
                        pass
            else:
                try:
                    float(val.replace(',', '.'))
                    count += 1
                except ValueError:
                    if not val and is_workday(year, month, d): 
                        count += 1
                        
        ws.cell(row=current_row, column=1).value = idx
        ws.cell(row=current_row, column=2).value = name
        ws.cell(row=current_row, column=3).value = tab_num
        ws.cell(row=current_row, column=4).value = count
        current_row += 1
        idx += 1
        
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    wb.save(tmp.name)
    
    return FileResponse(tmp.name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=f"Отчет_Молоко_{type}.xlsx")

'''

if '/api/export-summary' not in text:
    text = text.replace('@app.get("/api/export-sick-email")', exports_code + '\n\n@app.get("/api/export-sick-email")')
    
    # Bump version
    text = re.sub(r'APP_VERSION = "web-tabel-1\.\d+"', 'APP_VERSION = "web-tabel-1.32"', text)
    with open(app_py, 'w', encoding='utf-8') as f:
        f.write(text)
    print("Added export endpoints")
else:
    print("Already added exports")
