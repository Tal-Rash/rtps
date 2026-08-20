from __future__ import annotations

import datetime as dt
import base64
import html
import hashlib
import hmac
import io
import json
import zlib
import os
import secrets
import sqlite3
import threading
import webbrowser
import sys
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from threading import Lock
from urllib.parse import parse_qs, urlparse


ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT.parent))
from rtps_common import connect_sqlite, module_role, resolve_user_access

SHARED_DATA_DIR = ROOT.parent / "data"
LEGACY_WEB_SECRET_FILE = ROOT / "data" / "web_secret.txt"
WEB_SECRET_FILE = SHARED_DATA_DIR / "web_secret.txt"
ACCESS_STATE_FILE = SHARED_DATA_DIR / "web_access.json"
DB_FILE = ROOT.parent / "base" / "common_database.db"
SESSION_COOKIE = "rtps_session"
SESSION_TTL_SECONDS = 7 * 24 * 60 * 60
APP_PREFIX = "/zamer-kp"
APP_VERSION = "web-zkp-2.14"
DB_LOCK = Lock()
MAIN_LOGIN_URL = os.environ.get("MAIN_LOGIN_URL", "http://yrtps.ru/login")

INPUT_ROWS = 12
INPUT_DATA_COLS = 10
DEFAULT_REPAIR_OPTIONS = {
    "tem": ["", "ТО2", "ТО3", "ТО4", "ТР1", "ТР2", "ТР3", "СР", "КР"],
    "pe": ["", "ТО", "ТР", "СР", "КР"],
}
DEFAULT_NORMS = [
    ("max_prokat", "Прокат", "больше или равно", "6", "7"),
    ("min_greben", "Толщина гребня", "меньше или равно", "26", "25"),
    ("min_krut", "Крутизна гребня", "меньше или равно", "7", "6"),
    ("min_bandage_thickness", "Толщина бандажа", "меньше или равно", "", ""),
    ("max_diameter_diff", "Наибольшая разница диаметров бандажей в комплекте, мм", "больше или равно", "", ""),
    ("prokat_6_count", "Число КП с прокатом 6 мм и более", "больше или равно", "", ""),
]
ARCHIVE_EXCEL_HEADERS = [
    "Дата замера",
    "Локомотив",
    "Вид ремонта",
    "Серия",
    "Секция",
    "Номер КП",
    "Прокат лев",
    "Прокат прав",
    "Толщина гребня лев",
    "Толщина гребня прав",
    "Крутизна гребня лев",
    "Крутизна гребня прав",
    "Толщина бандажа лев",
    "Толщина бандажа прав",
    "Диаметр бандажа лев",
    "Диаметр бандажа прав",
]
WEAR_TREND_METRICS = [
    {
        "key": "prokat",
        "label": "Прокат",
        "columns": (2, 3),
        "aggregate": "max",
        "worse_when": "higher",
    },
    {
        "key": "greben",
        "label": "Гребень",
        "columns": (4, 5),
        "aggregate": "min",
        "worse_when": "lower",
    },
    {
        "key": "krut",
        "label": "Крутизна",
        "columns": (6, 7),
        "aggregate": "min",
        "worse_when": "lower",
    },
    {
        "key": "bandage_thickness",
        "label": "Бандаж",
        "columns": (8, 9),
        "aggregate": "min",
        "worse_when": "lower",
    },
    {
        "key": "diameter",
        "label": "Диаметр",
        "columns": (10, 11),
        "aggregate": "side",
        "worse_when": "lower",
    },
]


ROOT = Path(__file__).parent
SHARED_DATA_DIR = ROOT.parent / "data"

def load_web_secret() -> str:
    if WEB_SECRET_FILE.exists():
        return WEB_SECRET_FILE.read_text(encoding="utf-8").strip()
    return "opYbo6NB8pb7dChYQkmHEvUH6K4hAHjuzi2qEYOC024"

WEB_SECRET = load_web_secret()
LEGACY_WEB_SECRET = ""


def connect() -> sqlite3.Connection:
    DB_FILE.parent.mkdir(parents=True, exist_ok=True)
    return connect_sqlite(DB_FILE)


def ensure_db() -> None:
    with DB_LOCK, connect() as conn:
        cur = conn.cursor()

        cur.execute(
            "CREATE TABLE IF NOT EXISTS input_meta (y INT, locomotive TEXT, measurement_date TEXT, wheel_pair_count INT, section_count INT, PRIMARY KEY(y, locomotive))"
        )
        cur.execute(
            "CREATE TABLE IF NOT EXISTS input_data (y INT, locomotive TEXT, r INT, c INT, v TEXT, PRIMARY KEY(y, locomotive, r, c))"
        )
        cur.execute("CREATE TABLE IF NOT EXISTS inventory (y INT, ser TEXT, num TEXT, inv TEXT, PRIMARY KEY(y, ser, num))")
        _ensure_inventory_sync_columns(conn)
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS kp_data (
                locomotive TEXT, r INT, c INT, v TEXT,
                PRIMARY KEY(locomotive, r, c)
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS kp_data_versions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                locomotive TEXT NOT NULL,
                valid_to TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS kp_data_version_cells (
                version_id INT NOT NULL,
                r INT NOT NULL,
                c INT NOT NULL,
                v TEXT,
                PRIMARY KEY(version_id, r, c)
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS archive_data (
                y INT,
                measurement_date TEXT,
                locomotive TEXT,
                repair_type TEXT,
                r INT,
                c INT,
                v TEXT,
                PRIMARY KEY(y, measurement_date, locomotive, repair_type, r, c)
            )
            """
        )
        _migrate_archive_repair_types(conn)
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS kp_norms_data (
                metric_key TEXT PRIMARY KEY,
                label TEXT,
                condition TEXT,
                yellow_value TEXT,
                red_value TEXT
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS app_meta (
                k TEXT PRIMARY KEY,
                v TEXT
            )
            """
        )
        existing_input_meta_cols = {row[1] for row in cur.execute("PRAGMA table_info(input_meta)").fetchall()}
        if "wheel_pair_count" not in existing_input_meta_cols:
            cur.execute("ALTER TABLE input_meta ADD COLUMN wheel_pair_count INT")
        if "section_count" not in existing_input_meta_cols:
            cur.execute("ALTER TABLE input_meta ADD COLUMN section_count INT")
        existing_inventory_cols = {row[1] for row in cur.execute("PRAGMA table_info(inventory)").fetchall()}
        if "sort_order" not in existing_inventory_cols:
            cur.execute("ALTER TABLE inventory ADD COLUMN sort_order INT")
        if "updated_at" not in existing_inventory_cols:
            cur.execute("ALTER TABLE inventory ADD COLUMN updated_at INT")
            cur.execute("UPDATE inventory SET updated_at = 0 WHERE updated_at IS NULL")
        if "deleted_at" not in existing_inventory_cols:
            cur.execute("ALTER TABLE inventory ADD COLUMN deleted_at INT")
            cur.execute("UPDATE inventory SET deleted_at = 0 WHERE deleted_at IS NULL")
        if "wheel_pair_count" not in existing_inventory_cols:
            cur.execute("ALTER TABLE inventory ADD COLUMN wheel_pair_count INT")
        if "section_count" not in existing_inventory_cols:
            cur.execute("ALTER TABLE inventory ADD COLUMN section_count INT")
        if "eight_digit_number" not in existing_inventory_cols:
            cur.execute("ALTER TABLE inventory ADD COLUMN eight_digit_number TEXT")
        if "manufacture_year" not in existing_inventory_cols:
            cur.execute("ALTER TABLE inventory ADD COLUMN manufacture_year TEXT")
        if "service_life" not in existing_inventory_cols:
            cur.execute("ALTER TABLE inventory ADD COLUMN service_life TEXT")

        cur.execute("UPDATE inventory SET sort_order = rowid WHERE sort_order IS NULL OR sort_order <= 0")
        cur.executemany(
            "INSERT OR IGNORE INTO kp_norms_data(metric_key, label, condition, yellow_value, red_value) VALUES(?,?,?,?,?)",
            DEFAULT_NORMS,
        )
        conn.commit()


def _migrate_archive_repair_types(conn: sqlite3.Connection) -> None:
    cur = conn.cursor()
    rows = cur.execute(
        "SELECT y, measurement_date, locomotive, repair_type, r, c FROM archive_data WHERE repair_type LIKE '%-%' OR repair_type LIKE '% %'"
    ).fetchall()
    for row in rows:
        year = int(row["y"] or 0)
        measurement_date = text(row["measurement_date"]).strip()
        locomotive = text(row["locomotive"]).strip()
        repair_type = text(row["repair_type"]).strip()
        normalized = normalize_repair_type(repair_type)
        if not normalized or normalized == repair_type:
            continue
        r = int(row["r"] or 0)
        c = int(row["c"] or 0)
        exists = cur.execute(
            "SELECT 1 FROM archive_data WHERE y=? AND measurement_date=? AND locomotive=? AND repair_type=? AND r=? AND c=?",
            (year, measurement_date, locomotive, normalized, r, c),
        ).fetchone()
        if exists:
            cur.execute(
                "DELETE FROM archive_data WHERE y=? AND measurement_date=? AND locomotive=? AND repair_type=? AND r=? AND c=?",
                (year, measurement_date, locomotive, repair_type, r, c),
            )
        else:
            cur.execute(
                "UPDATE archive_data SET repair_type=? WHERE y=? AND measurement_date=? AND locomotive=? AND repair_type=? AND r=? AND c=?",
                (normalized, year, measurement_date, locomotive, repair_type, r, c),
            )
    conn.commit()





def _ensure_inventory_sync_columns(conn: sqlite3.Connection) -> None:
    columns = {text(row[1]).strip().lower() for row in conn.execute("PRAGMA table_info(inventory)").fetchall()}
    if "updated_at" not in columns:
        conn.execute("ALTER TABLE inventory ADD COLUMN updated_at INT NOT NULL DEFAULT 0")
    if "deleted_at" not in columns:
        conn.execute("ALTER TABLE inventory ADD COLUMN deleted_at INT NOT NULL DEFAULT 0")


def text(value) -> str:
    return "" if value is None else str(value)


def normalize_repair_type(value) -> str:
    return text(value).strip().upper().replace(" ", "").replace("-", "")


def parse_cookie_values(handler: BaseHTTPRequestHandler, name: str) -> list[str]:
    values: list[str] = []
    for part in handler.headers.get("Cookie", "").split(";"):
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        if key.strip() == name:
            values.append(value.strip())
    return values


def verify_cookie(value: str) -> tuple[str, str, str, str] | None:
    for sep in (":", "|"):
        try:
            parts = value.rsplit(sep, 5)
            if len(parts) == 6:
                user_id, role, modules, safe_name, expiry_text, sig = parts
                payload = f"{user_id}{sep}{role}{sep}{modules}{sep}{safe_name}{sep}{expiry_text}"
            elif len(parts) == 5:
                user_id, role, modules, safe_name, sig = parts
                payload = f"{user_id}{sep}{role}{sep}{modules}{sep}{safe_name}"
                expiry_text = "2000000000"
            else:
                continue
                
            secrets_to_try = [WEB_SECRET]
            if LEGACY_WEB_SECRET and LEGACY_WEB_SECRET not in secrets_to_try:
                secrets_to_try.append(LEGACY_WEB_SECRET)
                
            matched = False
            import hmac
            import hashlib
            for secret in secrets_to_try:
                expected = hmac.new(secret.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
                if hmac.compare_digest(expected, sig):
                    matched = True
                    break
                    
            if not matched:
                continue
            import datetime as dt
            if float(expiry_text) < dt.datetime.now().timestamp():
                return None
                
            import urllib.parse
            return urllib.parse.unquote(user_id), urllib.parse.unquote(role), urllib.parse.unquote(modules), urllib.parse.unquote(safe_name)
        except Exception:
            continue
    return None

def current_session(handler) -> tuple[str, str, str, str] | None:
    for token in parse_cookie_values(handler, SESSION_COOKIE):
        session = verify_cookie(token)
        if session:
            return session
    return None

def get_mod_role(session: tuple[str, str, str, str] | None, mod_name: str) -> str | None:
    if not session: return None
    username = session[0]
    role = session[1]
    modules = session[2]

    resolved = resolve_user_access(ROOT.parent / "base" / "web_users.db", username, role, modules)
    if not resolved:
        return None
    role, modules = resolved
    return module_role(role, modules, mod_name)

def require_auth(handler, need_edit: bool = False) -> tuple[str, str, str, str] | None:
    session = current_session(handler)
    if not session:
        handler.send_error(HTTPStatus.UNAUTHORIZED, "Unauthorized")
        return None
    mod_role = get_mod_role(session, "zamer_kp")
    if not mod_role:
        handler.send_error(HTTPStatus.FORBIDDEN, "Forbidden")
        return None
    if need_edit and mod_role not in ("edit", "editor", "admin"):
        handler.send_error(HTTPStatus.FORBIDDEN, "Forbidden")
        return None
    return session

def route_path(path: str) -> str:
    if path == APP_PREFIX:
        return "/"
    if path.startswith(APP_PREFIX + "/"):
        return path[len(APP_PREFIX) :] or "/"
    return path


def send_html(handler: BaseHTTPRequestHandler, body: str, status: int = 200) -> None:
    data = body.encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "text/html; charset=utf-8")
    handler.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
    handler.send_header("X-Content-Type-Options", "nosniff")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def send_json(handler: BaseHTTPRequestHandler, payload, status: int = 200) -> None:
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Cache-Control", "no-store")
    handler.send_header("X-Content-Type-Options", "nosniff")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def send_file(handler: BaseHTTPRequestHandler, data: bytes, filename: str, content_type: str) -> None:
    safe_filename = filename.replace('"', "").replace("\r", "").replace("\n", "")
    encoded = "".join(f"%{byte:02X}" for byte in safe_filename.encode("utf-8"))
    
    ascii_filename = "".join(c if ord(c) < 128 else "_" for c in safe_filename)
    
    handler.send_response(HTTPStatus.OK)
    handler.send_header("Content-Type", content_type)
    handler.send_header("Cache-Control", "no-store")
    handler.send_header("X-Content-Type-Options", "nosniff")
    handler.send_header("Content-Disposition", f'attachment; filename="{ascii_filename}"; filename*=UTF-8\'\'{encoded}')
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def redirect(handler: BaseHTTPRequestHandler, location: str, cookie: str | None = None) -> None:
    handler.send_response(HTTPStatus.SEE_OTHER)
    handler.send_header("Location", location)
    handler.send_header("Cache-Control", "no-store")
    if cookie:
        handler.send_header("Set-Cookie", cookie)
    handler.end_headers()



def normalize_text(value: str) -> str:
    text_value = text(value).strip().lower()
    text_value = text_value.replace("ё", "е")
    return text_value


def series_for_locomotive(cur: sqlite3.Cursor, locomotive: str) -> str:
    locomotive = text(locomotive).strip()
    if not locomotive:
        return ""
    row = cur.execute(
        "SELECT ser FROM inventory WHERE TRIM(COALESCE(num, ''))=? ORDER BY COALESCE(sort_order, 0) ASC, COALESCE(updated_at, y) DESC, y DESC, rowid DESC LIMIT 1",
        (locomotive,),
    ).fetchone()
    if row:
        return text(row["ser"]).strip()
    return ""


def locomotive_axis_count(series: str, locomotive: str) -> int:
    normalized = normalize_text(series + " " + locomotive)
    if "пэ-2м" in normalized or "пэ2м" in normalized or "пэ 2м" in normalized or "pe-2m" in normalized or "pe2m" in normalized:
        return 12
    if "тэм" in normalized or "tem" in normalized:
        return 6
    return 12


def default_section_count(axis_count: int) -> int:
    return 1 if int(axis_count or 0) <= 6 else 3


def allowed_repairs(series: str, locomotive: str) -> list[str]:
    normalized = normalize_text(series + " " + locomotive)
    if "пэ-2м" in normalized or "пэ2м" in normalized or "пэ 2м" in normalized or "pe-2m" in normalized or "pe2m" in normalized:
        return DEFAULT_REPAIR_OPTIONS["pe"]
    return DEFAULT_REPAIR_OPTIONS["tem"]


def load_locomotives(cur: sqlite3.Cursor) -> list[dict[str, str]]:
    return load_inventory_records(cur, include_deleted=False)


def load_inventory_records(cur: sqlite3.Cursor, include_deleted: bool = False) -> list[dict[str, str]]:
    query = """
        SELECT y, ser, num, inv, COALESCE(sort_order, 0) AS sort_order, COALESCE(updated_at, 0) AS updated_at, COALESCE(deleted_at, 0) AS deleted_at, COALESCE(wheel_pair_count, 0) AS wheel_pair_count, COALESCE(section_count, 0) AS section_count, COALESCE(eight_digit_number, '') AS eight_digit_number, COALESCE(manufacture_year, '') AS manufacture_year, COALESCE(service_life, '') AS service_life
        FROM inventory
        WHERE TRIM(COALESCE(num, '')) <> ''
    """
    if not include_deleted:
        query += " AND COALESCE(deleted_at, 0) = 0"
    query += " ORDER BY COALESCE(sort_order, 0) ASC, COALESCE(updated_at, 0) DESC, rowid DESC"
    rows = cur.execute(query).fetchall()

    best_rows: dict[str, sqlite3.Row] = {}
    for row in rows:
        number = text(row["num"]).strip()
        if not number:
            continue
        series = normalize_text(row["ser"]).strip().upper()
        key = f"{series}|{number}"
        current = best_rows.get(key)
        if current is None:
            best_rows[key] = row
            continue
        current_rank = (
            int(current["updated_at"] or 0),
            int(current["deleted_at"] or 0),
            int(current["sort_order"] or 0),
            int(current["rowid"] or 0),
        )
        row_rank = (
            int(row["updated_at"] or 0),
            int(row["deleted_at"] or 0),
            int(row["sort_order"] or 0),
            int(row["rowid"] or 0),
        )
        if row_rank > current_rank:
            best_rows[key] = row

    result: list[dict[str, str]] = []
    for row in sorted(
        best_rows.values(),
        key=lambda item: (
            int(item["sort_order"] or 0),
            -int(item["updated_at"] or 0),
            -int(item["deleted_at"] or 0),
            normalize_text(item["ser"]).strip().upper(),
            text(item["num"]).strip(),
        ),
    ):
        number = text(row["num"]).strip()
        if not number:
            continue
        deleted_at = int(row["deleted_at"] or 0)
        if deleted_at > 0 and not include_deleted:
            continue
        series = normalize_text(row["ser"]).strip().upper()
        inv = text(row["inv"]).strip()
        sort_order = int(row["sort_order"] or 0)
        updated_at = int(row["updated_at"] or 0)
        wheel_pair_count = int(row["wheel_pair_count"] or 0)
        section_count = int(row["section_count"] or 0)
        eight_digit_number = text(row["eight_digit_number"]).strip()
        manufacture_year = text(row["manufacture_year"]).strip()
        service_life = text(row["service_life"]).strip()
        label = f"{series} {number}".strip()
        result.append(
            {
                "series": series,
                "number": number,
                "label": label,
                "inventoryNumber": inv,
                "sortOrder": sort_order,
                "updatedAt": updated_at,
                "deletedAt": deleted_at,
                "wheelPairCount": wheel_pair_count,
                "sectionCount": section_count,
                "eightDigitNumber": eight_digit_number,
                "manufactureYear": manufacture_year,
                "serviceLife": service_life,
            }
        )
    return result


def empty_measurements() -> list[list[str]]:
    return [["" for _ in range(INPUT_DATA_COLS)] for _ in range(INPUT_ROWS)]


def row_to_index(row_value: int) -> int | None:
    idx = int(row_value) - 2
    return idx if 0 <= idx < INPUT_ROWS else None


def load_state(locomotive: str | None = None, wheel_pair_count: int | None = None, section_count: int | None = None) -> dict:
    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        locomotives = load_locomotives(cur)
        if not locomotive:
            locomotive = locomotives[0]["number"] if locomotives else ""
        locomotive = text(locomotive).strip()

        series = series_for_locomotive(cur, locomotive)
        axis_count = locomotive_axis_count(series, locomotive)
        repair_options = allowed_repairs(series, locomotive)
        locomotive_record = next((item for item in locomotives if text(item.get("number")).strip() == locomotive), None)

        meta = None
        if locomotive:
            meta = cur.execute(
                "SELECT y, measurement_date, wheel_pair_count, section_count FROM input_meta WHERE locomotive=? ORDER BY y DESC LIMIT 1",
                (locomotive,),
            ).fetchone()

        measurement_date = dt.date.today().isoformat()
        year = dt.date.today().year
        if wheel_pair_count is None:
            wheel_pair_count = int(locomotive_record.get("wheelPairCount") or 0) if locomotive_record else 0
            if wheel_pair_count <= 0:
                wheel_pair_count = axis_count
        if section_count is None:
            section_count = int(locomotive_record.get("sectionCount") or 0) if locomotive_record else 0
            if section_count <= 0:
                section_count = default_section_count(axis_count)
        has_manual_meta = False
        if meta:
            year = int(meta["y"] or year)
            measurement_date = text(meta["measurement_date"]).strip() or measurement_date
            try:
                year = int(measurement_date[:4])
            except Exception:
                pass
            try:
                meta_wheel_pairs = int(meta["wheel_pair_count"] or 0)
            except Exception:
                meta_wheel_pairs = 0
            try:
                meta_sections = int(meta["section_count"] or 0)
            except Exception:
                meta_sections = 0
            if meta_wheel_pairs > 0:
                wheel_pair_count = meta_wheel_pairs
            if meta_sections > 0:
                section_count = meta_sections
            has_manual_meta = meta_wheel_pairs > 0 or meta_sections > 0
        elif measurement_date:
            try:
                year = int(measurement_date[:4])
            except Exception:
                year = dt.date.today().year

        rows = empty_measurements()
        if locomotive:
            db_rows = cur.execute(
                "SELECT r, c, v FROM input_data WHERE y=? AND locomotive=? ORDER BY r, c",
                (year, locomotive),
            ).fetchall()
            for row in db_rows:
                idx = row_to_index(int(row["r"]))
                col = int(row["c"]) - 2
                if idx is None or not (0 <= col < INPUT_DATA_COLS):
                    continue
                rows[idx][col] = text(row["v"])

        kp_rows = cur.execute("SELECT r, c, v FROM kp_data WHERE locomotive=? ORDER BY r, c", (locomotive,)).fetchall()
        if not kp_rows:
            kp_rows = cur.execute("SELECT r, c, v FROM kp_data WHERE locomotive='' ORDER BY r, c").fetchall()

        kp_map: dict[int, dict[int, str]] = {}
        for row in kp_rows:
            kp_map.setdefault(int(row["r"]), {})[int(row["c"])] = text(row["v"])

        norms = {
            row["metric_key"]: {
                "label": text(row["label"]),
                "condition": text(row["condition"]),
                "yellow_value": text(row["yellow_value"]),
                "red_value": text(row["red_value"]),
            }
            for row in cur.execute(
                "SELECT metric_key, label, condition, yellow_value, red_value FROM kp_norms_data ORDER BY rowid"
            ).fetchall()
        }

    return {
        "locomotive": locomotive,
        "series": series,
        "axis_count": axis_count,
        "measurement_date": measurement_date,
        "repair_type": "",
        "repair_options": repair_options,
        "locomotives": locomotives,
        "measurements": rows,
        "kp": kp_map,
        "norms": norms,
        "year": year,
        "wheel_pair_count": int(wheel_pair_count or axis_count),
        "section_count": int(section_count or default_section_count(axis_count)),
        "has_manual_meta": has_manual_meta,
    }


def kp_completion_state(cur: sqlite3.Cursor, locomotive: str) -> str | None:
    locomotive = text(locomotive).strip()
    if not locomotive:
        return None

    series = series_for_locomotive(cur, locomotive)
    expected_rows = locomotive_axis_count(series, locomotive)
    rows = cur.execute(
        "SELECT r, c, v FROM kp_data WHERE locomotive=? AND c IN (2, 3)",
        (locomotive,),
    ).fetchall()
    if not rows:
        return None

    values: dict[tuple[int, int], str] = {}
    for row in rows:
        values[(int(row["r"]), int(row["c"]))] = text(row["v"]).strip()

    for row_index in range(expected_rows):
        if not values.get((row_index, 2), "") or not values.get((row_index, 3), ""):
            return "yellow"
    return "green"


def load_kp_versions(cur: sqlite3.Cursor, locomotive: str) -> list[dict]:
    rows = cur.execute(
        """
        SELECT id, valid_to, created_at
        FROM kp_data_versions
        WHERE locomotive=?
        ORDER BY COALESCE(valid_to, '') DESC, id DESC
        """,
        (locomotive,),
    ).fetchall()
    return [
        {
            "id": int(row["id"]),
            "valid_to": text(row["valid_to"]).strip(),
            "created_at": text(row["created_at"]).strip(),
        }
        for row in rows
    ]


def load_kp_view(selected_locomotive: str = "", version_id: int | None = None) -> dict:
    selected_locomotive = text(selected_locomotive).strip()
    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        locomotives = load_locomotives(cur)
        if not selected_locomotive and locomotives:
            selected_locomotive = locomotives[0]["number"]

        all_mode = selected_locomotive == "Все локомотивы"
        kp_map: dict[int, dict[int, str]] = {}
        rows: list[dict[str, object]] = []
        axis_count = 0
        status = None
        selected_version = None

        if all_mode:
            all_values = cur.execute(
                "SELECT locomotive, r, c, v FROM kp_data WHERE TRIM(COALESCE(locomotive, '')) <> '' ORDER BY locomotive, r, c"
            ).fetchall()
            kp_values: dict[tuple[str, int, int], str] = {}
            for row in all_values:
                kp_values[(text(row["locomotive"]).strip(), int(row["r"]), int(row["c"]))] = text(row["v"])

            for loco in locomotives:
                number = loco["number"]
                series = loco["series"]
                count = locomotive_axis_count(series, number)
                for row_index in range(count):
                    values = [
                        number,
                        kp_values.get((number, row_index, 0), str(row_index + 1)),
                        kp_values.get((number, row_index, 1), ""),
                        kp_values.get((number, row_index, 2), ""),
                        kp_values.get((number, row_index, 3), ""),
                    ]
                    rows.append(
                        {
                            "locomotive": number,
                            "row": row_index,
                            "values": values,
                            "search": " ".join(text(v).strip().lower() for v in values),
                            "editable": False,
                        }
                    )
            status = "all"
        else:
            series = series_for_locomotive(cur, selected_locomotive)
            axis_count = locomotive_axis_count(series, selected_locomotive)
            if version_id:
                version_row = cur.execute(
                    """
                    SELECT id, valid_to, created_at
                    FROM kp_data_versions
                    WHERE id=? AND locomotive=?
                    """,
                    (version_id, selected_locomotive),
                ).fetchone()
                if version_row:
                    selected_version = {
                        "id": int(version_row["id"]),
                        "valid_to": text(version_row["valid_to"]).strip(),
                        "created_at": text(version_row["created_at"]).strip(),
                    }
                    kp_rows = cur.execute(
                        "SELECT r, c, v FROM kp_data_version_cells WHERE version_id=? ORDER BY r, c",
                        (version_id,),
                    ).fetchall()
                else:
                    kp_rows = []
            else:
                kp_rows = cur.execute(
                    "SELECT r, c, v FROM kp_data WHERE locomotive=? ORDER BY r, c",
                    (selected_locomotive,),
                ).fetchall()
                if not kp_rows:
                    kp_rows = cur.execute(
                        "SELECT r, c, v FROM kp_data WHERE locomotive='' ORDER BY r, c"
                    ).fetchall()
            for row in kp_rows:
                kp_map.setdefault(int(row["r"]), {})[int(row["c"])] = text(row["v"])

            for row_index in range(axis_count):
                values = [
                    kp_map.get(row_index, {}).get(0, str(row_index + 1)),
                    kp_map.get(row_index, {}).get(1, ""),
                    kp_map.get(row_index, {}).get(2, ""),
                    kp_map.get(row_index, {}).get(3, ""),
                ]
                rows.append(
                    {
                        "locomotive": selected_locomotive,
                        "row": row_index,
                        "values": values,
                        "search": " ".join(text(v).strip().lower() for v in values),
                        "editable": selected_version is None,
                    }
                )
            status = kp_completion_state(cur, selected_locomotive) if selected_version is None else None

        return {
            "selected_locomotive": selected_locomotive,
            "all_mode": all_mode,
            "axis_count": axis_count,
            "locomotives": locomotives,
            "rows": rows,
            "kp_map": kp_map,
            "status": status,
            "versions": load_kp_versions(cur, selected_locomotive) if selected_locomotive and not all_mode else [],
            "selected_version": selected_version,
        }


def save_kp_data(payload: dict) -> dict:
    locomotive = text(payload.get("locomotive")).strip()
    if not locomotive or locomotive == "Все локомотивы":
        return {"error": "Выберите конкретный локомотив."}, 400

    rows = payload.get("rows") or []
    if not isinstance(rows, list):
        return {"error": "Некорректные данные КП."}, 400

    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        cur.execute("BEGIN")
        cur.execute("DELETE FROM kp_data WHERE locomotive=?", (locomotive,))

        insert_rows: list[tuple[str, int, int, str]] = []
        for r, row in enumerate(rows):
            values = list(row or []) + [""] * 4
            for c, value in enumerate(values[:4]):
                value = text(value).strip()
                if value:
                    insert_rows.append((locomotive, r, c, value))
        cur.executemany("INSERT INTO kp_data (locomotive, r, c, v) VALUES (?, ?, ?, ?)", insert_rows)
        conn.commit()

    return load_kp_view(locomotive)


def create_kp_data_version(payload: dict) -> dict:
    locomotive = text(payload.get("locomotive")).strip()
    valid_to = text(payload.get("valid_to")).strip()
    if not locomotive or locomotive == "Все локомотивы":
        return {"error": "Выберите конкретный локомотив."}, 400
    try:
        dt.date.fromisoformat(valid_to)
    except Exception:
        return {"error": "Укажите дату, до которой действовали старые данные."}, 400

    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        current_rows = cur.execute(
            "SELECT r, c, v FROM kp_data WHERE locomotive=? ORDER BY r, c",
            (locomotive,),
        ).fetchall()
        if not current_rows:
            return {"error": "Нет текущих данных для сохранения в историю."}, 400

        cur.execute(
            "INSERT INTO kp_data_versions(locomotive, valid_to, created_at) VALUES(?,?,?)",
            (locomotive, valid_to, dt.datetime.now().isoformat(timespec="seconds")),
        )
        version_id = int(cur.lastrowid)
        cur.executemany(
            "INSERT INTO kp_data_version_cells(version_id, r, c, v) VALUES(?,?,?,?)",
            [
                (version_id, int(row["r"]), int(row["c"]), text(row["v"]))
                for row in current_rows
            ],
        )
        cur.execute("DELETE FROM kp_data WHERE locomotive=?", (locomotive,))
        new_rows = [
            (locomotive, int(row["r"]), int(row["c"]), text(row["v"]))
            for row in current_rows
            if int(row["c"]) in (0, 1) and text(row["v"]).strip()
        ]
        cur.executemany(
            "INSERT INTO kp_data(locomotive, r, c, v) VALUES(?,?,?,?)",
            new_rows,
        )
        conn.commit()
    return load_kp_view(locomotive)


def update_kp_version(payload: dict) -> dict:
    locomotive = text(payload.get("locomotive")).strip()
    valid_to = text(payload.get("valid_to")).strip()
    try:
        version_id = int(payload.get("version_id"))
        dt.date.fromisoformat(valid_to)
    except Exception:
        return {"error": "Некорректная дата версии КП."}, 400
    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        cur.execute(
            "UPDATE kp_data_versions SET valid_to=? WHERE id=? AND locomotive=?",
            (valid_to, version_id, locomotive),
        )
        if cur.rowcount <= 0:
            return {"error": "Версия КП не найдена."}, 404
        conn.commit()
    return load_kp_view(locomotive, version_id)


def save_state(payload: dict, full_name: str = "") -> dict:
    locomotive = text(payload.get("locomotive")).strip()
    measurement_date = text(payload.get("measurement_date")).strip() or dt.date.today().isoformat()
    rows = payload.get("measurements") or []
    try:
        wheel_pair_count = int(payload.get("wheel_pair_count") or 0)
    except Exception:
        wheel_pair_count = 0
    try:
        section_count = int(payload.get("section_count") or 0)
    except Exception:
        section_count = 0

    try:
        year = int(measurement_date[:4])
    except Exception:
        year = dt.date.today().year

    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        cur.execute("BEGIN")
        cur.execute("DELETE FROM input_data WHERE y=? AND locomotive=?", (year, locomotive))
        insert_rows: list[tuple[int, str, int, int, str]] = []
        for r, row in enumerate(rows):
            row = list(row or []) + [""] * INPUT_DATA_COLS
            for c, value in enumerate(row[:INPUT_DATA_COLS]):
                value = text(value).strip()
                if value:
                    insert_rows.append((year, locomotive, r + 2, c + 2, value))
        cur.executemany("INSERT INTO input_data(y, locomotive, r, c, v) VALUES(?,?,?,?,?)", insert_rows)
        cur.execute(
            "INSERT OR REPLACE INTO input_meta(y, locomotive, measurement_date, wheel_pair_count, section_count) VALUES(?,?,?,?,?)",
            (
                year,
                locomotive,
                measurement_date,
                wheel_pair_count or None,
                section_count or None,
            ),
        )
        conn.commit()
    return load_state(locomotive)


def build_archive_rows(payload: dict) -> tuple[dict, list[tuple[int, str, str, str, int, int, str]], list[str]]:
    locomotive = text(payload.get("locomotive")).strip()
    measurement_date = text(payload.get("measurement_date")).strip() or dt.date.today().isoformat()
    repair_type = normalize_repair_type(payload.get("repair_type"))
    rows = payload.get("measurements") or []

    try:
        year = int(measurement_date[:4])
    except Exception:
        year = dt.date.today().year

    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        series = series_for_locomotive(cur, locomotive)
        series_text = normalize_text(series + " " + locomotive)
        axis_count = locomotive_axis_count(series, locomotive)
        try:
            wheel_pair_count = int(payload.get("wheel_pair_count") or axis_count)
        except Exception:
            wheel_pair_count = axis_count
        try:
            section_count = int(payload.get("section_count") or default_section_count(axis_count))
        except Exception:
            section_count = default_section_count(axis_count)
        visible_rows = max(1, min(wheel_pair_count, INPUT_ROWS))
        if "пэ-2м" in series_text or "пэ2м" in series_text or "пэ 2м" in series_text or "pe-2m" in series_text or "pe2m" in series_text:
            required_columns = [0, 1, 2, 3, 6, 7]
        else:
            required_columns = [0, 1, 2, 3, 4, 5, 6, 7]

        kp_rows = cur.execute("SELECT r, c, v FROM kp_data WHERE locomotive=? ORDER BY r, c", (locomotive,)).fetchall()
        if not kp_rows:
            kp_rows = cur.execute("SELECT r, c, v FROM kp_data WHERE locomotive='' ORDER BY r, c").fetchall()
        kp_map: dict[int, dict[int, str]] = {}
        for row in kp_rows:
            kp_map.setdefault(int(row["r"]), {})[int(row["c"])] = text(row["v"])

        def parse_float(v):
            try:
                return float(text(v).replace(",", "."))
            except Exception:
                return None

        missing_cells: list[str] = []
        normalized_rows: list[list[str]] = []
        section_sizes: list[int] = []
        base = visible_rows // max(1, section_count)
        remainder = visible_rows % max(1, section_count)
        for i in range(max(1, section_count)):
            section_sizes.append(base + (1 if i < remainder else 0))
        for row_index in range(visible_rows):
            row = list(rows[row_index] if row_index < len(rows) else []) + [""] * INPUT_DATA_COLS
            normalized_rows.append(row[:INPUT_DATA_COLS])
            for col in required_columns:
                value = text(row[col]).strip()
                if not value:
                    axis_number = row_index + 1
                    missing_cells.append(f"ось {axis_number}, колонка {col + 1}")

        if missing_cells:
            return {"error": "missing", "measurement_date": measurement_date, "locomotive": locomotive, "repair_type": repair_type}, [], missing_cells

        if not repair_type:
            return {"error": "repair", "measurement_date": measurement_date, "locomotive": locomotive, "repair_type": repair_type}, [], []

        archive_rows: list[tuple[int, str, str, str, int, int, str]] = []
        for row_index in range(visible_rows):
            row = normalized_rows[row_index]
            table_row = row_index + 2
            running = 0
            section_value = "1"
            for section_index, span in enumerate(section_sizes, start=1):
                running += span
                if row_index < running:
                    section_value = str(section_index)
                    break
            kp_row = kp_map.get(row_index, {})

            left_band = parse_float(row[6])
            right_band = parse_float(row[7])
            left_kp = parse_float(kp_row.get(2, ""))
            right_kp = parse_float(kp_row.get(3, ""))
            left_diam = "" if left_kp is None or left_band is None else str(int(round(left_kp + left_band * 2)))
            right_diam = "" if right_kp is None or right_band is None else str(int(round(right_kp + right_band * 2)))

            values = {
                0: section_value,
                1: str(row_index + 1),
                2: row[0],
                3: row[1],
                4: row[2],
                5: row[3],
                6: row[4],
                7: row[5],
                8: row[6],
                9: row[7],
                10: left_diam,
                11: right_diam,
            }
            for col, value in values.items():
                value = text(value).strip()
                if value:
                    archive_rows.append((year, measurement_date, locomotive, repair_type, table_row, col, value))

        return {
            "year": year,
            "measurement_date": measurement_date,
            "locomotive": locomotive,
            "repair_type": repair_type,
            "axis_count": axis_count,
        }, archive_rows, []


def save_archive(payload: dict) -> dict:
    meta, archive_rows, missing_cells = build_archive_rows(payload)
    if meta.get("error") == "missing":
        return {"error": "Не все обязательные ячейки заполнены.", "missing_cells": missing_cells}, 400
    if meta.get("error") == "repair":
        return {"error": "Выберите вид ремонта."}, 400

    year = int(meta["year"])
    measurement_date = meta["measurement_date"]
    locomotive = meta["locomotive"]
    repair_type = meta["repair_type"]
    overwrite = bool(payload.get("overwrite"))
    try:
        wheel_pair_count = int(payload.get("wheel_pair_count") or 0)
    except Exception:
        wheel_pair_count = 0
    try:
        section_count = int(payload.get("section_count") or 0)
    except Exception:
        section_count = 0

    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        existing = cur.execute(
            "SELECT COUNT(*) FROM archive_data WHERE y=? AND measurement_date=? AND locomotive=? AND repair_type=?",
            (year, measurement_date, locomotive, repair_type),
        ).fetchone()[0]
        if existing and not overwrite:
            return {
                "error": "duplicate",
                "message": "Запись с таким локомотивом, датой и видом ремонта уже есть в архиве.",
            }, 409

        cur.execute("BEGIN")
        cur.execute(
            "DELETE FROM archive_data WHERE y=? AND measurement_date=? AND locomotive=? AND repair_type=?",
            (year, measurement_date, locomotive, repair_type),
        )
        cur.executemany(
            "INSERT INTO archive_data (y, measurement_date, locomotive, repair_type, r, c, v) VALUES (?, ?, ?, ?, ?, ?, ?)",
            archive_rows,
        )
        cur.execute("DELETE FROM input_data WHERE y=? AND locomotive=?", (year, locomotive))
        cur.execute(
            "INSERT OR REPLACE INTO input_meta(y, locomotive, measurement_date, wheel_pair_count, section_count) VALUES(?,?,?,?,?)",
            (
                year,
                locomotive,
                measurement_date,
                wheel_pair_count or None,
                section_count or None,
            ),
        )
        conn.commit()

    return load_state(locomotive)


def update_archive_cells(payload: dict) -> dict:
    changes = payload.get("changes") or []
    if not isinstance(changes, list) or not changes:
        return {"error": "Нет изменений для сохранения."}, 400

    normalized: list[tuple[int, str, str, str, int, int, str]] = []
    for change in changes:
        if not isinstance(change, dict):
            return {"error": "Некорректный формат изменений."}, 400
        try:
            year = int(change.get("year"))
            measurement_date = text(change.get("measurement_date")).strip()
            locomotive = text(change.get("locomotive")).strip()
            repair_type = normalize_repair_type(change.get("repair_type"))
            source_r = int(change.get("source_r"))
            display_col = int(change.get("display_col"))
            if display_col < 10 or display_col > 19:
                return {"error": "Можно редактировать только правую часть таблицы архива."}, 400
            source_c = display_col - 8
            value = text(change.get("value"))
        except Exception:
            return {"error": "Некорректные данные изменения архива."}, 400

        if not measurement_date or not locomotive or not repair_type:
            return {"error": "Не удалось определить строку архива для изменения."}, 400

        normalized.append((year, measurement_date, locomotive, repair_type, source_r, source_c, value))

    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        cur.execute("BEGIN")
        for year, measurement_date, locomotive, repair_type, source_r, source_c, value in normalized:
            cur.execute(
                "INSERT OR REPLACE INTO archive_data (y, measurement_date, locomotive, repair_type, r, c, v) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (year, measurement_date, locomotive, repair_type, source_r, source_c, value),
            )
        conn.commit()

    return {"ok": True}


def delete_archive_measurement(payload: dict) -> dict:
    try:
        year = int(payload.get("year"))
        measurement_date = text(payload.get("measurement_date")).strip()
        locomotive = text(payload.get("locomotive")).strip()
        repair_type = normalize_repair_type(payload.get("repair_type"))
    except Exception:
        return {"error": "Некорректные данные для удаления архива."}, 400

    if not measurement_date or not locomotive or not repair_type:
        return {"error": "Не удалось определить замер для удаления."}, 400

    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        cur.execute(
            "DELETE FROM archive_data WHERE y=? AND measurement_date=? AND locomotive=? AND repair_type=?",
            (year, measurement_date, locomotive, repair_type),
        )
        deleted = cur.rowcount
        conn.commit()

    return {"ok": True, "deleted": deleted}


def load_norms_rows() -> list[dict[str, str | bool]]:
    default_keys = [item[0] for item in DEFAULT_NORMS]
    with DB_LOCK, connect() as conn:
        rows = conn.execute(
            "SELECT metric_key, label, condition, yellow_value, red_value FROM kp_norms_data ORDER BY rowid"
        ).fetchall()

    db_norms = {text(row["metric_key"]): row for row in rows}
    result: list[dict[str, str | bool]] = []
    for metric_key, label, condition, yellow, red in DEFAULT_NORMS:
        row = db_norms.get(metric_key)
        result.append(
            {
                "metric_key": metric_key,
                "label": label,
                "condition": text(row["condition"] if row else condition),
                "yellow_value": text(row["yellow_value"] if row else yellow),
                "red_value": text(row["red_value"] if row else red),
                "is_default": True,
            }
        )

    for row in rows:
        metric_key = text(row["metric_key"])
        if metric_key in default_keys:
            continue
        result.append(
            {
                "metric_key": metric_key,
                "label": text(row["label"]),
                "condition": text(row["condition"]),
                "yellow_value": text(row["yellow_value"]),
                "red_value": text(row["red_value"]),
                "is_default": False,
            }
        )
    return result


def save_norms_rows(payload: dict) -> dict:
    rows = payload.get("rows")
    if not isinstance(rows, list):
        return {"error": "rows"}, HTTPStatus.BAD_REQUEST

    normalized: list[tuple[str, str, str, str, str]] = []
    seen: set[str] = set()
    default_keys = {item[0] for item in DEFAULT_NORMS}
    conditions = {"меньше или равно", "больше или равно"}
    for index, row in enumerate(rows):
        if not isinstance(row, dict):
            continue
        metric_key = text(row.get("metric_key")).strip()
        label = text(row.get("label")).strip()
        condition = text(row.get("condition")).strip()
        yellow = text(row.get("yellow_value")).strip()
        red = text(row.get("red_value")).strip()
        if not label:
            continue
        if not metric_key:
            metric_key = f"custom_{index + 1}"
        if metric_key in seen:
            suffix = 2
            base_key = metric_key
            while f"{base_key}_{suffix}" in seen:
                suffix += 1
            metric_key = f"{base_key}_{suffix}"
        if condition not in conditions:
            condition = "меньше или равно"
        if metric_key in default_keys:
            default_label = next((item[1] for item in DEFAULT_NORMS if item[0] == metric_key), label)
            label = default_label
        seen.add(metric_key)
        normalized.append((metric_key, label, condition, yellow, red))

    with DB_LOCK, connect() as conn:
        conn.execute("DELETE FROM kp_norms_data")
        conn.executemany(
            """
            INSERT INTO kp_norms_data(metric_key, label, condition, yellow_value, red_value)
            VALUES(?,?,?,?,?)
            """,
            normalized,
        )
        conn.commit()

    return {"ok": True, "rows": load_norms_rows()}
def phone_json_number(value):
    raw = text(value).strip()
    if not raw:
        return None
    try:
        number = float(raw.replace(",", "."))
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def decode_phone_payload_text(raw_text: str) -> dict | None:
    raw_text = text(raw_text).strip()
    if not raw_text:
        return None
    try:
        decoded_bytes = base64.b64decode(raw_text.encode("utf-8"))
        decompressed = zlib.decompress(decoded_bytes)
        payload = json.loads(decompressed.decode("utf-8"))
        if isinstance(payload, dict):
            return payload
    except Exception:
        pass
    try:
        payload = json.loads(raw_text)
        if isinstance(payload, dict):
            return payload
    except Exception:
        pass
    return None


def require_qrcode():
    try:
        import qrcode
        from qrcode.image.svg import SvgImage
    except ImportError as exc:
        raise RuntimeError("Для QR нужен пакет qrcode.") from exc
    return qrcode, SvgImage


def build_phone_reference_payload(selected_numbers: list[str] | None = None) -> dict:
    selected = {text(number).strip() for number in (selected_numbers or []) if text(number).strip()}
    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        locomotives = load_inventory_records(cur, include_deleted=True)
        if selected:
            locomotives = [item for item in locomotives if item["number"] in selected]

        exported = []
        for loco in locomotives:
            number = loco["number"]
            series = loco.get("series", "")
            axis_count = locomotive_axis_count(series, number)
            kp_rows = cur.execute(
                "SELECT r, c, v FROM kp_data WHERE locomotive=? AND c IN (0, 1, 2, 3) ORDER BY r, c",
                (number,),
            ).fetchall()
            kp_values = {(int(row["r"]), int(row["c"])): text(row["v"]).strip() for row in kp_rows}
            wheel_pairs = []
            for row_index in range(axis_count):
                wheel_pairs.append(
                    {
                        "number": phone_json_number(kp_values.get((row_index, 0), str(row_index + 1))),
                        "axisNumber": phone_json_number(kp_values.get((row_index, 1), str(row_index + 1))),
                        "diameterLeft": phone_json_number(kp_values.get((row_index, 2), "")),
                        "diameterRight": phone_json_number(kp_values.get((row_index, 3), "")),
                    }
                )
            exported.append(
                {
                    "series": series,
                    "number": number,
                    "wheelPairCount": axis_count,
                    "wheelPairs": wheel_pairs,
                    "manufactureYear": text(loco.get("manufactureYear") or loco.get("manufacture_year") or ""),
                    "serviceLife": text(loco.get("serviceLife") or loco.get("service_life") or ""),
                    "sortOrder": int(loco.get("sortOrder") or 0),
                    "updatedAt": int(loco.get("updatedAt") or 0),
                    "deletedAt": int(loco.get("deletedAt") or 0),
                }
            )

    return {
        "formatVersion": 2,
        "exportType": "referenceData",
        "exportedAt": dt.datetime.now().astimezone().isoformat(timespec="seconds"),
        "locomotives": exported,
    }


def build_phone_archive_measurement_record(year_value, measurement_date, locomotive_value, repair_type_value, rows_by_index) -> dict | None:
    locomotive_value = text(locomotive_value).strip()
    repair_type_value = normalize_repair_type(repair_type_value)
    measurement_date = text(measurement_date).strip()
    if not locomotive_value or not repair_type_value or not measurement_date:
        return None

    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        series = series_for_locomotive(cur, locomotive_value)
        wheel_pair_count = locomotive_axis_count(series, locomotive_value)
        wheel_pairs = []
        for row_index in sorted(rows_by_index.keys()):
            if row_index < 2:
                continue
            values = rows_by_index.get(row_index, {})
            pair_number = row_index - 1
            left = {
                "flangeThickness": phone_json_number(values.get(4, "")),
                "flangeWear": phone_json_number(values.get(2, "")),
                "flangeSteepness": phone_json_number(values.get(6, "")),
                "bandageThickness": phone_json_number(values.get(8, "")),
                "bandageDiameter": phone_json_number(values.get(10, "")),
            }
            right = {
                "flangeThickness": phone_json_number(values.get(5, "")),
                "flangeWear": phone_json_number(values.get(3, "")),
                "flangeSteepness": phone_json_number(values.get(7, "")),
                "bandageThickness": phone_json_number(values.get(9, "")),
                "bandageDiameter": phone_json_number(values.get(11, "")),
            }
            wheel_pairs.append({"number": pair_number, "left": left, "right": right})

    return {
        "createdAt": dt.datetime.now().astimezone().isoformat(timespec="seconds"),
        "measurementId": f"pc-{year_value}-{measurement_date}-{locomotive_value}-{repair_type_value}",
        "locomotive": {
            "series": series,
            "number": locomotive_value,
            "wheelPairCount": wheel_pair_count,
            "comment": "",
            "isNew": False,
        },
        "repairType": repair_type_value,
        "measurementDate": measurement_date,
        "wheelPairs": wheel_pairs,
        "source": "pc",
    }


def build_phone_archive_payload(selected_locomotives: list[str] | None = None, selected_period: tuple[str, str] | None = None) -> dict:
    locomotives_filter = {text(item).strip() for item in (selected_locomotives or []) if text(item).strip()}
    date_from = date_to = ""
    if selected_period:
        date_from, date_to = selected_period
    with DB_LOCK, connect() as conn:
        query = """
            SELECT y, measurement_date, locomotive, repair_type, r, c, v
            FROM archive_data
            WHERE TRIM(COALESCE(measurement_date, '')) <> ''
        """
        params: list[str] = []
        if locomotives_filter:
            placeholders = ",".join("?" for _ in locomotives_filter)
            query += f" AND locomotive IN ({placeholders})"
            params.extend(sorted(locomotives_filter))
        if date_from:
            query += " AND measurement_date >= ?"
            params.append(date_from)
        if date_to:
            query += " AND measurement_date <= ?"
            params.append(date_to)
        query += " ORDER BY y DESC, measurement_date, locomotive, repair_type, r, c"
        rows = conn.execute(query, params).fetchall()

    grouped: dict[tuple[int, str, str, str], dict[int, dict[int, str]]] = {}
    for row in rows:
        r = int(row["r"])
        if r < 2:
            continue
        key = (int(row["y"] or 0), text(row["measurement_date"]), text(row["locomotive"]), text(row["repair_type"]))
        grouped.setdefault(key, {})
        grouped[key].setdefault(r, {})[int(row["c"])] = text(row["v"])

    archive = []
    for (year, measurement_date, locomotive, repair_type), values_by_row in grouped.items():
        record = build_phone_archive_measurement_record(year, measurement_date, locomotive, repair_type, values_by_row)
        if record:
            archive.append(record)

    return {
        "formatVersion": 1,
        "exportType": "archiveData",
        "exportedAt": dt.datetime.now().astimezone().isoformat(timespec="seconds"),
        "archive": archive,
    }


def build_phone_qr_frames(payload: dict) -> list[str]:
    qrcode, SvgImage = require_qrcode()
    json_data = json.dumps(payload, separators=(",", ":"), ensure_ascii=False)
    compressed = zlib.compress(json_data.encode("utf-8"), level=9)
    base64_str = base64.b64encode(compressed).decode("utf-8")
    chunk_size = 400
    chunks = [base64_str[i : i + chunk_size] for i in range(0, len(base64_str), chunk_size)] or [""]
    frames = []
    for index, chunk in enumerate(chunks):
        frame_header = f"{index + 1:02d}/{len(chunks):02d}|{chunk}"
        qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
        qr.add_data(frame_header)
        qr.make(fit=True)
        image = qr.make_image(image_factory=SvgImage)
        buf = io.BytesIO()
        image.save(buf)
        frames.append("data:image/svg+xml;base64," + base64.b64encode(buf.getvalue()).decode("ascii"))
    return frames


def build_phone_export_payload(kind: str, selected_locomotives: list[str] | None = None, selected_period: tuple[str, str] | None = None) -> dict:
    kind = text(kind).strip().lower()
    if kind == "reference":
        return build_phone_reference_payload(selected_locomotives)
    return build_phone_archive_payload(selected_locomotives, selected_period)


def upsert_inventory_locomotive(
    cur: sqlite3.Cursor,
    series: str,
    locomotive: str,
    inv: str = "",
    wheel_pair_count: int = 0,
    section_count: int = 0,
    eight_digit_number: str = "",
    manufacture_year: str = "",
    service_life: str = "",
    sort_order: int = 0,
    updated_at: int | None = None,
    deleted_at: int | None = None,
) -> None:
    locomotive = text(locomotive).strip()
    if not locomotive:
        return
    series = text(series).strip().upper()
    inv = text(inv).strip()
    eight_digit_number = text(eight_digit_number).strip()
    manufacture_year = text(manufacture_year).strip()
    service_life = text(service_life).strip()
    year = dt.date.today().year
    sort_order_value = int(sort_order or 0)
    updated_at = int(updated_at or int(dt.datetime.now().timestamp() * 1000))
    if sort_order_value <= 0:
        existing = cur.execute(
            "SELECT COALESCE(sort_order, 0) AS sort_order FROM inventory WHERE UPPER(TRIM(COALESCE(num, ''))) = UPPER(TRIM(?)) ORDER BY COALESCE(updated_at, 0) DESC, COALESCE(deleted_at, 0) DESC, COALESCE(sort_order, 0) ASC, rowid DESC LIMIT 1",
            (locomotive,),
        ).fetchone()
        if existing:
            sort_order_value = int(existing["sort_order"] or 0)
    if sort_order_value <= 0:
        max_row = cur.execute(
            "SELECT COALESCE(MAX(sort_order), 0) AS max_sort_order FROM inventory"
        ).fetchone()
        sort_order_value = int(max_row["max_sort_order"] or 0) + 1
    exact = cur.execute(
        "SELECT rowid, inv, COALESCE(sort_order, 0) AS sort_order, COALESCE(wheel_pair_count, 0) AS wheel_pair_count, COALESCE(section_count, 0) AS section_count, COALESCE(eight_digit_number, '') AS eight_digit_number, COALESCE(manufacture_year, '') AS manufacture_year, COALESCE(service_life, '') AS service_life, COALESCE(deleted_at, 0) AS deleted_at "
        "FROM inventory WHERE UPPER(TRIM(COALESCE(ser, ''))) = UPPER(TRIM(?)) AND TRIM(COALESCE(num, ''))=? "
        "ORDER BY COALESCE(updated_at, 0) DESC, COALESCE(deleted_at, 0) DESC, y DESC, rowid DESC LIMIT 1",
        (series, locomotive),
    ).fetchone()
    if exact:
        if sort_order_value <= 0:
            sort_order_value = int(exact["sort_order"] or 0)
        if not inv:
            inv = text(exact["inv"]).strip()
        if not wheel_pair_count:
            wheel_pair_count = int(exact["wheel_pair_count"] or 0)
        if not section_count:
            section_count = int(exact["section_count"] or 0)
        if not eight_digit_number:
            eight_digit_number = text(exact["eight_digit_number"]).strip()
        if not manufacture_year:
            manufacture_year = text(exact["manufacture_year"]).strip()
        if not service_life:
            service_life = text(exact["service_life"]).strip()
        if deleted_at is None:
            deleted_at = int(exact["deleted_at"] or 0)
        cur.execute(
            """
            UPDATE inventory
            SET ser=?, num=?, inv=?, wheel_pair_count=?, section_count=?, eight_digit_number=?, manufacture_year=?, service_life=?, sort_order=?, updated_at=?, deleted_at=?
            WHERE rowid=?
            """,
            (series, locomotive, inv, wheel_pair_count or None, section_count or None, eight_digit_number, manufacture_year, service_life, sort_order_value, updated_at, int(deleted_at or 0), int(exact["rowid"])),
        )
        cur.execute(
            "DELETE FROM inventory WHERE UPPER(TRIM(COALESCE(ser, ''))) = UPPER(TRIM(?)) AND TRIM(COALESCE(num, ''))=? AND rowid<>?",
            (series, locomotive, int(exact["rowid"])),
        )
        return
    if deleted_at is None:
        deleted_at = 0
    cur.execute(
        """
        INSERT INTO inventory (y, ser, num, inv, wheel_pair_count, section_count, eight_digit_number, manufacture_year, service_life, sort_order, updated_at, deleted_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (year, series, locomotive, inv, wheel_pair_count or None, section_count or None, eight_digit_number, manufacture_year, service_life, sort_order_value, updated_at, int(deleted_at or 0)),
    )

    if int(deleted_at or 0) <= 0:
        for index in range(max(1, int(wheel_pair_count or 1))):
            cur.execute("INSERT OR IGNORE INTO kp_data (locomotive, r, c, v) VALUES (?, ?, 0, ?)", (locomotive, index, str(index + 1)))
            cur.execute("INSERT OR IGNORE INTO kp_data (locomotive, r, c, v) VALUES (?, ?, 1, ?)", (locomotive, index, str(index + 1)))


def ensure_phone_locomotive(cur: sqlite3.Cursor, series: str, loco_number: str, wheel_pair_count: int) -> None:
    loco_number = text(loco_number).strip()
    if not loco_number:
        return
    existing = cur.execute(
        "SELECT inv, COALESCE(wheel_pair_count, 0) AS wheel_pair_count, COALESCE(section_count, 0) AS section_count, COALESCE(eight_digit_number, '') AS eight_digit_number, COALESCE(manufacture_year, '') AS manufacture_year, COALESCE(service_life, '') AS service_life, COALESCE(sort_order, 0) AS sort_order, COALESCE(deleted_at, 0) AS deleted_at "
        "FROM inventory WHERE UPPER(TRIM(COALESCE(num, ''))) = UPPER(TRIM(?)) LIMIT 1",
        (loco_number,),
    ).fetchone()
    if existing and int(existing["deleted_at"] or 0) > 0:
        return
    if existing:
        upsert_inventory_locomotive(
            cur,
            text(series).strip().upper(),
            loco_number,
            inv=text(existing["inv"]).strip(),
            wheel_pair_count=int(existing["wheel_pair_count"] or 0),
            section_count=int(existing["section_count"] or 0),
            eight_digit_number=text(existing["eight_digit_number"]).strip(),
            manufacture_year=text(existing["manufacture_year"]).strip(),
            service_life=text(existing["service_life"]).strip(),
            sort_order=int(existing["sort_order"] or 0),
            deleted_at=0,
        )
        return
    upsert_inventory_locomotive(
        cur,
        text(series).strip().upper(),
        loco_number,
        wheel_pair_count=wheel_pair_count,
        section_count=default_section_count(wheel_pair_count),
        deleted_at=0,
    )


def phone_measurement_missing_fields(payload: dict) -> list[str]:
    missing = []

    def has_value(value):
        if value is None:
            return False
        if isinstance(value, str):
            return value.strip() != ""
        return True

    if not has_value(payload.get("formatVersion")):
        missing.append("formatVersion")
    if not has_value(payload.get("createdAt")):
        missing.append("createdAt")
    if not has_value(payload.get("measurementId")):
        missing.append("measurementId")
    if not has_value(payload.get("repairType")):
        missing.append("repairType")
    if not has_value(payload.get("measurementDate")):
        missing.append("measurementDate")

    locomotive = payload.get("locomotive")
    if not isinstance(locomotive, dict):
        missing.append("locomotive")
        locomotive = {}

    for key in ("series", "number", "wheelPairCount"):
        if not has_value(locomotive.get(key)):
            missing.append(f"locomotive.{key}")

    wheel_pairs = payload.get("wheelPairs")
    if not isinstance(wheel_pairs, list) or not wheel_pairs:
        missing.append("wheelPairs")
        return missing

    side_fields = ("flangeThickness", "flangeWear", "flangeSteepness", "bandageThickness")
    for pair_index, pair in enumerate(wheel_pairs, start=1):
        if not isinstance(pair, dict):
            missing.append(f"wheelPairs[{pair_index}]")
            continue
        if not has_value(pair.get("number")):
            missing.append(f"wheelPairs[{pair_index}].number")
        for side_name in ("left", "right"):
            side = pair.get(side_name)
            if not isinstance(side, dict):
                missing.append(f"wheelPairs[{pair_index}].{side_name}")
                continue
            for field_name in side_fields:
                if not has_value(side.get(field_name)):
                    missing.append(f"wheelPairs[{pair_index}].{side_name}.{field_name}")

    return missing


def phone_archive_rows_from_payload(
    measurement_date: str,
    loco_number: str,
    repair_type: str,
    wheel_pairs: list[dict],
    kp_values: dict[tuple[int, int], str] | None = None,
) -> list[tuple[int, str, str, str, int, int, str]]:
    year_value = int(measurement_date[:4]) if len(measurement_date) >= 4 and measurement_date[:4].isdigit() else dt.date.today().year
    rows: list[tuple[int, str, str, str, int, int, str]] = []
    kp_values = kp_values or {}

    def parse_float(v):
        if v is None:
            return None
        try:
            return float(text(v).replace(",", "."))
        except ValueError:
            return None

    axis_count = len(wheel_pairs)
    for pair in wheel_pairs:
        pair_number = int(pair.get("number") or 1)
        kp_row_index = pair_number - 1
        table_row = pair_number + 1
        section_value = "1" if axis_count == 6 else str(((pair_number - 1) // 4) + 1)
        left = pair.get("left") or {}
        right = pair.get("right") or {}

        left_thick = parse_float(left.get("bandageThickness"))
        right_thick = parse_float(right.get("bandageThickness"))

        left_diam = left.get("bandageDiameter")
        if left_diam is None or text(left_diam).strip() == "":
            kp_left = parse_float(kp_values.get((kp_row_index, 2), ""))
            if kp_left is not None and left_thick is not None:
                left_diam = str(int(round(kp_left + left_thick * 2)))
            else:
                left_diam = ""

        right_diam = right.get("bandageDiameter")
        if right_diam is None or text(right_diam).strip() == "":
            kp_right = parse_float(kp_values.get((kp_row_index, 3), ""))
            if kp_right is not None and right_thick is not None:
                right_diam = str(int(round(kp_right + right_thick * 2)))
            else:
                right_diam = ""

        values = {
            0: section_value,
            1: str(pair_number),
            2: phone_json_number(left.get("flangeWear")),
            3: phone_json_number(right.get("flangeWear")),
            4: phone_json_number(left.get("flangeThickness")),
            5: phone_json_number(right.get("flangeThickness")),
            6: phone_json_number(left.get("flangeSteepness")),
            7: phone_json_number(right.get("flangeSteepness")),
            8: phone_json_number(left.get("bandageThickness")),
            9: phone_json_number(right.get("bandageThickness")),
            10: phone_json_number(left_diam),
            11: phone_json_number(right_diam),
        }

        for col, value in values.items():
            if value != "" and value is not None:
                rows.append((year_value, measurement_date, loco_number, repair_type, table_row, col, text(value)))

    return rows


def import_phone_reference_payload(payload: dict) -> dict:
    locomotives = payload.get("locomotives", [])
    if not isinstance(locomotives, list) or not locomotives:
        return {"error": "В файле справочника нет локомотивов."}, HTTPStatus.BAD_REQUEST

    grouped: dict[str, dict[str, object]] = {}
    for loco in locomotives:
        if not isinstance(loco, dict):
            continue
        series = text(loco.get("series")).strip().upper()
        number = text(loco.get("number")).strip()
        if not number:
            continue
        key = f"{series}|{number}"
        candidate = {
            "series": series,
            "number": number,
            "sortOrder": int(loco.get("sortOrder") or 0),
            "updatedAt": int(loco.get("updatedAt") or 0),
            "deletedAt": int(loco.get("deletedAt") or 0),
            "wheelPairCount": int(loco.get("wheelPairCount") or 6),
            "wheelPairs": loco.get("wheelPairs", []),
        }
        current = grouped.get(key)
        if current is None or (
            candidate["updatedAt"],
            candidate["deletedAt"],
            candidate["sortOrder"],
        ) > (
            current["updatedAt"],
            current["deletedAt"],
            current["sortOrder"],
        ):
            grouped[key] = candidate

    imported_count = 0
    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        cur.execute("BEGIN")
        for loco in grouped.values():
            series = text(loco["series"]).strip().upper()
            number = text(loco["number"]).strip()
            sort_order = int(loco["sortOrder"] or 0)
            updated_at = int(loco["updatedAt"] or 0)
            deleted_at = int(loco["deletedAt"] or 0)
            wheel_pair_count = int(loco["wheelPairCount"] or 6)
            upsert_inventory_locomotive(
                cur,
                series,
                number,
                wheel_pair_count=wheel_pair_count,
                sort_order=sort_order,
                deleted_at=deleted_at,
            )
            wheel_pairs = loco.get("wheelPairs", [])
            if not isinstance(wheel_pairs, list):
                continue
            if deleted_at > 0:
                cur.execute("DELETE FROM kp_data WHERE locomotive=?", (number,))
                imported_count += 1
                continue
            cur.execute("DELETE FROM kp_data WHERE locomotive=?", (number,))
            for pair in wheel_pairs:
                if not isinstance(pair, dict):
                    continue
                try:
                    r = int(pair.get("number") or 1) - 1
                except Exception:
                    continue
                kp_number = text(pair.get("number") or r + 1)
                axis_number = text(pair.get("axisNumber") or kp_number)
                diameter_left = text(pair.get("diameterLeft") or "").strip()
                diameter_right = text(pair.get("diameterRight") or "").strip()
                cur.execute("INSERT OR REPLACE INTO kp_data (locomotive, r, c, v) VALUES (?, ?, 0, ?)", (number, r, kp_number))
                cur.execute("INSERT OR REPLACE INTO kp_data (locomotive, r, c, v) VALUES (?, ?, 1, ?)", (number, r, axis_number))
                if diameter_left:
                    cur.execute("INSERT OR REPLACE INTO kp_data (locomotive, r, c, v) VALUES (?, ?, 2, ?)", (number, r, diameter_left))
                if diameter_right:
                    cur.execute("INSERT OR REPLACE INTO kp_data (locomotive, r, c, v) VALUES (?, ?, 3, ?)", (number, r, diameter_right))
            imported_count += 1
        conn.commit()

    return {"ok": True, "imported_count": imported_count}


def import_phone_archive_payload(payload: dict) -> dict:
    archive_items = payload.get("archive", [])
    if not isinstance(archive_items, list) or not archive_items:
        return {"error": "В архивном JSON нет данных."}, HTTPStatus.BAD_REQUEST

    imported_measurements = 0
    imported_cells = 0
    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        cur.execute("BEGIN")
        for item in archive_items:
            if not isinstance(item, dict):
                continue
            locomotive = item.get("locomotive") or {}
            if not isinstance(locomotive, dict):
                continue
            series = text(locomotive.get("series")).strip()
            loco_number = text(locomotive.get("number")).strip()
            measurement_date = text(item.get("measurementDate")).strip()
            repair_type = normalize_repair_type(item.get("repairType"))
            wheel_pairs = item.get("wheelPairs") or []
            if not loco_number or not measurement_date or not repair_type or not isinstance(wheel_pairs, list) or not wheel_pairs:
                continue
            wheel_pair_count = int(locomotive.get("wheelPairCount") or len(wheel_pairs))
            ensure_phone_locomotive(cur, series, loco_number, wheel_pair_count)
            kp_rows = cur.execute("SELECT r, c, v FROM kp_data WHERE locomotive=?", (loco_number,)).fetchall()
            kp_values = {(int(r), int(c)): text(v) for r, c, v in kp_rows}
            rows = phone_archive_rows_from_payload(measurement_date, loco_number, repair_type, wheel_pairs, kp_values)
            if not rows:
                continue
            year_value = int(measurement_date[:4]) if len(measurement_date) >= 4 and measurement_date[:4].isdigit() else dt.date.today().year
            cur.execute(
                "DELETE FROM archive_data WHERE y=? AND measurement_date=? AND locomotive=? AND repair_type=?",
                (year_value, measurement_date, loco_number, repair_type),
            )
            cur.executemany(
                "INSERT OR REPLACE INTO archive_data (y, measurement_date, locomotive, repair_type, r, c, v) VALUES (?, ?, ?, ?, ?, ?, ?)",
                rows,
            )
            imported_measurements += 1
            imported_cells += len(rows)
        conn.commit()

    return {"ok": True, "imported_measurements": imported_measurements, "imported_cells": imported_cells}


def import_phone_measurement_payload(payload: dict) -> dict:
    if not isinstance(payload, dict):
        return {"error": "Некорректный JSON."}, HTTPStatus.BAD_REQUEST

    missing = phone_measurement_missing_fields(payload)
    if missing:
        return {"error": "В замере не заполнены обязательные данные.", "missing": missing[:20]}, HTTPStatus.BAD_REQUEST

    locomotive = payload.get("locomotive") or {}
    series = text(locomotive.get("series")).strip()
    loco_number = text(locomotive.get("number")).strip()
    measurement_date = text(payload.get("measurementDate")).strip()
    repair_type = normalize_repair_type(payload.get("repairType"))
    wheel_pairs = payload.get("wheelPairs") or []
    wheel_pair_count = int(locomotive.get("wheelPairCount") or len(wheel_pairs))

    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        cur.execute("BEGIN")
        ensure_phone_locomotive(cur, series, loco_number, wheel_pair_count)
        kp_rows = cur.execute("SELECT r, c, v FROM kp_data WHERE locomotive=?", (loco_number,)).fetchall()
        kp_values = {(int(r), int(c)): text(v) for r, c, v in kp_rows}
        rows = phone_archive_rows_from_payload(measurement_date, loco_number, repair_type, wheel_pairs, kp_values)
        if not rows:
            return {"error": "Не удалось собрать строки замера."}, HTTPStatus.BAD_REQUEST
        year_value = int(measurement_date[:4]) if len(measurement_date) >= 4 and measurement_date[:4].isdigit() else dt.date.today().year
        cur.execute(
            "DELETE FROM archive_data WHERE y=? AND measurement_date=? AND locomotive=? AND repair_type=?",
            (year_value, measurement_date, loco_number, repair_type),
        )
        cur.executemany(
            "INSERT OR REPLACE INTO archive_data (y, measurement_date, locomotive, repair_type, r, c, v) VALUES (?, ?, ?, ?, ?, ?, ?)",
            rows,
        )
        conn.commit()

    return {"ok": True, "imported_measurements": 1, "imported_cells": len(rows)}


def import_phone_payload(payload: dict) -> dict:
    if not isinstance(payload, dict):
        return {"error": "Некорректный JSON."}, HTTPStatus.BAD_REQUEST
    if payload.get("exportType") == "archiveData" or "archive" in payload:
        return import_phone_archive_payload(payload)
    if payload.get("exportType") == "referenceData" or "locomotives" in payload:
        return import_phone_reference_payload(payload)
    if int(payload.get("formatVersion", 0) or 0) == 1:
        return import_phone_measurement_payload(payload)
    return {"error": "Нераспознанный формат телефона."}, HTTPStatus.BAD_REQUEST


def require_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise RuntimeError("Для работы с Excel нужен пакет openpyxl.") from exc
    return Workbook, load_workbook, Alignment, Border, Font, PatternFill, Side, DataValidation


def build_archive_workbook():
    Workbook, _, Alignment, Border, Font, PatternFill, Side, DataValidation = require_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "Архив"
    ws.append(ARCHIVE_EXCEL_HEADERS)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="BFBFBF")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = [14, 12, 14, 12, 10, 10, 14, 14, 18, 18, 18, 18, 18, 18, 18, 18]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, index).column_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:P1"
    ws.row_dimensions[1].height = 34

    return wb

def archive_excel_template_bytes() -> bytes:
    wb = build_archive_workbook()
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def parse_excel_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    raw = text(value).strip()
    if not raw:
        return ""
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            pass
    return raw


def format_excel_export_date(value: str) -> str:
    raw = text(value).strip()
    if not raw:
        return ""
    try:
        return dt.datetime.strptime(raw, "%Y-%m-%d").strftime("%d.%m.%Y")
    except ValueError:
        return raw


def parse_excel_int(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    raw = text(value).strip().replace(",", ".")
    if not raw:
        return None
    try:
        number = float(raw)
    except ValueError:
        return None
    return int(number) if number.is_integer() else None


def parse_float_value(value) -> float | None:
    raw = text(value).strip().replace(",", ".")
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def excel_cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return text(value).strip()


def excel_num_text(value) -> str:
    raw = text(value).strip()
    return raw.replace(".", ",") if raw else ""


def normalize_excel_header(value) -> str:
    return "".join(ch for ch in text(value).strip().lower() if ch.isalnum())


def archive_export_locomotives(cur: sqlite3.Cursor) -> list[str]:
    rows = cur.execute(
        "SELECT DISTINCT locomotive FROM archive_data WHERE TRIM(COALESCE(locomotive, '')) <> '' ORDER BY locomotive"
    ).fetchall()
    return [text(row["locomotive"]).strip() for row in rows if text(row["locomotive"]).strip()]


def load_effective_kp_values_for_date(
    cur: sqlite3.Cursor,
    locomotive: str,
    measurement_date: str,
    cache: dict[tuple[str, str], dict[tuple[int, int], str]] | None = None,
) -> dict[tuple[int, int], str]:
    loco = text(locomotive).strip()
    date_key = text(measurement_date).strip()
    cache_key = (loco, date_key)
    if cache is not None and cache_key in cache:
        return cache[cache_key]

    kp_values: dict[tuple[int, int], str] = {}
    if loco:
        version_row = None
        if date_key:
            version_row = cur.execute(
                """
                SELECT id
                FROM kp_data_versions
                WHERE locomotive=? AND TRIM(COALESCE(valid_to, '')) <> '' AND valid_to >= ?
                ORDER BY valid_to ASC, id ASC
                LIMIT 1
                """,
                (loco, date_key),
            ).fetchone()

        if version_row:
            version_rows = cur.execute(
                "SELECT r, c, v FROM kp_data_version_cells WHERE version_id=? AND c IN (2, 3) ORDER BY r, c",
                (int(version_row["id"]),),
            ).fetchall()
            for row in version_rows:
                kp_values[(int(row["r"]), int(row["c"]))] = text(row["v"]).strip()

        current_rows = cur.execute(
            "SELECT r, c, v FROM kp_data WHERE locomotive=? AND c IN (2, 3) ORDER BY r, c",
            (loco,),
        ).fetchall()
        for row in current_rows:
            kp_values.setdefault((int(row["r"]), int(row["c"])), text(row["v"]).strip())

    if cache is not None:
        cache[cache_key] = kp_values
    return kp_values


def resolve_archive_diameter_pair(
    cur: sqlite3.Cursor,
    locomotive: str,
    measurement_date: str,
    pair_row_index: int,
    values: list[str],
    kp_cache: dict[tuple[str, str], dict[tuple[int, int], str]] | None = None,
) -> tuple[str, str, str, str]:
    bandage_left = values[8] or ""
    bandage_right = values[9] or ""
    fallback_left = values[10] or ""
    fallback_right = values[11] or ""
    kp_values = load_effective_kp_values_for_date(cur, locomotive, measurement_date, kp_cache)

    resolved_left = fallback_left
    left_bandage_value = parse_float_value(bandage_left)
    if left_bandage_value is not None:
        kp_left = parse_float_value(kp_values.get((pair_row_index, 2), ""))
        if kp_left is not None:
            resolved_left = str(int(round(kp_left + left_bandage_value * 2)))

    resolved_right = fallback_right
    right_bandage_value = parse_float_value(bandage_right)
    if right_bandage_value is not None:
        kp_right = parse_float_value(kp_values.get((pair_row_index, 3), ""))
        if kp_right is not None:
            resolved_right = str(int(round(kp_right + right_bandage_value * 2)))

    return bandage_left, bandage_right, resolved_left, resolved_right


def build_archive_export_rows(selected_locomotives: list[str] | None = None, date_from: str = "", date_to: str = "") -> list[list[str]]:
    locomotives_filter = {text(item).strip() for item in selected_locomotives or [] if text(item).strip()}
    date_from = text(date_from).strip()
    date_to = text(date_to).strip()
    query = """
        SELECT y, measurement_date, locomotive, repair_type, r, c, v
        FROM archive_data
        WHERE TRIM(COALESCE(measurement_date, '')) <> ''
    """
    params: list[str] = []
    if locomotives_filter:
        query += f" AND locomotive IN ({','.join('?' for _ in locomotives_filter)})"
        params.extend(sorted(locomotives_filter))
    if date_from:
        query += " AND measurement_date >= ?"
        params.append(date_from)
    if date_to:
        query += " AND measurement_date <= ?"
        params.append(date_to)
    query += " ORDER BY y, measurement_date, locomotive, repair_type, r, c"

    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        rows = cur.execute(query, params).fetchall()
        grouped: dict[tuple[int, str, str, str, int], list[str]] = {}
        for row in rows:
            r = int(row["r"])
            if r < 2:
                continue
            key = (int(row["y"] or 0), text(row["measurement_date"]), text(row["locomotive"]), normalize_repair_type(row["repair_type"]), r)
            grouped.setdefault(key, [""] * 12)
            c = int(row["c"])
            if 0 <= c < 12:
                grouped[key][c] = text(row["v"])

        kp_cache: dict[tuple[str, str], dict[tuple[int, int], str]] = {}
        export_rows: list[list[str]] = []
        for (_, measurement_date, locomotive, repair_type, r), values in sorted(grouped.items()):
            series = series_for_locomotive(cur, locomotive)
            bandage_left, bandage_right, diameter_left, diameter_right = resolve_archive_diameter_pair(
                cur,
                locomotive,
                measurement_date,
                r - 2,
                values,
                kp_cache,
            )

            export_rows.append(
                [
                    format_excel_export_date(measurement_date),
                    locomotive,
                    normalize_repair_type(repair_type),
                    series,
                    values[0] or "1",
                    values[1] or str(r - 1),
                    excel_num_text(values[2]),
                    excel_num_text(values[3]),
                    excel_num_text(values[4]),
                    excel_num_text(values[5]),
                    excel_num_text(values[6]),
                    excel_num_text(values[7]),
                    excel_num_text(bandage_left),
                    excel_num_text(bandage_right),
                    excel_num_text(diameter_left),
                    excel_num_text(diameter_right),
                ]
            )
        return export_rows


def archive_excel_export_bytes(selected_locomotives: list[str] | None = None, date_from: str = "", date_to: str = "") -> tuple[bytes, int]:
    _, _, Alignment, _, _, _, _, _ = require_openpyxl()

    wb = build_archive_workbook()
    ws = wb.active
    rows = build_archive_export_rows(selected_locomotives, date_from, date_to)
    out_row = 2
    for row_data in rows:
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(out_row, col, value)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        out_row += 1
    ws.auto_filter.ref = f"A1:P{max(1, out_row - 1)}"
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), len(rows)




def phone_reference_export_payload() -> dict:
    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        locomotives = load_inventory_records(cur, include_deleted=True)
        kp_rows = cur.execute(
            "SELECT locomotive, r, c, v FROM kp_data WHERE TRIM(COALESCE(locomotive, '')) <> '' ORDER BY locomotive, r, c"
        ).fetchall()

    kp_map: dict[str, dict[int, dict[int, str]]] = {}
    for row in kp_rows:
        locomotive = text(row["locomotive"]).strip()
        kp_map.setdefault(locomotive, {}).setdefault(int(row["r"]), {})[int(row["c"])] = text(row["v"])

    payload_locomotives: list[dict[str, object]] = []
    for locomotive in locomotives:
        number = text(locomotive.get("number")).strip()
        series = text(locomotive.get("series")).strip()
        wheel_pair_count = max(1, locomotive_axis_count(series, number))
        row_map = kp_map.get(number, {})
        wheel_pairs: list[dict[str, object]] = []
        for pair_index in range(wheel_pair_count):
            row = row_map.get(pair_index, {})
            wheel_pairs.append(
                {
                    "number": pair_index + 1,
                    "axisNumber": parse_excel_int(row.get(1, "")) or (pair_index + 1),
                    "diameterLeft": parse_float_value(row.get(2, "")),
                    "diameterRight": parse_float_value(row.get(3, "")),
                }
            )
        payload_locomotives.append(
            {
                "series": series,
                "number": number,
                "wheelPairCount": wheel_pair_count,
                "inventoryNumber": text(locomotive.get("inventoryNumber") or locomotive.get("inv") or ""),
                "invNumber": text(locomotive.get("inventoryNumber") or locomotive.get("inv") or ""),
                "eightDigitNumber": text(locomotive.get("eightDigitNumber") or ""),
                "manufactureYear": text(locomotive.get("manufactureYear") or locomotive.get("manufacture_year") or ""),
                "sortOrder": int(locomotive.get("sortOrder") or 0),
                "updatedAt": int(locomotive.get("updatedAt") or 0),
                "deletedAt": int(locomotive.get("deletedAt") or 0),
                "wheelPairs": wheel_pairs,
            }
        )

    return {
        "formatVersion": 2,
        "exportType": "referenceData",
        "exportedAt": dt.datetime.now().isoformat(timespec="seconds"),
        "locomotives": payload_locomotives,
    }


def phone_archive_export_payload(selected_locomotives: list[str] | None = None, date_from: str = "", date_to: str = "") -> dict:
    locomotives_filter = {text(item).strip() for item in selected_locomotives or [] if text(item).strip()}
    date_from = text(date_from).strip()
    date_to = text(date_to).strip()
    query = """
        SELECT y, measurement_date, locomotive, repair_type, r, c, v
        FROM archive_data
        WHERE TRIM(COALESCE(measurement_date, '')) <> ''
    """
    params: list[str] = []
    if locomotives_filter:
        query += f" AND locomotive IN ({','.join('?' for _ in locomotives_filter)})"
        params.extend(sorted(locomotives_filter))
    if date_from:
        query += " AND measurement_date >= ?"
        params.append(date_from)
    if date_to:
        query += " AND measurement_date <= ?"
        params.append(date_to)
    query += " ORDER BY y, measurement_date, locomotive, repair_type, r, c"

    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        rows = cur.execute(query, params).fetchall()
        grouped: dict[tuple[int, str, str, str], dict[int, list[str]]] = {}
        series_cache: dict[str, str] = {}
        kp_cache: dict[tuple[str, str], dict[tuple[int, int], str]] = {}
        for row in rows:
            key = (
                int(row["y"] or 0),
                text(row["measurement_date"]).strip(),
                text(row["locomotive"]).strip(),
                normalize_repair_type(row["repair_type"]),
            )
            grouped.setdefault(key, {})[int(row["r"])] = grouped.setdefault(key, {}).get(int(row["r"]), [""] * 12)
            grouped[key][int(row["r"])][int(row["c"])] = text(row["v"])

        archive_items: list[dict[str, object]] = []
        for (year, measurement_date, locomotive, repair_type), rows_by_r in sorted(grouped.items()):
            series = series_cache.get(locomotive)
            if series is None:
                series = series_for_locomotive(cur, locomotive)
                series_cache[locomotive] = series
            wheel_pair_count = max(1, locomotive_axis_count(series, locomotive))
            wheel_pairs: list[dict[str, object]] = []
            for row_index in sorted(rows_by_r):
                row = rows_by_r[row_index]
                pair_number = parse_excel_int(row[1]) or max(1, row_index - 1)
                left_band, right_band, diameter_left, diameter_right = resolve_archive_diameter_pair(
                    cur,
                    locomotive,
                    measurement_date,
                    pair_number - 1,
                    row,
                    kp_cache,
                )

                wheel_pairs.append(
                    {
                        "number": pair_number,
                        "left": {
                            "flangeThickness": parse_float_value(row[4]),
                            "flangeWear": parse_float_value(row[2]),
                            "flangeSteepness": parse_float_value(row[6]),
                            "bandageThickness": parse_float_value(left_band),
                            "bandageDiameter": parse_float_value(diameter_left),
                        },
                        "right": {
                            "flangeThickness": parse_float_value(row[5]),
                            "flangeWear": parse_float_value(row[3]),
                            "flangeSteepness": parse_float_value(row[7]),
                            "bandageThickness": parse_float_value(right_band),
                            "bandageDiameter": parse_float_value(diameter_right),
                        },
                    }
                )

            archive_items.append(
                {
                    "formatVersion": 1,
                    "createdAt": dt.datetime.now().isoformat(timespec="seconds"),
                    "measurementId": f"{year}:{measurement_date}:{locomotive}:{repair_type}",
                    "locomotive": {
                        "series": series,
                        "number": locomotive,
                        "wheelPairCount": wheel_pair_count,
                        "comment": "",
                        "isNew": False,
                    },
                    "repairType": repair_type,
                    "measurementDate": measurement_date,
                    "wheelPairs": wheel_pairs,
                }
            )

    return {
        "formatVersion": 1,
        "exportType": "archiveData",
        "exportedAt": dt.datetime.now().isoformat(timespec="seconds"),
        "archive": archive_items,
    }


def phone_export_payload(kind: str, selected_locomotives: list[str] | None = None, date_from: str = "", date_to: str = "") -> dict:
    kind = text(kind).strip().lower()
    if kind == "reference":
        return phone_reference_export_payload()
    if kind == "archive":
        return phone_archive_export_payload(selected_locomotives, date_from, date_to)
    raise ValueError("Неизвестный тип экспорта.")


def parse_phone_json_payload(raw: bytes) -> object:
    data = json.loads(raw.decode("utf-8"))
    if isinstance(data, dict) and "payload" in data:
        payload = data.get("payload")
        if isinstance(payload, str):
            return json.loads(payload)
        return payload
    if isinstance(data, str):
        text_data = data.strip()
        if text_data.startswith("{") or text_data.startswith("["):
            try:
                return json.loads(text_data)
            except Exception:
                pass
        try:
            decoded = base64.b64decode(text_data)
            try:
                return json.loads(zlib.decompress(decoded).decode("utf-8"))
            except Exception:
                return json.loads(decoded.decode("utf-8"))
        except Exception:
            return data
    return data


def import_phone_reference_payload(payload: dict) -> dict:
    reference = payload.get("referenceData") if isinstance(payload.get("referenceData"), dict) else payload
    locomotives = reference.get("locomotives") if isinstance(reference, dict) else None
    if not isinstance(locomotives, list):
        return {"error": "Некорректный справочник локомотивов."}, HTTPStatus.BAD_REQUEST

    grouped: dict[str, dict[str, object]] = {}
    for item in locomotives:
        if not isinstance(item, dict):
            continue
        series = text(item.get("series")).strip().upper()
        number = text(item.get("number")).strip()
        if not number:
            continue
        key = f"{series}|{number}"
        candidate = {
            "series": series,
            "number": number,
            "wheelPairCount": parse_excel_int(item.get("wheelPairCount")) or len(item.get("wheelPairs") or []) or locomotive_axis_count(series, number),
            "sortOrder": parse_excel_int(item.get("sortOrder")) or 0,
            "updatedAt": parse_excel_int(item.get("updatedAt")) or 0,
            "deletedAt": parse_excel_int(item.get("deletedAt")) or 0,
            "eightDigitNumber": text(item.get("eightDigitNumber")).strip(),
            "manufactureYear": text(item.get("manufactureYear") or item.get("manufacture_year")).strip(),
            "serviceLife": text(item.get("serviceLife") or item.get("service_life")).strip(),
            "wheelPairs": item.get("wheelPairs") if isinstance(item.get("wheelPairs"), list) else [],
        }
        current = grouped.get(key)
        if current is None or (
            candidate["updatedAt"],
            candidate["deletedAt"],
            candidate["sortOrder"],
        ) > (
            current["updatedAt"],
            current["deletedAt"],
            current["sortOrder"],
        ):
            grouped[key] = candidate

    imported_locomotives = 0
    imported_wheel_pairs = 0
    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        cur.execute("BEGIN")
        for item in grouped.values():
            series = text(item["series"]).strip().upper()
            number = text(item["number"]).strip()
            wheel_pair_count = int(item["wheelPairCount"] or 0)
            sort_order = int(item["sortOrder"] or 0)
            updated_at = int(item["updatedAt"] or 0)
            deleted_at = int(item["deletedAt"] or 0)
            eight_digit_number = text(item["eightDigitNumber"]).strip()
            manufacture_year = text(item["manufactureYear"]).strip()
            service_life = text(item["serviceLife"]).strip()
            upsert_inventory_locomotive(
                cur,
                series,
                number,
                "",
                wheel_pair_count=wheel_pair_count,
                eight_digit_number=eight_digit_number,
                manufacture_year=manufacture_year,
                service_life=service_life,
                sort_order=sort_order,
                updated_at=updated_at or None,
                deleted_at=deleted_at,
            )
            if deleted_at <= 0:
                cur.execute("DELETE FROM kp_data WHERE locomotive=?", (number,))
            wheel_pairs = item.get("wheelPairs")
            if not isinstance(wheel_pairs, list) or not wheel_pairs:
                wheel_pairs = [{"number": index + 1, "axisNumber": index + 1} for index in range(max(1, wheel_pair_count))]
            for pair in wheel_pairs:
                if not isinstance(pair, dict):
                    continue
                pair_number = parse_excel_int(pair.get("number")) or 0
                if pair_number <= 0:
                    continue
                axis_number = parse_excel_int(pair.get("axisNumber")) or pair_number
                cur.execute(
                    "INSERT OR REPLACE INTO kp_data(locomotive, r, c, v) VALUES(?, ?, ?, ?)",
                    (number, pair_number - 1, 1, str(axis_number)),
                )
                diameter_left = parse_float_value(pair.get("diameterLeft"))
                diameter_right = parse_float_value(pair.get("diameterRight"))
                if diameter_left is not None:
                    cur.execute(
                        "INSERT OR REPLACE INTO kp_data(locomotive, r, c, v) VALUES(?, ?, ?, ?)",
                        (number, pair_number - 1, 2, str(diameter_left)),
                    )
                if diameter_right is not None:
                    cur.execute(
                        "INSERT OR REPLACE INTO kp_data(locomotive, r, c, v) VALUES(?, ?, ?, ?)",
                        (number, pair_number - 1, 3, str(diameter_right)),
                    )
            imported_locomotives += 1
            imported_wheel_pairs += len(wheel_pairs)
        conn.commit()

    return {"ok": True, "imported_locomotives": imported_locomotives, "imported_wheel_pairs": imported_wheel_pairs}


def import_phone_measurement_payload(payload: dict) -> dict:
    measurement = payload.get("measurement") if isinstance(payload.get("measurement"), dict) else payload
    if not isinstance(measurement, dict):
        return {"error": "Некорректный замер."}, HTTPStatus.BAD_REQUEST

    locomotive = measurement.get("locomotive")
    if not isinstance(locomotive, dict):
        return {"error": "Не найден локомотив в замере."}, HTTPStatus.BAD_REQUEST

    series = text(locomotive.get("series")).strip()
    number = text(locomotive.get("number")).strip()
    measurement_date = text(measurement.get("measurementDate")).strip() or dt.date.today().isoformat()
    repair_type = normalize_repair_type(measurement.get("repairType"))
    measurement_id = text(measurement.get("measurementId")).strip() or f"{measurement_date}:{number}:{repair_type}"
    wheel_pairs = measurement.get("wheelPairs")
    if not isinstance(wheel_pairs, list) or not wheel_pairs:
        return {"error": "В замере нет колесных пар."}, HTTPStatus.BAD_REQUEST

    try:
        year = int(measurement_date[:4])
    except Exception:
        year = dt.date.today().year

    wheel_pair_count = parse_excel_int(locomotive.get("wheelPairCount")) or len(wheel_pairs) or locomotive_axis_count(series, number)
    section_count = default_section_count(wheel_pair_count)
    section_sizes: list[int] = []
    base = max(1, wheel_pair_count) // max(1, section_count)
    remainder = max(1, wheel_pair_count) % max(1, section_count)
    for index in range(max(1, section_count)):
        section_sizes.append(base + (1 if index < remainder else 0))

    archive_rows: list[tuple[int, str, str, str, int, int, str]] = []
    imported_cells = 0
    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        cur.execute("BEGIN")
        upsert_inventory_locomotive(cur, series, number)
        cur.execute(
            "DELETE FROM archive_data WHERE y=? AND measurement_date=? AND locomotive=? AND repair_type=?",
            (year, measurement_date, number, repair_type),
        )
        for pair in wheel_pairs:
            if not isinstance(pair, dict):
                continue
            pair_number = parse_excel_int(pair.get("number")) or 0
            if pair_number <= 0:
                continue
            row_index = pair_number - 1
            running = 0
            section_value = "1"
            for section_index, span in enumerate(section_sizes, start=1):
                running += span
                if row_index < running:
                    section_value = str(section_index)
                    break
            left = pair.get("left") if isinstance(pair.get("left"), dict) else {}
            right = pair.get("right") if isinstance(pair.get("right"), dict) else {}
            values = {
                0: section_value,
                1: str(pair_number),
                2: excel_num_text(left.get("flangeWear")),
                3: excel_num_text(right.get("flangeWear")),
                4: excel_num_text(left.get("flangeThickness")),
                5: excel_num_text(right.get("flangeThickness")),
                6: excel_num_text(left.get("flangeSteepness")),
                7: excel_num_text(right.get("flangeSteepness")),
                8: excel_num_text(left.get("bandageThickness")),
                9: excel_num_text(right.get("bandageThickness")),
                10: excel_num_text(left.get("bandageDiameter")),
                11: excel_num_text(right.get("bandageDiameter")),
            }
            for col, value in values.items():
                value = text(value).strip()
                if value:
                    archive_rows.append((year, measurement_date, number, repair_type, row_index + 2, col, value))
                    imported_cells += 1
        cur.executemany(
            "INSERT OR REPLACE INTO archive_data (y, measurement_date, locomotive, repair_type, r, c, v) VALUES (?, ?, ?, ?, ?, ?, ?)",
            archive_rows,
        )
        conn.commit()

    return {
        "ok": True,
        "imported_measurements": 1,
        "imported_cells": imported_cells,
        "measurement_id": measurement_id,
    }


def import_phone_payload(payload: object) -> dict:
    if isinstance(payload, dict):
        export_type = text(payload.get("exportType")).strip()
        if payload.get("archiveData") is not None or export_type == "archiveData":
            archive = payload.get("archiveData") if isinstance(payload.get("archiveData"), dict) else payload
            archive_items = archive.get("archive") if isinstance(archive, dict) else None
            if isinstance(archive_items, list):
                imported_measurements = 0
                imported_cells = 0
                for item in archive_items:
                    result = import_phone_measurement_payload(item)
                    if isinstance(result, tuple):
                        return result[0], result[1]
                    imported_measurements += int(result.get("imported_measurements", 0))
                    imported_cells += int(result.get("imported_cells", 0))
                return {"ok": True, "imported_measurements": imported_measurements, "imported_cells": imported_cells}
            return import_phone_measurement_payload(archive if isinstance(archive, dict) else payload)
        if payload.get("referenceData") is not None or export_type == "referenceData":
            reference = payload.get("referenceData") if isinstance(payload.get("referenceData"), dict) else payload
            return import_phone_reference_payload(reference)
        if payload.get("measurement") is not None:
            measurement = payload.get("measurement")
            if isinstance(measurement, dict):
                return import_phone_measurement_payload(measurement)
        if "locomotives" in payload:
            return import_phone_reference_payload(payload)
        if "wheelPairs" in payload and "measurementDate" in payload:
            return import_phone_measurement_payload(payload)
    return {"error": "Не удалось распознать телефонные данные."}, HTTPStatus.BAD_REQUEST


def ensure_import_locomotive(cur: sqlite3.Cursor, series: str, locomotive: str, wheel_pair_count: int) -> None:
    locomotive = text(locomotive).strip()
    if not locomotive:
        return
    series = text(series).strip().upper()
    upsert_inventory_locomotive(cur, series, locomotive)
    for index in range(max(1, int(wheel_pair_count or 1))):
        cur.execute("INSERT OR IGNORE INTO kp_data (locomotive, r, c, v) VALUES (?, ?, 0, ?)", (locomotive, index, str(index + 1)))
        cur.execute("INSERT OR IGNORE INTO kp_data (locomotive, r, c, v) VALUES (?, ?, 1, ?)", (locomotive, index, str(index + 1)))


def import_archive_excel_bytes(data: bytes) -> dict:
    _, load_workbook, *_ = require_openpyxl()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    aliases = {
        "measurement_date": {"датазамера", "датавыполненияобмера", "датаобмера", "measurementdate"},
        "series": {"серия", "series"},
        "locomotive": {"локомотив", "номерлокомотива", "locomotive"},
        "repair_type": {"видремонта", "видрем", "repairtype"},
        "wheel_pair": {"номеркп", "wheelpair", "wheelpairnumber", "ось", "номероси", "kp"},
        "section": {"секция", "вагон", "section"},
        "prokat_left": {"прокатлев", "левпрокат", "prokatleft", "flangewearleft"},
        "prokat_right": {"прокатправ", "правпрокат", "prokatright", "flangewearright"},
        "greben_left": {"толщинагребнялев", "левтолщинагребня", "grebenleft", "flangethicknessleft"},
        "greben_right": {"толщинагребняправ", "правтолщинагребня", "grebenright", "flangethicknessright"},
        "krut_left": {"крутизнагребнялев", "левкрутизнагребня", "krutleft", "flangesteepnessleft"},
        "krut_right": {"крутизнагребняправ", "правкрутизнагребня", "krutright", "flangesteepnessright"},
        "bandage_thickness_left": {"толщинабандажалева", "толщинабандажалев", "leftbandagethickness", "bandagethicknessleft"},
        "bandage_thickness_right": {"толщинабандажаправ", "rightbandagethickness", "bandagethicknessright"},
        "bandage_diameter_left": {"диаметрбандажалева", "диаметрбандажалев", "leftbandagediameter", "bandagediameterleft"},
        "bandage_diameter_right": {"диаметрбандажаправ", "rightbandagediameter", "bandagediameterright"},
    }

    header_row_index = None
    header_columns: dict[str, int] = {}
    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        normalized = [normalize_excel_header(value) for value in row]
        found: dict[str, int] = {}
        for key, alias_set in aliases.items():
            for idx, header in enumerate(normalized):
                if header in alias_set:
                    found[key] = idx
                    break
        if len(found) >= 5:
            header_row_index = row_index
            header_columns = found
            break

    if header_row_index is None:
        return {"error": "Не найдена строка заголовков."}, HTTPStatus.BAD_REQUEST

    missing_headers = [name for name in ("measurement_date", "locomotive", "repair_type", "wheel_pair") if name not in header_columns]
    if missing_headers:
        return {"error": "В Excel не найдены обязательные колонки: " + ", ".join(missing_headers)}, HTTPStatus.BAD_REQUEST

    required_metric_columns = {
        "prokat_left": 2,
        "prokat_right": 3,
        "greben_left": 4,
        "greben_right": 5,
        "krut_left": 6,
        "krut_right": 7,
        "bandage_thickness_left": 8,
        "bandage_thickness_right": 9,
    }
    optional_metric_columns = {"bandage_diameter_left": 10, "bandage_diameter_right": 11}

    measurements: dict[tuple[str, str, str], dict] = {}
    errors: list[str] = []
    current_series = current_date = current_loco = current_repair = current_section = ""

    def row_value(row, column_name: str) -> str:
        idx = header_columns.get(column_name)
        if idx is None or idx >= len(row):
            return ""
        return excel_cell_text(row[idx])

    for row_index, row in enumerate(ws.iter_rows(min_row=header_row_index + 1, values_only=True), start=header_row_index + 1):
        if not row or all(cell is None or text(cell).strip() == "" for cell in row):
            continue
        row_date = parse_excel_date(row_value(row, "measurement_date")) or current_date
        row_loco = row_value(row, "locomotive") or current_loco
        row_repair = row_value(row, "repair_type") or current_repair
        row_series = row_value(row, "series") or current_series
        row_section = row_value(row, "section") or current_section
        wheel_pair_number = parse_excel_int(row_value(row, "wheel_pair"))

        if row_value(row, "measurement_date"):
            current_date = row_date
        if row_value(row, "locomotive"):
            current_loco = row_loco
        if row_value(row, "repair_type"):
            current_repair = row_repair
        if row_value(row, "series"):
            current_series = row_series
        if row_value(row, "section"):
            current_section = row_section

        if not row_date or not row_loco or not row_repair:
            errors.append(f"Строка {row_index}: не заполнены дата, локомотив или вид ремонта.")
            continue
        if wheel_pair_number is None or wheel_pair_number <= 0:
            errors.append(f"Строка {row_index}: не указан корректный номер КП.")
            continue

        values: dict[int, str] = {}
        missing_fields = []
        for column_name, db_col in required_metric_columns.items():
            value = row_value(row, column_name)
            if value == "":
                missing_fields.append(column_name)
            else:
                values[db_col] = value
        for column_name, db_col in optional_metric_columns.items():
            value = row_value(row, column_name)
            if value != "":
                values[db_col] = value
        if missing_fields:
            errors.append(f"Строка {row_index}: не заполнены поля " + ", ".join(missing_fields))
            continue

        group = measurements.setdefault((row_date, row_loco, row_repair), {"series": row_series, "rows": {}, "sections": {}})
        if row_series and not group["series"]:
            group["series"] = row_series
        if row_section:
            group["sections"][wheel_pair_number] = row_section
        group["rows"][wheel_pair_number] = values

    if not measurements:
        return {"error": "Не удалось импортировать ни одного замера.", "errors": errors[:20]}, HTTPStatus.BAD_REQUEST

    imported_measurements = 0
    imported_cells = 0
    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        kp_cache: dict[str, dict[tuple[int, int], str]] = {}
        cur.execute("BEGIN")
        for (measurement_date, locomotive, repair_type), meta in sorted(measurements.items()):
            wheel_pair_numbers = sorted(meta["rows"].keys())
            if not wheel_pair_numbers:
                continue
            year = int(measurement_date[:4]) if len(measurement_date) >= 4 and measurement_date[:4].isdigit() else dt.date.today().year
            series = meta["series"] or series_for_locomotive(cur, locomotive)
            ensure_import_locomotive(cur, series, locomotive, max(wheel_pair_numbers))
            if locomotive not in kp_cache:
                kp_rows = cur.execute("SELECT r, c, v FROM kp_data WHERE locomotive=? AND c IN (2, 3)", (locomotive,)).fetchall()
                kp_cache[locomotive] = {(int(row["r"]), int(row["c"])): text(row["v"]).strip() for row in kp_rows}

            axis_count = locomotive_axis_count(series_for_locomotive(cur, locomotive), locomotive)
            db_rows: list[tuple[int, str, str, str, int, int, str]] = []
            for wheel_pair_number in wheel_pair_numbers:
                row_values = meta["rows"][wheel_pair_number]
                table_row = wheel_pair_number + 1
                section_value = meta["sections"].get(wheel_pair_number, "1" if axis_count == 6 else str(((wheel_pair_number - 1) // 4) + 1))
                kp_index = wheel_pair_number - 1
                bandage_left = row_values.get(8, "")
                bandage_right = row_values.get(9, "")
                diameter_left = row_values.get(10, "")
                diameter_right = row_values.get(11, "")
                if not diameter_left:
                    kp_left = parse_float_value(kp_cache[locomotive].get((kp_index, 2), ""))
                    bandage_left_value = parse_float_value(bandage_left)
                    if kp_left is not None and bandage_left_value is not None:
                        diameter_left = str(int(round(kp_left + bandage_left_value * 2)))
                if not diameter_right:
                    kp_right = parse_float_value(kp_cache[locomotive].get((kp_index, 3), ""))
                    bandage_right_value = parse_float_value(bandage_right)
                    if kp_right is not None and bandage_right_value is not None:
                        diameter_right = str(int(round(kp_right + bandage_right_value * 2)))

                full_values = {
                    0: section_value,
                    1: str(wheel_pair_number),
                    2: row_values.get(2, ""),
                    3: row_values.get(3, ""),
                    4: row_values.get(4, ""),
                    5: row_values.get(5, ""),
                    6: row_values.get(6, ""),
                    7: row_values.get(7, ""),
                    8: bandage_left,
                    9: bandage_right,
                    10: diameter_left,
                    11: diameter_right,
                }
                for col, value in full_values.items():
                    if value != "":
                        db_rows.append((year, measurement_date, locomotive, repair_type, table_row, col, value))
            if not db_rows:
                continue
            cur.execute(
                "DELETE FROM archive_data WHERE y=? AND measurement_date=? AND locomotive=? AND repair_type=?",
                (year, measurement_date, locomotive, repair_type),
            )
            cur.executemany(
                "INSERT OR REPLACE INTO archive_data (y, measurement_date, locomotive, repair_type, r, c, v) VALUES (?, ?, ?, ?, ?, ?, ?)",
                db_rows,
            )
            imported_measurements += 1
            imported_cells += len(db_rows)
        conn.commit()

    return {"ok": True, "imported_measurements": imported_measurements, "imported_cells": imported_cells, "errors": errors[:10]}


def load_archive_rows(locomotive: str = "", search_text: str = "", sort_desc: bool = True) -> list[dict]:
    locomotive = text(locomotive).strip()
    search_text = text(search_text).strip().lower()
    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        order_direction = "DESC" if sort_desc else "ASC"
        params: list[str] = []
        query = """
            SELECT y, measurement_date, locomotive, repair_type, r, c, v
            FROM archive_data
        """
        where: list[str] = []
        if locomotive and locomotive != "Все локомотивы":
            where.append("locomotive = ?")
            params.append(locomotive)
        if where:
            query += " WHERE " + " AND ".join(where)
        query += f" ORDER BY y {order_direction}, measurement_date {order_direction}, locomotive ASC, repair_type ASC, r ASC, c ASC"
        rows = cur.execute(query, params).fetchall()

        grouped: dict[tuple[int, str, str, str, int], list[str]] = {}
        for row in rows:
            r = int(row["r"])
            if r < 2:
                continue
            key = (
                int(row["y"] or 0),
                text(row["measurement_date"]).strip(),
                text(row["locomotive"]).strip(),
                text(row["repair_type"]).strip(),
                r,
            )
            grouped.setdefault(key, [""] * 12)
            c = int(row["c"])
            if 0 <= c < 12:
                grouped[key][c] = text(row["v"])

        kp_cache: dict[tuple[str, str], dict[tuple[int, int], str]] = {}

        def to_float(v):
            try:
                return float(text(v).replace(",", "."))
            except Exception:
                return None

        def fmt_num(v, is_prokat=False):
            if v is None:
                return ""
            try:
                value = float(v)
            except Exception:
                return text(v).strip()
            if value.is_integer():
                if is_prokat:
                    return f"{int(value)},0"
                return str(int(value))
            return str(value).rstrip("0").rstrip(".").replace(".", ",")

        def fmt_raw(v, is_prokat=False):
            if v is None or text(v).strip() == "":
                return ""
            try:
                value = float(text(v).replace(",", "."))
            except Exception:
                return text(v).strip()
            if value.is_integer():
                if is_prokat:
                    return f"{int(value)},0"
                return str(int(value))
            return text(v).strip().replace(".", ",")

        stats_by_section: dict[tuple[int, str, str, str, str], dict[str, str]] = {}
        for (year, measurement_date, loco, repair_type, row_index), values in grouped.items():
            section_value = values[0] or "1"
            key = (year, measurement_date, loco, repair_type, section_value)
            stats = stats_by_section.setdefault(
                key,
                {
                    "max_prokat": "",
                    "min_greben": "",
                    "min_krut": "",
                    "min_bandage_thickness": "",
                    "min_diameter": "",
                    "max_diameter": "",
                    "max_diameter_diff": "",
                    "prokat_6_count": "0",
                    "bandage_limit_count": "0",
                },
            )
            prokat_pair = [v for v in [to_float(values[2]), to_float(values[3])] if v is not None]
            greben_pair = [v for v in [to_float(values[4]), to_float(values[5])] if v is not None]
            krut_pair = [v for v in [to_float(values[6]), to_float(values[7])] if v is not None]
            bandage_pair = [v for v in [to_float(values[8]), to_float(values[9])] if v is not None]
            _, _, resolved_left, resolved_right = resolve_archive_diameter_pair(
                cur,
                loco,
                measurement_date,
                row_index - 2,
                values,
                kp_cache,
            )
            diameter_pair = [v for v in [to_float(resolved_left), to_float(resolved_right)] if v is not None]

            if prokat_pair:
                current = to_float(stats["max_prokat"])
                stats["max_prokat"] = fmt_num(max(prokat_pair), True) if current is None else fmt_num(max([current, *prokat_pair]), True)
                if max(prokat_pair) >= 6:
                    stats["prokat_6_count"] = fmt_num((to_float(stats["prokat_6_count"]) or 0) + 1)
            if greben_pair:
                current = to_float(stats["min_greben"])
                stats["min_greben"] = fmt_num(min(greben_pair)) if current is None else fmt_num(min([current, *greben_pair]))
            if krut_pair:
                current = to_float(stats["min_krut"])
                stats["min_krut"] = fmt_num(min(krut_pair)) if current is None else fmt_num(min([current, *krut_pair]))
            if bandage_pair:
                current = to_float(stats["min_bandage_thickness"])
                stats["min_bandage_thickness"] = fmt_num(min(bandage_pair)) if current is None else fmt_num(min([current, *bandage_pair]))
            if diameter_pair:
                current_min = to_float(stats["min_diameter"])
                current_max = to_float(stats["max_diameter"])
                pair_min = min(diameter_pair)
                pair_max = max(diameter_pair)
                stats["min_diameter"] = fmt_num(pair_min) if current_min is None else fmt_num(min([current_min, *diameter_pair]))
                stats["max_diameter"] = fmt_num(pair_max) if current_max is None else fmt_num(max([current_max, *diameter_pair]))
                min_diameter = to_float(stats["min_diameter"])
                max_diameter = to_float(stats["max_diameter"])
                if min_diameter is not None and max_diameter is not None:
                    stats["max_diameter_diff"] = fmt_num(max_diameter - min_diameter)
            if bandage_pair:
                check_text = normalize_text(loco)
                is_limit = False
                if any(x in check_text for x in ["пэ-2м", "пэ2м", "пэ 2м", "pe-2m", "pe2m"]):
                    is_limit = any(v < 51 for v in bandage_pair)
                elif "тэм" in check_text or "tem" in check_text:
                    is_limit = any(v < 41 for v in bandage_pair)
                if is_limit:
                    stats["bandage_limit_count"] = fmt_num((to_float(stats["bandage_limit_count"]) or 0) + 1)

        def sort_key(item):
            (year, measurement_date, loco, repair_type, row_index), _ = item
            try:
                year_key = int(year)
            except Exception:
                year_key = 0
            try:
                date_key = int((measurement_date or "").replace("-", ""))
            except Exception:
                date_key = 0
            if sort_desc:
                return (-year_key, -date_key, loco, repair_type, row_index)
            return (year_key, date_key, loco, repair_type, row_index)

        archive_rows: list[dict] = []
        for (year, measurement_date, loco, repair_type, row_index), values in sorted(grouped.items(), key=sort_key):
            if search_text:
                probe = " ".join(
                    [
                        text(year),
                        measurement_date,
                        loco,
                        repair_type,
                        *[text(v) for v in values],
                    ]
                ).lower()
                if search_text not in probe:
                    continue

            formatted_date = measurement_date
            if measurement_date and len(measurement_date) >= 10:
                parts = measurement_date.split("-")
                if len(parts) == 3:
                    formatted_date = f"{parts[2]}.{parts[1]}.{parts[0][-2:]}"

            date_and_repair = f"{loco}\n{formatted_date}"
            if repair_type:
                date_and_repair += f"\n{normalize_repair_type(repair_type)}"

            section_value = values[0] or "1"
            stats = stats_by_section.get((year, measurement_date, loco, normalize_repair_type(repair_type), section_value), {})
            _, _, resolved_left, resolved_right = resolve_archive_diameter_pair(
                cur,
                loco,
                measurement_date,
                row_index - 2,
                values,
                kp_cache,
            )
            row_values = [
                date_and_repair,
                section_value,
                stats.get("max_prokat", ""),
                stats.get("min_greben", ""),
                stats.get("min_krut", ""),
                stats.get("min_bandage_thickness", ""),
                stats.get("max_diameter_diff", ""),
                stats.get("bandage_limit_count", "0"),
                stats.get("prokat_6_count", "0"),
                values[1],
                fmt_raw(values[2], is_prokat=True),
                fmt_raw(values[3], is_prokat=True),
                fmt_raw(values[4]),
                fmt_raw(values[5]),
                fmt_raw(values[6]),
                fmt_raw(values[7]),
                fmt_raw(values[8]),
                fmt_raw(values[9]),
                fmt_raw(resolved_left),
                fmt_raw(resolved_right),
            ]
            archive_rows.append({
                "values": row_values,
                "year": year,
                "measurement_date": measurement_date,
                "locomotive": loco,
                "repair_type": normalize_repair_type(repair_type),
                "source_r": row_index,
                "section": section_value,
                "stats": stats,
            })

        return archive_rows


def format_trend_number(value: float | int | None) -> str:
    if value is None:
        return ""
    try:
        number = float(value)
    except Exception:
        return text(value).strip()
    if abs(number - round(number)) < 1e-9:
        return str(int(round(number)))
    return str(round(number, 2)).rstrip("0").rstrip(".").replace(".", ",")


def format_trend_date(value: str) -> str:
    value = text(value).strip()
    if len(value) >= 10 and value[4:5] == "-" and value[7:8] == "-":
        parts = value.split("-")
        if len(parts) == 3:
            return f"{parts[2]}.{parts[1]}.{parts[0]}"
    return value


def _wear_trend_compare(metric_key: str, latest: float | None, previous: float | None) -> tuple[str, float | None]:
    if latest is None or previous is None:
        return "none", None
    delta = latest - previous
    if abs(delta) < 1e-9:
        return "stable", delta
    metric = next((item for item in WEAR_TREND_METRICS if item["key"] == metric_key), None)
    worse_when = text(metric.get("worse_when") if metric else "").strip().lower()
    is_worse = (delta > 0 and worse_when == "higher") or (delta < 0 and worse_when == "lower")
    is_better = (delta < 0 and worse_when == "higher") or (delta > 0 and worse_when == "lower")
    if is_worse:
        return "worse", delta
    if is_better:
        return "better", delta
    return "stable", delta


def _wear_pair_number_from_row(values: list[str], source_row: int) -> int:
    raw = text(values[1] if len(values) > 1 else "").strip()
    try:
        pair_number = int(raw)
        if pair_number > 0:
            return pair_number
    except Exception:
        pass
    if source_row >= 2:
        return max(1, source_row - 1)
    return 1


def _wear_session_metrics(values: list[str]) -> dict[str, dict[str, float | None] | float | None]:
    def value_at(col: int) -> float | None:
        if col >= len(values):
            return None
        return parse_float_value(values[col])

    return {
        "prokat": {"left": value_at(2), "right": value_at(3)},
        "greben": {"left": value_at(4), "right": value_at(5)},
        "krut": {"left": value_at(6), "right": value_at(7)},
        "bandage_thickness": {"left": value_at(8), "right": value_at(9)},
        "diameter": {"left": value_at(10), "right": value_at(11)},
    }


def load_wear_analysis_rows(locomotive: str = "", date_from: str = "", date_to: str = "") -> dict:
    date_from = text(date_from).strip()
    date_to = text(date_to).strip()
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    with DB_LOCK, connect() as conn:
        cur = conn.cursor()
        locomotives = load_locomotives(cur)
        selected = text(locomotive).strip()
        if not selected and locomotives:
            selected = text(locomotives[0].get("number")).strip()

        locomotive_record = next((item for item in locomotives if text(item.get("number")).strip() == selected), None)
        series = series_for_locomotive(cur, selected)
        axis_count = locomotive_axis_count(series, selected)
        wheel_pair_count = int(locomotive_record.get("wheelPairCount") or 0) if locomotive_record else 0
        if wheel_pair_count <= 0:
            wheel_pair_count = axis_count or 12
        wheel_pair_count = max(1, wheel_pair_count)

        query = [
            "SELECT y, measurement_date, locomotive, repair_type, r, c, v",
            "FROM archive_data",
            "WHERE locomotive=? AND TRIM(COALESCE(measurement_date, '')) <> ''",
        ]
        params: list[str] = [selected]
        if date_from:
            query.append("AND measurement_date >= ?")
            params.append(date_from)
        if date_to:
            query.append("AND measurement_date <= ?")
            params.append(date_to)
        query.append("ORDER BY measurement_date ASC, y ASC, repair_type ASC, r ASC, c ASC")
        rows = cur.execute("\n".join(query), params).fetchall()

        sessions: dict[tuple[str, str, int], dict[str, object]] = {}
        for row in rows:
            measurement_date = text(row["measurement_date"]).strip()
            repair_type = normalize_repair_type(row["repair_type"])
            source_row = int(row["r"] or 0)
            key = (measurement_date, repair_type, source_row)
            session = sessions.setdefault(
                key,
                {
                    "year": int(row["y"] or 0),
                    "measurement_date": measurement_date,
                    "repair_type": repair_type,
                    "source_row": source_row,
                    "values": [""] * 12,
                },
            )
            values = session["values"]
            if isinstance(values, list):
                col = int(row["c"] or 0)
                if 0 <= col < len(values):
                    values[col] = text(row["v"]).strip()

        sessions_by_pair: dict[int, list[dict[str, object]]] = {pair: [] for pair in range(1, wheel_pair_count + 1)}
        for session in sessions.values():
            values = session.get("values")
            if not isinstance(values, list):
                continue
            pair_number = _wear_pair_number_from_row(values, int(session.get("source_row") or 0))
            pair_number = max(1, min(wheel_pair_count, pair_number))
            metrics = _wear_session_metrics(values)
            sessions_by_pair.setdefault(pair_number, []).append(
                {
                    "year": int(session.get("year") or 0),
                    "measurement_date": text(session.get("measurement_date")).strip(),
                    "repair_type": text(session.get("repair_type")).strip(),
                    "metrics": metrics,
                }
            )

        chart_pairs: list[dict[str, object]] = []
        for pair_number in range(1, wheel_pair_count + 1):
            points = sessions_by_pair.get(pair_number, [])
            points.sort(
                key=lambda item: (
                    text(item.get("measurement_date")).strip(),
                    int(item.get("year") or 0),
                    text(item.get("repair_type")).strip(),
                )
            )
            chart_pairs.append(
                {
                    "wheel_pair": pair_number,
                    "points": points,
                }
            )

        result_rows: list[dict] = []
        for pair_number in range(1, wheel_pair_count + 1):
            pair_sessions = sessions_by_pair.get(pair_number, [])
            pair_sessions.sort(
                key=lambda item: (
                    text(item.get("measurement_date")).strip(),
                    int(item.get("year") or 0),
                    text(item.get("repair_type")).strip(),
                )
            )
            metric_payload: dict[str, dict[str, object]] = {}
            worse_count = 0
            better_count = 0
            stable_count = 0
            total_compared = 0

            for metric in WEAR_TREND_METRICS:
                key = metric["key"]
                side_payload: dict[str, dict[str, object]] = {}
                for side in ("left", "right"):
                    history: list[float] = []
                    for session in pair_sessions:
                        metrics = session.get("metrics")
                        if isinstance(metrics, dict):
                            metric_value = metrics.get(key)
                            if isinstance(metric_value, dict):
                                value = metric_value.get(side)
                                if value is not None:
                                    history.append(value)
                    if len(history) >= 2:
                        first_value = history[0]
                        latest = history[-1]
                        trend, delta = _wear_trend_compare(key, latest, first_value)
                        total_compared += 1
                        if trend == "worse":
                            worse_count += 1
                        elif trend == "better":
                            better_count += 1
                        elif trend == "stable":
                            stable_count += 1
                    else:
                        first_value = history[0] if history else None
                        latest = first_value
                        trend, delta = "none", None
                    side_payload[side] = {
                        "latest": latest,
                        "previous": first_value,
                        "delta": delta,
                        "trend": trend,
                    }
                metric_payload[key] = {
                    "label": metric["label"],
                    "left": side_payload["left"],
                    "right": side_payload["right"],
                }

            if total_compared <= 0:
                status_key = "none"
                status_label = "Недостаточно данных"
            elif worse_count > better_count:
                status_key = "worse"
                status_label = "Износ растет"
            elif better_count > worse_count:
                status_key = "better"
                status_label = "Износ снижается"
            elif stable_count > 0 and worse_count == 0 and better_count == 0:
                status_key = "stable"
                status_label = "Стабильно"
            else:
                status_key = "mixed"
                status_label = "Смешанный тренд"

            last_session = pair_sessions[-1] if pair_sessions else None
            result_rows.append(
                {
                    "wheel_pair": pair_number,
                    "session_count": len(pair_sessions),
                    "last_measurement_date": text(last_session.get("measurement_date") if last_session else "").strip(),
                    "last_repair_type": text(last_session.get("repair_type") if last_session else "").strip(),
                    "status_key": status_key,
                    "status_label": status_label,
                    "metrics": metric_payload,
                }
            )

        return {
            "locomotive": selected,
            "series": series,
            "wheel_pair_count": wheel_pair_count,
            "date_from": date_from,
            "date_to": date_to,
            "rows": result_rows,
            "chart": {
                "metrics": [
                    {
                        "key": metric["key"],
                        "label": metric["label"],
                    }
                    for metric in WEAR_TREND_METRICS
                ],
                "pairs": chart_pairs,
            },
        }


def render_page(role: str, template_name: str = "index.html", extra: dict[str, str] | None = None) -> str:
    with DB_LOCK, connect() as conn:
        loco_choices = load_locomotives(conn.cursor())
    with open(ROOT / "templates" / template_name, "r", encoding="utf-8") as f:
        html = f.read()
    replacements = {
        "{{APP_PREFIX}}": APP_PREFIX,
        "{{APP_VERSION}}": APP_VERSION,
        "{{CAN_EDIT}}": "true" if role in ("edit", "editor", "admin") else "false",
        "{{TAB_INPUT_STYLE}}": "" if role in ("edit", "editor", "admin") else 'style="display:none"',
        "{{LOCOMOTIVE_CHOICES}}": json.dumps(loco_choices, ensure_ascii=False),
    }
    if extra:
        replacements.update(extra)
    return (
        html
        .replace("{{APP_PREFIX}}", replacements["{{APP_PREFIX}}"])
        .replace("{{APP_VERSION}}", replacements["{{APP_VERSION}}"])
        .replace("{{CAN_EDIT}}", replacements["{{CAN_EDIT}}"])
        .replace("{{TAB_INPUT_STYLE}}", replacements["{{TAB_INPUT_STYLE}}"])
        .replace("{{LOCOMOTIVE_CHOICES}}", replacements["{{LOCOMOTIVE_CHOICES}}"])
        .replace("{{WEAR_LOCOMOTIVE}}", replacements.get("{{WEAR_LOCOMOTIVE}}", ""))
        .replace("{{WEAR_DATE_FROM}}", replacements.get("{{WEAR_DATE_FROM}}", ""))
        .replace("{{WEAR_DATE_TO}}", replacements.get("{{WEAR_DATE_TO}}", ""))
        .replace("{{WEAR_CHART_MODE}}", replacements.get("{{WEAR_CHART_MODE}}", "pair"))
    )



from fastapi import FastAPI, Request, Response, Depends, Form, HTTPException, Cookie, status
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import uvicorn
import urllib.parse
import traceback

app = FastAPI(title="RTPS Zamer KP")

@app.middleware("http")
async def strip_prefix(request: Request, call_next):
    if request.scope["path"].startswith(APP_PREFIX + "/"):
        request.scope["path"] = request.scope["path"][len(APP_PREFIX):]
    return await call_next(request)

app.mount("/static", StaticFiles(directory=str(ROOT / "static")), name="static")

def _cookie_value_fastapi(username: str, role: str) -> str:
    payload = f"{username}:{role}:{int(dt.datetime.now().timestamp())}"
    signature = hmac.new(WEB_SECRET.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    return f"{payload}:{signature}"

def _verify_cookie_fastapi(value: str) -> tuple[str, str, str, str] | None:
    for sep in (":", "|"):
        try:
            parts = value.rsplit(sep, 5)
            if len(parts) == 6:
                user_id, role, modules, safe_name, expiry_text, sig = parts
                payload = f"{user_id}{sep}{role}{sep}{modules}{sep}{safe_name}{sep}{expiry_text}"
            elif len(parts) == 5:
                user_id, role, modules, safe_name, sig = parts
                payload = f"{user_id}{sep}{role}{sep}{modules}{sep}{safe_name}"
                expiry_text = "2000000000"
            else:
                continue
                
            secrets_to_try = [WEB_SECRET]
                
            matched = False
            for secret in secrets_to_try:
                expected = hmac.new(secret.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
                if hmac.compare_digest(expected, sig):
                    matched = True
                    break
                    
            if not matched:
                continue
            if float(expiry_text) < dt.datetime.now().timestamp():
                return None
                
            return urllib.parse.unquote(user_id), urllib.parse.unquote(role), urllib.parse.unquote(modules), urllib.parse.unquote(safe_name)
        except Exception:
            continue
    return None

def get_current_session_fastapi(request: Request):
    cookie = request.cookies.get(SESSION_COOKIE)
    if cookie:
        return _verify_cookie_fastapi(cookie)
    return None

def get_mod_role_fastapi(session: tuple[str, str, str, str] | None, module: str) -> str:
    if not session:
        return ""
    username = session[0]
    role = session[1]
    modules = session[2]

    resolved = resolve_user_access(ROOT.parent / "base" / "web_users.db", username, role, modules)
    if not resolved:
        return ""
    role, modules = resolved
    return module_role(role, modules, module) or ""

def require_auth_fastapi(request: Request, need_edit: bool = False):
    session = get_current_session_fastapi(request)
    role = get_mod_role_fastapi(session, "zamer_kp")
    if not role:
        return False, None
    if need_edit and role not in ("edit", "editor", "admin"):
        return False, None
    return True, session

def json_response(data: dict | list, status_code: int = 200) -> JSONResponse:
    return JSONResponse(content=data, status_code=status_code)

@app.get("/zamer-kp", response_class=HTMLResponse)
@app.get("/", response_class=HTMLResponse)
async def home_route(request: Request):
    session = get_current_session_fastapi(request)
    mod_role = get_mod_role_fastapi(session, "zamer_kp")
    
    if not session or not mod_role:
        return RedirectResponse(f"{MAIN_LOGIN_URL}?next=/zamer-kp", status_code=303)
        
    html_content = render_page(mod_role)
    response = HTMLResponse(content=html_content)
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response

@app.get("/wear-charts", response_class=HTMLResponse)
async def wear_charts_route(request: Request, locomotive: str = "", date_from: str = "", date_to: str = "", mode: str = "pair"):
    session = get_current_session_fastapi(request)
    mod_role = get_mod_role_fastapi(session, "zamer_kp")

    if not session or not mod_role:
        return RedirectResponse(f"{MAIN_LOGIN_URL}?next=/zamer-kp/wear-charts", status_code=303)

    extra = {
        "{{WEAR_LOCOMOTIVE}}": html.escape(text(locomotive), quote=True),
        "{{WEAR_DATE_FROM}}": html.escape(text(date_from), quote=True),
        "{{WEAR_DATE_TO}}": html.escape(text(date_to), quote=True),
        "{{WEAR_CHART_MODE}}": "all" if text(mode).strip().lower() == "all" else "pair",
    }
    html_content = render_page(mod_role, "wear_charts.html", extra)
    response = HTMLResponse(content=html_content)
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response

@app.get("/logout")
async def logout_route():
    resp = RedirectResponse(MAIN_LOGIN_URL, status_code=303)
    resp.delete_cookie(SESSION_COOKIE, path="/", httponly=True, samesite="lax")
    return resp

@app.get("/api/state")
async def get_state(request: Request, locomotive: str = ""):
    auth_ok, session = require_auth_fastapi(request)
    if not auth_ok:
        return Response("Unauthorized", status_code=401)
    return json_response(load_state(locomotive.strip()))

@app.get("/api/archive")
async def get_archive(request: Request, locomotive: str = "", search: str = "", sort: str = "desc"):
    auth_ok, session = require_auth_fastapi(request)
    if not auth_ok:
        return Response("Unauthorized", status_code=401)
    rows = load_archive_rows(locomotive.strip(), search.strip(), sort.strip().lower() != "asc")
    return json_response({"rows": rows})

@app.get("/api/kp-data")
async def get_kp_data(request: Request, locomotive: str = "", version_id: int | None = None):
    auth_ok, session = require_auth_fastapi(request)
    if not auth_ok:
        return Response("Unauthorized", status_code=401)
    return json_response(load_kp_view(locomotive.strip(), version_id))

@app.get("/api/wear-analysis")
async def get_wear_analysis(request: Request, locomotive: str = "", date_from: str = "", date_to: str = ""):
    auth_ok, session = require_auth_fastapi(request)
    if not auth_ok:
        return Response("Unauthorized", status_code=401)
    return json_response(load_wear_analysis_rows(locomotive.strip(), date_from.strip(), date_to.strip()))

@app.get("/api/norms")
async def get_norms(request: Request):
    auth_ok, session = require_auth_fastapi(request)
    if not auth_ok:
        return Response("Unauthorized", status_code=401)
    return json_response({"rows": load_norms_rows()})

@app.get("/api/archive-excel-template")
async def export_archive_template(request: Request):
    auth_ok, session = require_auth_fastapi(request)
    if not auth_ok:
        return Response("Unauthorized", status_code=401)
    try:
        data = archive_excel_template_bytes()
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename*=UTF-8''%D0%A8%D0%B0%D0%B1%D0%BB%D0%BE%D0%BD_%D0%B8%D0%BC%D0%BF%D0%BE%D1%80%D1%82%D0%B0_%D0%B0%D1%80%D1%85%D0%B8%D0%B2%D0%B0.xlsx"}
        )
    except Exception as exc:
        return json_response({"error": str(exc)}, status_code=400)

@app.get("/api/phone-export")
async def export_phone(request: Request, kind: str = "archive", date_from: str = "", date_to: str = ""):
    try:
        query_params = request.query_params
        selected_locomotives = [item.strip() for item in query_params.getlist("locomotive") if item.strip()]
        payload = phone_export_payload(kind.strip().lower(), selected_locomotives, date_from.strip(), date_to.strip())
        return json_response(payload)
    except Exception as exc:
        return json_response({"error": str(exc)}, status_code=400)

@app.get("/api/archive-excel-export")
async def export_archive_excel(request: Request, date_from: str = "", date_to: str = ""):
    auth_ok, session = require_auth_fastapi(request)
    if not auth_ok:
        return Response("Unauthorized", status_code=401)
    try:
        query_params = request.query_params
        selected_locomotives = [item.strip() for item in query_params.getlist("locomotive") if item.strip()]
        data, row_count = archive_excel_export_bytes(selected_locomotives, date_from.strip(), date_to.strip())
        if row_count <= 0:
            return json_response({"error": "По выбранным фильтрам данных нет."}, status_code=400)
        filename = f"Экспорт_архива_{dt.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        from urllib.parse import quote
        safe_filename = quote(filename)
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{safe_filename}"}
        )
    except Exception as exc:
        return json_response({"error": str(exc)}, status_code=400)

@app.post("/api/state")
async def post_state(request: Request):
    auth_ok, session = require_auth_fastapi(request, need_edit=True)
    if not auth_ok:
        return json_response({"error": "Unauthorized"}, status_code=401)
    try:
        payload = await request.json()
        saved = save_state(payload, full_name=session[3] if session and len(session) > 3 else "")
        return json_response(saved)
    except Exception as exc:
        return json_response({"error": str(exc)}, status_code=400)

@app.post("/api/archive")
async def post_archive(request: Request):
    auth_ok, session = require_auth_fastapi(request, need_edit=True)
    if not auth_ok:
        return json_response({"error": "Unauthorized"}, status_code=401)
    try:
        payload = await request.json()
        if payload.get("action") == "delete":
            result = delete_archive_measurement(payload)
        elif payload.get("changes"):
            result = update_archive_cells(payload)
        else:
            result = save_archive(payload)
            
        if isinstance(result, tuple):
            return json_response(result[0], status_code=result[1])
        return json_response(result)
    except Exception as exc:
        return json_response({"error": str(exc)}, status_code=400)

@app.post("/api/kp-data")
async def post_kp_data(request: Request):
    auth_ok, session = require_auth_fastapi(request, need_edit=True)
    if not auth_ok:
        return json_response({"error": "Unauthorized"}, status_code=401)
    try:
        payload = await request.json()
        action = text(payload.get("action")).strip()
        if action == "new_version":
            result = create_kp_data_version(payload)
        elif action == "update_version":
            result = update_kp_version(payload)
        else:
            result = save_kp_data(payload)
        if isinstance(result, tuple):
            return json_response(result[0], status_code=result[1])
        return json_response(result)
    except Exception as exc:
        return json_response({"error": str(exc)}, status_code=400)

@app.post("/api/norms")
async def post_norms(request: Request):
    auth_ok, session = require_auth_fastapi(request, need_edit=True)
    if not auth_ok:
        return json_response({"error": "Unauthorized"}, status_code=401)
    try:
        payload = await request.json()
        result = save_norms_rows(payload)
        if isinstance(result, tuple):
            return json_response(result[0], status_code=result[1])
        return json_response(result)
    except Exception as exc:
        return json_response({"error": str(exc)}, status_code=400)

@app.post("/api/archive-excel-import")
async def post_archive_import(request: Request):
    auth_ok, session = require_auth_fastapi(request, need_edit=True)
    if not auth_ok:
        return json_response({"error": "Unauthorized"}, status_code=401)
    try:
        payload = await request.json()
        encoded = text(payload.get("data")).strip()
        if not encoded:
            return json_response({"error": "Файл Excel не передан."}, status_code=400)
        import base64
        data = base64.b64decode(encoded)
        result = import_archive_excel_bytes(data)
        if isinstance(result, tuple):
            return json_response(result[0], status_code=result[1])
        return json_response(result)
    except Exception as exc:
        return json_response({"error": str(exc)}, status_code=400)

@app.post("/api/phone-import")
async def post_phone_import(request: Request):
    try:
        raw = await request.body()
        payload = parse_phone_json_payload(raw)
        result = import_phone_payload(payload)
        if isinstance(result, tuple):
            return json_response(result[0], status_code=result[1])
        return json_response(result)
    except Exception as exc:
        return json_response({"error": str(exc)}, status_code=400)

def main() -> None:
    ensure_db()
    host = os.environ.get("WEB_HOST", "127.0.0.1")
    port = int(os.environ.get("WEB_PORT", "8003"))
    url = f"http://{host}:{port}{APP_PREFIX}"
    print(f"Замер КП ready (FastAPI): {url}")
    if host in {"127.0.0.1", "localhost", "0.0.0.0"}:
        import threading, webbrowser
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    uvicorn.run("app:app", host=host, port=port, reload=False)

if __name__ == "__main__":
    main()
