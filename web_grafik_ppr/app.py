from __future__ import annotations

import calendar
import datetime as dt
import hashlib
import json
import os
from io import BytesIO
import shutil
import hmac
import secrets
import sqlite3
import threading
import webbrowser
import sys
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from threading import Lock, RLock
from urllib.parse import parse_qs, quote, urlparse

APP_VERSION = "web-gpp-1.7"
MONTHS_RU = [
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]
TEM_NORM_ROWS = ["ТО2", "ТО3", "ТР1", "ТР2", "ТР3", "СР", "КР"]
AGR_NORM_ROWS = ["ТО", "ТР", "КР"]
REPAIR_SCHEDULE_COLUMN_CODES = [
    "ТР1", "ТР2", "ТР1", "ТР3", "ТР1", "ТР2", "ТР1", "СР",
    "ТР1", "ТР2", "ТР1", "ТР3", "ТР1", "ТР2", "ТР1", "КР",
]
REPORT_TEMPLATE_NAME = "Отчет_шаблон.xlsx"
TU28_TEMPLATE_NAME = "ТУ-28_шаблон.xlsx"
MONTH_DAY_LIMIT_FOR_REPORT = 25
TEP_REPORT_FACTORS = {"ТО2": 1, "ТО3": 2, "ТР1": 5, "ТР2": 10, "ТР3": 15}
AGR_REPORT_FACTORS = {"ТО": 1, "ТР": 5}
TEP_HOUR_FACTORS = {"ТО2": 1, "ТО3": 2, "ТР1": 5, "ТР2": 10, "ТР3": 15}
AGR_HOUR_FACTORS = {"ТО": 1, "ТР": 5}
TU28_REPAIR_CODES = {"ТО3", "ТР1", "ТР2", "ТР3", "СР", "КР"}
FIXED_HOLIDAYS = {
    (1, 1), (1, 2), (1, 3), (1, 4), (1, 5), (1, 6), (1, 7), (1, 8),
    (2, 23), (3, 8), (5, 1), (5, 9), (6, 12), (11, 4),
}
TRANSFER_HOLIDAYS_BY_YEAR = {
    2025: {(5, 2), (5, 8), (6, 13), (11, 3), (12, 31)},
}

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT.parent))
from rtps_common import connect_sqlite, module_role, resolve_user_access

DATA_DIR = ROOT / "data"
DB_FILE = DATA_DIR / "grafik_ppr_web.db"
SHARED_DATA_DIR = ROOT.parent / "data"
AUTH_FILE = SHARED_DATA_DIR / "web_auth.json"
WEB_SECRET_FILE = SHARED_DATA_DIR / "web_secret.txt"
SOURCE_DB = ROOT.parent / "base" / "common_database.db"
SOURCE_DIR = ROOT.parent / "src" / "График ППР"
ACT_TEMPLATE_NAME = "Акт_шаблон.xlsx"

DB_LOCK = RLock()
SERVER_STARTED_AT = None
SESSION_COOKIE = "rtps_session"
SESSION_TTL_SECONDS = 7 * 24 * 60 * 60


def format_n(value) -> str:
    try:
        number = float(str(value).replace(",", "."))
    except Exception:
        return "0,00"
    if number.is_integer():
        return f"{int(number)},00"
    return f"{number:.2f}".replace(".", ",")


def load_web_secret() -> str:
    return "opYbo6NB8pb7dChYQkmHEvUH6K4hAHjuzi2qEYOC024"


WEB_SECRET = load_web_secret()
SESSIONS: dict[str, tuple[str, str, str, str, float]] = {}


def load_auth_config() -> tuple[str, str, str]:
    user = os.environ.get("WEB_USER", "admin").strip() or "admin"
    view_password = os.environ.get("WEB_VIEW_PASSWORD", "").strip()
    edit_password = (
        os.environ.get("WEB_EDIT_PASSWORD", "").strip()
        or os.environ.get("WEB_PASSWORD", "").strip()
    )
    if view_password and edit_password:
        return user, view_password, edit_password
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if AUTH_FILE.exists():
        try:
            payload = json.loads(AUTH_FILE.read_text(encoding="utf-8"))
            file_user = str(payload.get("user", user)).strip() or user
            legacy_password = str(payload.get("password", "")).strip()
            file_view = str(payload.get("view_password", legacy_password)).strip()
            file_edit = str(payload.get("edit_password", "")).strip()
            if not file_edit and legacy_password and not payload.get("view_password"):
                file_edit = secrets.token_urlsafe(8)
            if not file_edit:
                file_edit = legacy_password
            if file_view and not file_edit:
                file_edit = secrets.token_urlsafe(8)
            if file_edit and not file_view:
                file_view = secrets.token_urlsafe(8)
            if file_view and file_edit and file_view == file_edit:
                file_edit = secrets.token_urlsafe(8)
            if file_view and file_edit:
                AUTH_FILE.write_text(
                    json.dumps(
                        {"user": file_user, "view_password": file_view, "edit_password": file_edit},
                        ensure_ascii=False,
                        indent=2,
                    ),
                    encoding="utf-8",
                )
                return file_user, file_view, file_edit
        except Exception:
            pass
    if not view_password:
        view_password = secrets.token_urlsafe(8)
    if not edit_password:
        edit_password = secrets.token_urlsafe(8)
    AUTH_FILE.write_text(
        json.dumps(
            {"user": user, "view_password": view_password, "edit_password": edit_password},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"Web auth created: user={user} view_password={view_password} edit_password={edit_password}")
    return user, view_password, edit_password


WEB_USER, WEB_VIEW_PASSWORD, WEB_EDIT_PASSWORD = load_auth_config()
AUTH_ENABLED = True
APP_PREFIX = "/grafik-ppr"


def ensure_database() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if not DB_FILE.exists() and SOURCE_DB.exists():
        shutil.copy2(SOURCE_DB, DB_FILE)
    with sqlite3.connect(DB_FILE) as conn:
        ensure_schema(conn.cursor())
        conn.commit()


def conn() -> sqlite3.Connection:
    c = connect_sqlite(DB_FILE)
    ensure_schema(c.cursor())
    return c


def ensure_schema(cur: sqlite3.Cursor) -> None:
    cur.execute("CREATE TABLE IF NOT EXISTS repairs (y INT, m TEXT, t TEXT, r INT, c INT, v TEXT, PRIMARY KEY(y,m,t,r,c))")
    cur.execute("CREATE TABLE IF NOT EXISTS norms (y INT, cat TEXT, k TEXT, v TEXT, PRIMARY KEY(y,cat,k))")
    cur.execute("CREATE TABLE IF NOT EXISTS inventory (y INT, ser TEXT, num TEXT, inv TEXT, PRIMARY KEY(y,ser,num))")
    cur.execute("CREATE TABLE IF NOT EXISTS repair_settings (k TEXT PRIMARY KEY, v TEXT)")
    cur.execute("CREATE TABLE IF NOT EXISTS acts_state (y INT, m TEXT, act_num TEXT, is_done INT, sap_order_done INT DEFAULT 0, PRIMARY KEY(y, m, act_num))")
    cur.execute("CREATE TABLE IF NOT EXISTS report_notes (y INT, m TEXT, k TEXT, v TEXT, PRIMARY KEY(y,m,k))")
    cur.execute("CREATE TABLE IF NOT EXISTS tu28_data (y INT, m TEXT, r INT, k TEXT, v TEXT, PRIMARY KEY(y,m,r,k))")
    cur.execute("CREATE TABLE IF NOT EXISTS repair_schedule (y INT, r INT, k TEXT, v TEXT, PRIMARY KEY(y,r,k))")


def default_repair_schedule_state() -> dict:
    columns = [{"code": code} for code in REPAIR_SCHEDULE_COLUMN_CODES]
    return {
        "columns": columns,
        "periodicity": {
            "series": ["ТЭМ-2УМ", "ТЭМ-2", ""],
            "values": [["", "", "", "", ""] for _ in range(3)],
        },
        "objects": [
            {
                "series": "",
                "number": "",
                "kr": {"plan": "", "fact": ""},
                "plan": ["" for _ in columns],
                "fact": ["" for _ in columns],
            }
        ],
    }


def _repair_schedule_parse_date(value) -> dt.date | None:
    text = s(value).strip()
    if not text:
        return None
    try:
        day_s, month_s, year_s = text.split(".")
        day = int(day_s)
        month = int(month_s)
        year = int(year_s)
        result = dt.date(year, month, day)
    except Exception:
        return None
    return result


def _repair_schedule_format_date(value: dt.date | None) -> str:
    if not isinstance(value, dt.date):
        return ""
    return value.strftime("%d.%m.%Y")


def _repair_schedule_add_months(value: dt.date | None, months: float) -> dt.date | None:
    if not isinstance(value, dt.date):
        return None
    try:
        months_total = float(months)
    except Exception:
        return None
    if months_total <= 0:
        return value
    whole = int(months_total)
    fraction = months_total - whole
    year = value.year
    month = value.month + whole
    while month > 12:
        year += 1
        month -= 12
    while month < 1:
        year -= 1
        month += 12
    day = min(value.day, calendar.monthrange(year, month)[1])
    result = dt.date(year, month, day)
    if fraction > 0:
        result += dt.timedelta(days=round(30 * fraction))
    return result


def _repair_schedule_period_row(periodicity: dict, series_name: str) -> list:
    series_list = periodicity.get("series", []) if isinstance(periodicity, dict) else []
    values_list = periodicity.get("values", []) if isinstance(periodicity, dict) else []
    target = s(series_name).strip().upper()
    row_index = 0
    for idx, item in enumerate(series_list):
        if s(item).strip().upper() == target:
            row_index = idx
            break
    if not isinstance(values_list, list) or row_index >= len(values_list):
        return []
    row = values_list[row_index]
    return row if isinstance(row, list) else []


def _repair_schedule_period_months(periodicity: dict, code: str, series_name: str) -> float:
    row = _repair_schedule_period_row(periodicity, series_name)
    index_map = {"ТР1": 0, "ТР2": 1, "ТР3": 2, "СР": 3, "КР": 4}
    idx = index_map.get(normalize_repair_code(code))
    if idx is None or idx >= len(row):
        return 0.0
    try:
        value = float(str(row[idx]).replace(",", ".").strip())
    except Exception:
        return 0.0
    return value if value > 0 else 0.0


def _repair_schedule_code_factor(code: str) -> float:
    normalized = normalize_repair_code(code)
    if normalized == "ТР1":
        return 1.0
    if normalized == "ТР2":
        return 0.5
    if normalized == "ТР3":
        return 0.25
    if normalized == "СР":
        return 0.125
    if normalized == "КР":
        return 1.0
    return 0.0


def compute_repair_schedule_derived(schedule: dict) -> dict:
    if not isinstance(schedule, dict):
        return schedule
    columns = schedule.get("columns", []) if isinstance(schedule.get("columns", []), list) else []
    periodicity = schedule.get("periodicity", {}) if isinstance(schedule.get("periodicity", {}), dict) else {}
    objects = schedule.get("objects", []) if isinstance(schedule.get("objects", []), list) else []
    for row in objects:
        if not isinstance(row, dict):
            continue
        if not isinstance(row.get("kr"), dict):
            row["kr"] = {"plan": "", "fact": ""}
        fact_date = _repair_schedule_parse_date(row["kr"].get("fact"))
        plan_date = fact_date or _repair_schedule_parse_date(row["kr"].get("plan"))
        row["kr"]["plan"] = _repair_schedule_format_date(plan_date)
        source_date = fact_date or plan_date
        plan_list = row.get("plan", []) if isinstance(row.get("plan", []), list) else []
        fact_list = row.get("fact", []) if isinstance(row.get("fact", []), list) else []
        while len(plan_list) < len(columns):
            plan_list.append("")
        while len(fact_list) < len(columns):
            fact_list.append("")
        for cidx, col in enumerate(columns):
            code = s((col or {}).get("code"))
            period_months = _repair_schedule_period_months(periodicity, code, row.get("series", ""))
            target_date = None
            if source_date and period_months > 0:
                target_date = _repair_schedule_add_months(source_date, period_months * _repair_schedule_code_factor(code))
            planned = _repair_schedule_format_date(target_date)
            plan_list[cidx] = planned
            fact_date_col = _repair_schedule_parse_date(fact_list[cidx])
            source_date = fact_date_col or target_date or source_date
        row["plan"] = plan_list
        row["fact"] = fact_list
    schedule["objects"] = objects
    return schedule


def s(value) -> str:
    return "" if value is None else str(value)


def normalize_repair_code(value: str) -> str:
    text = s(value).strip().upper().replace(" ", "").replace("-", "")
    latin_map = str.maketrans({
        "A": "А", "B": "В", "C": "С", "E": "Е", "H": "Н", "K": "К",
        "M": "М", "O": "О", "P": "Р", "T": "Т", "X": "Х", "Y": "У",
    })
    text = text.translate(latin_map)
    if any("А" <= c <= "Я" for c in text):
        return text
    if any(c.isdigit() for c in text):
        return "".join(filter(str.isdigit, text))
    return text


def month_index(month_name: str) -> int:
    return MONTHS_RU.index(month_name) + 1


def load_system_dates(year: int) -> dict[str, list[tuple[int, int]]]:
    transfer_dates: set[tuple[int, int]] = set(TRANSFER_HOLIDAYS_BY_YEAR.get(year, set()))
    holiday_dates: set[tuple[int, int]] = set(FIXED_HOLIDAYS)
    if not SOURCE_DB.exists():
        return {
            "transfer": sorted(transfer_dates),
            "holiday": sorted(holiday_dates),
        }

    try:
        with sqlite3.connect(SOURCE_DB) as conn:
            cur = conn.cursor()
            rows = cur.execute(
                "SELECT c, v FROM ts_norms_data WHERE y=? AND c IN (6, 7)",
                (year,),
            ).fetchall()
        for col_idx, raw_text in rows:
            if not raw_text:
                continue
            text = str(raw_text).replace(";", "\n").replace(",", "\n")
            for line in text.splitlines():
                parts = line.strip().split(".")
                if len(parts) < 2:
                    continue
                try:
                    day = int(parts[0])
                    month = int(parts[1])
                except ValueError:
                    continue
                if col_idx == 6:
                    transfer_dates.add((month, day))
                else:
                    holiday_dates.add((month, day))
    except Exception:
        pass

    return {
        "transfer": sorted(transfer_dates),
        "holiday": sorted(holiday_dates),
    }


def default_state(year: int) -> dict:
    months = []
    for month_num, month_name in enumerate(MONTHS_RU, 1):
        days = calendar.monthrange(year, month_num)[1]
        months.append(
            {
                "name": month_name,
                "month": month_num,
                "days": days,
                "plan": _default_table_rows(year, month_num, "plan"),
                "fact": _default_table_rows(year, month_num, "fact"),
            }
        )

    norms = {
        "h_tep": [{"k": label, "v": ""} for label in TEM_NORM_ROWS],
        "h_agr": [{"k": label, "v": ""} for label in AGR_NORM_ROWS],
        "p_tep": [{"k": MONTHS_RU[i], "v": ""} for i in range(12)],
        "p_agr": [{"k": MONTHS_RU[i], "v": ""} for i in range(12)],
    }
    inventory = []
    acts = {}
    notes = {}
    return {
        "year": year,
        "system_dates": load_system_dates(year),
        "months": months,
        "norms": norms,
        "acts": acts,
        "notes": notes,
        "repair_schedule": default_repair_schedule_state(),
        "repair_schedule_year": year,
    }


def load_repair_schedule_for_year(cur: sqlite3.Cursor, year: int) -> dict:
    schedule_rows = cur.execute("SELECT r, k, v FROM repair_schedule WHERE y=? ORDER BY r, k", (year,)).fetchall()
    schedule = default_repair_schedule_state()
    columns = schedule["columns"]
    objects: dict[int, dict[str, dict[int, str] | dict[str, str] | str]] = {}
    for row in schedule_rows:
        idx = int(row["r"])
        key = s(row["k"])
        value = s(row["v"])
        if idx == -1 and key.startswith("col_"):
            try:
                cidx = int(key[4:])
            except ValueError:
                continue
            if 0 <= cidx < len(columns) and value:
                columns[cidx]["code"] = normalize_repair_code(value)
            continue
        if idx >= 0 and key.startswith("periodicity_series_"):
            try:
                sidx = int(key.rsplit("_", 1)[1])
            except Exception:
                continue
            while len(schedule["periodicity"]["series"]) <= sidx:
                schedule["periodicity"]["series"].append("")
                schedule["periodicity"]["values"].append(["", "", "", "", ""])
            if 0 <= sidx < len(schedule["periodicity"]["series"]):
                schedule["periodicity"]["series"][sidx] = value
            continue
        if idx >= 0 and key.startswith("periodicity_value_"):
            try:
                _, _, r_str, c_str = key.split("_", 3)
                r_idx = int(r_str)
                c_idx = int(c_str)
            except Exception:
                continue
            while len(schedule["periodicity"]["values"]) <= r_idx:
                schedule["periodicity"]["series"].append("")
                schedule["periodicity"]["values"].append(["", "", "", "", ""])
            while len(schedule["periodicity"]["values"][r_idx]) <= c_idx:
                schedule["periodicity"]["values"][r_idx].append("")
            schedule["periodicity"]["values"][r_idx][c_idx] = value
            continue
        if idx < 0:
            continue
        obj = objects.setdefault(idx, {"series": "", "number": "", "plan": {}, "fact": {}})
        if key == "series":
            obj["series"] = value
        elif key == "number":
            obj["number"] = value
        elif key == "kr_plan":
            obj.setdefault("kr", {})["plan"] = value
        elif key == "kr_fact":
            obj.setdefault("kr", {})["fact"] = value
        elif key.startswith("plan_"):
            try:
                cidx = int(key[5:])
            except ValueError:
                continue
            obj["plan"][cidx] = value
        elif key.startswith("fact_"):
            try:
                cidx = int(key[5:])
            except ValueError:
                continue
            obj["fact"][cidx] = value
    schedule["columns"] = columns
    schedule["objects"] = []
    for idx in sorted(objects):
        obj = objects[idx]
        schedule["objects"].append({
            "series": s(obj["series"]),
            "number": s(obj["number"]),
            "kr": {
                "plan": s((obj.get("kr") or {}).get("plan", "")),
                "fact": s((obj.get("kr") or {}).get("fact", "")),
            },
            "plan": [s(obj["plan"].get(i, "")) for i in range(len(columns))],
            "fact": [s(obj["fact"].get(i, "")) for i in range(len(columns))],
        })
    if not schedule["objects"]:
        schedule["objects"] = default_repair_schedule_state()["objects"]
    return compute_repair_schedule_derived(schedule)


def _default_table_rows(year: int, month: int, table_type: str, rows: int = 14) -> list[dict]:
    days = calendar.monthrange(year, month)[1]
    result = []
    for idx in range(rows):
        result.append(
            {
                "excluded": False,
                "cells": [str(idx + 1), "", "", ""]
                + ["" for _ in range(days)]
                + [""],
            }
        )
    return result


def load_state(year: int, include_summary: bool = True) -> dict:
    state = default_state(year)
    with DB_LOCK, conn() as db:
        cur = db.cursor()
        repair_schedule_year = year
        try:
            row = cur.execute("SELECT v FROM repair_settings WHERE k='last_year'").fetchone()
            if row and s(row["v"]).strip():
                repair_schedule_year = int(s(row["v"]))
        except Exception:
            repair_schedule_year = year

        repairs = cur.execute(
            "SELECT m, t, r, c, v FROM repairs WHERE y=? ORDER BY m, t, r, c",
            (year,),
        ).fetchall()
        month_map = {m["name"]: m for m in state["months"]}
        for row in repairs:
            month_name = s(row["m"])
            table_type = s(row["t"])
            if month_name not in month_map or table_type not in {"plan", "fact"}:
                continue
            table = month_map[month_name][table_type]
            r = int(row["r"])
            c = int(row["c"])
            value = s(row["v"])
            while r >= len(table):
                table.append(_default_table_rows(year, month_index(month_name), table_type, 1)[0])
            if c == -1:
                table[r]["excluded"] = True
                continue
            if c == 999:
                table[r]["cells"][-1] = value
            elif 0 <= c <= 2:
                table[r]["cells"][c] = value
            elif 3 <= c < len(table[r]["cells"]) - 1:
                table[r]["cells"][c + 1] = value

        norms = cur.execute("SELECT cat, k, v FROM norms WHERE y=? ORDER BY cat, k", (year,)).fetchall()
        for row in norms:
            cat = s(row["cat"])
            if cat not in state["norms"]:
                continue
            key = s(row["k"])
            value = s(row["v"])
            if cat == "h_tep":
                idx = TEM_NORM_ROWS.index(key) if key in TEM_NORM_ROWS else -1
                if 0 <= idx < len(state["norms"][cat]):
                    state["norms"][cat][idx] = {"k": key or TEM_NORM_ROWS[idx], "v": value}
            elif cat == "h_agr":
                idx = AGR_NORM_ROWS.index(key) if key in AGR_NORM_ROWS else -1
                if 0 <= idx < len(state["norms"][cat]):
                    state["norms"][cat][idx] = {"k": key or AGR_NORM_ROWS[idx], "v": value}
            elif cat in {"p_tep", "p_agr"}:
                idx = -1
                if key in MONTHS_RU:
                    idx = MONTHS_RU.index(key)
                else:
                    try:
                        idx = int(key) - 1
                    except ValueError:
                        idx = -1
                if 0 <= idx < 12:
                    state["norms"][cat][idx] = {"k": key or MONTHS_RU[idx], "v": value}
            else:
                state["norms"][cat].append({"k": key, "v": value})

        acts = cur.execute("SELECT m, act_num, is_done, sap_order_done FROM acts_state WHERE y=? ORDER BY m, act_num", (year,)).fetchall()
        for row in acts:
            state["acts"].setdefault(s(row["m"]), {})[s(row["act_num"])] = {
                "is_done": bool(row["is_done"]),
                "sap_order_done": bool(row["sap_order_done"]),
            }

        notes = cur.execute("SELECT m, k, v FROM report_notes WHERE y=? ORDER BY m, k", (year,)).fetchall()
        for row in notes:
            state["notes"].setdefault(s(row["m"]), {})[s(row["k"])] = s(row["v"])

        state["repair_schedule_year"] = repair_schedule_year
        state["repair_schedule"] = load_repair_schedule_for_year(cur, repair_schedule_year)

        tu28_data = cur.execute("SELECT m, r, k, v FROM tu28_data WHERE y=? ORDER BY m, r, k", (year,)).fetchall()
        for row in tu28_data:
            month_name = s(row["m"])
            r = int(row["r"])
            k = s(row["k"])
            v = s(row["v"])
            if month_name not in month_map:
                continue
            table = month_map[month_name]["fact"]
            while r >= len(table):
                table.append(_default_table_rows(year, month_index(month_name), "fact", 1)[0])
            try:
                parsed_v = json.loads(v) if v else []
            except Exception:
                parsed_v = []
            if k == "tu28_extra":
                table[r]["tu28_extra"] = parsed_v
            elif k == "tu28_staff":
                table[r]["tu28_staff"] = parsed_v

        if include_summary:
            state["repair_summary"] = build_repair_summary_state(cur)

    return state


def _repair_summary_rows_from_month_state(state: dict) -> list[dict]:
    rows: list[dict] = []
    year = int(state.get("year") or 0)
    for month_index0, month in enumerate(state.get("months", []) or []):
        month_number = int(month.get("month") or month_index0 + 1)
        if not (1 <= month_number <= 12):
            continue
        month_days = int(month.get("days") or calendar.monthrange(year, month_number)[1])
        for row_index, row in enumerate(month.get("fact", []) or []):
            if not row or row.get("excluded"):
                continue
            key = report_unit_key(row)
            if not key:
                continue
            series, number = key
            cells = row.get("cells") or []
            for cell_index, value in enumerate(cells):
                if cell_index < 4 or cell_index >= 4 + month_days:
                    continue
                code = normalize_repair_code(value)
                if not code or not any("А" <= ch <= "Я" for ch in code):
                    continue
                day = cell_index - 3
                try:
                    date_value = dt.date(year, month_number, day)
                except Exception:
                    continue
                rows.append({
                    "rowIndex": row_index,
                    "locoKey": f"{series}|{number}",
                    "locoLabel": f"{series} {number}".strip(),
                    "series": series,
                    "number": number,
                    "repairCode": code,
                    "repairDate": _repair_schedule_format_date(date_value),
                    "repairDateSort": int(dt.datetime(year, month_number, day).timestamp() * 1000),
                    "columnIndex": cell_index,
                    "sourceKind": "month",
                })
    return rows


def _repair_summary_rows_from_schedule_state(state: dict) -> list[dict]:
    rows: list[dict] = []
    schedule = state.get("repair_schedule", {}) or {}
    columns = schedule.get("columns", []) if isinstance(schedule, dict) else []
    for row_index, row in enumerate(schedule.get("objects", []) or []):
        series = s(row.get("series")).strip()
        number = s(row.get("number")).strip()
        if not series and not number:
            continue
        loco_label = " ".join(part for part in [series, number] if part).strip()
        loco_key = f"{series}|{number}"

        def push_row(repair_code: str, date_value, column_index: int, source_kind: str):
            code = normalize_repair_code(repair_code)
            date_text = s(date_value).strip()
            if not code or not date_text or not any("А" <= ch <= "Я" for ch in code):
                return
            parsed = _repair_schedule_parse_date(date_text)
            if not parsed:
                return
            rows.append({
                "rowIndex": row_index,
                "locoKey": loco_key,
                "locoLabel": loco_label,
                "series": series,
                "number": number,
                "repairCode": code,
                "repairDate": date_text,
                "repairDateSort": int(dt.datetime(parsed.year, parsed.month, parsed.day).timestamp() * 1000),
                "columnIndex": column_index,
                "sourceKind": source_kind,
            })

        push_row("КР", (row.get("kr") or {}).get("fact", ""), -1, "kr")
        for cidx, col in enumerate(columns):
            push_row((col or {}).get("code", ""), (row.get("fact") or [])[cidx] if cidx < len(row.get("fact") or []) else "", cidx, "fact")
    return rows


def _repair_summary_pack(rows: list[dict]) -> dict:
    rows = list(rows or [])
    rows.sort(key=lambda item: (
        -int(item.get("repairDateSort") or 0),
        s(item.get("locoLabel")),
        s(item.get("repairCode")),
        int(item.get("columnIndex") or 0),
    ))
    types: list[str] = []
    locos: list[dict] = []
    seen_types: set[str] = set()
    seen_locos: set[str] = set()
    for row in rows:
        code = s(row.get("repairCode")).strip()
        if code and any("А" <= ch <= "Я" for ch in code) and code not in seen_types:
            seen_types.add(code)
            types.append(code)
        loco_key = s(row.get("locoKey")).strip()
        loco_label = s(row.get("locoLabel")).strip() or loco_key.replace("|", " ").strip()
        if loco_key and loco_key not in seen_locos:
            seen_locos.add(loco_key)
            locos.append({"key": loco_key, "label": loco_label})
    locos.sort(key=lambda item: s(item.get("label")).lower())
    return {"rows": rows, "types": types, "loco_options": locos}


def _load_kp_archive_measurements() -> list[dict]:
    if not SOURCE_DB.exists():
        return []
    try:
        with connect_sqlite(SOURCE_DB) as archive_conn:
            archive_rows = archive_conn.execute(
                """
                SELECT DISTINCT measurement_date, locomotive, repair_type
                FROM archive_data
                WHERE TRIM(COALESCE(measurement_date, '')) <> ''
                  AND TRIM(COALESCE(locomotive, '')) <> ''
                  AND TRIM(COALESCE(repair_type, '')) <> ''
                ORDER BY measurement_date
                """
            ).fetchall()
    except Exception:
        return []

    measurements: list[dict] = []
    for row in archive_rows:
        measurement_date = s(row["measurement_date"]).strip()
        locomotive = s(row["locomotive"]).strip()
        repair_code = normalize_repair_code(row["repair_type"])
        try:
            parsed_date = dt.date.fromisoformat(measurement_date)
        except Exception:
            parsed_date = _repair_schedule_parse_date(measurement_date)
        if not parsed_date or not locomotive or not repair_code:
            continue
        measurements.append({
            "number": locomotive,
            "repairCode": repair_code,
            "measurementDate": parsed_date.isoformat(),
        })
    return measurements


def build_repair_summary_state(cur: sqlite3.Cursor) -> dict:
    years: set[int] = set()
    for table in ("repairs", "repair_schedule"):
        try:
            fetched = cur.execute(f"SELECT DISTINCT y FROM {table}").fetchall()
        except Exception:
            fetched = []
        for row in fetched:
            try:
                years.add(int(row["y"]))
            except Exception:
                continue

    months_rows: list[dict] = []
    schedule_rows: list[dict] = []
    system_dates_by_year: dict[str, dict] = {}
    for year in sorted(years):
        try:
            year_state = load_state(year, include_summary=False)
        except Exception:
            continue
        system_dates_by_year[str(year)] = year_state.get("system_dates") or load_system_dates(year)
        months_rows.extend(_repair_summary_rows_from_month_state(year_state))
        schedule_rows.extend(_repair_summary_rows_from_schedule_state(year_state))

    return {
        "months": _repair_summary_pack(months_rows),
        "schedule": _repair_summary_pack(schedule_rows),
        "kp_measurements": _load_kp_archive_measurements(),
        "system_dates_by_year": system_dates_by_year,
    }


def find_act_template_path() -> Path | None:
    candidates = [
        ROOT / ACT_TEMPLATE_NAME,
        SOURCE_DIR / ACT_TEMPLATE_NAME,
        ROOT.parent / "dist" / "РТПС" / "_internal" / "График ППР" / ACT_TEMPLATE_NAME,
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def find_report_template_path() -> Path | None:
    candidates = [
        ROOT / REPORT_TEMPLATE_NAME,
        SOURCE_DIR / REPORT_TEMPLATE_NAME,
        ROOT.parent / "dist" / "РТПС" / "_internal" / "График ППР" / REPORT_TEMPLATE_NAME,
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def find_tu28_template_path() -> Path | None:
    candidates = [
        ROOT / TU28_TEMPLATE_NAME,
        SOURCE_DIR / TU28_TEMPLATE_NAME,
        ROOT.parent / "dist" / "РТПС" / "_internal" / "График ППР" / TU28_TEMPLATE_NAME,
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def read_norm_hours_from_state(items: list[dict]) -> dict[str, float]:
    norms: dict[str, float] = {}
    for item in items or []:
        key = s(item.get("k")).strip()
        value = s(item.get("v")).strip()
        if not key or not value:
            continue
        try:
            norms[key] = float(value.replace(",", "."))
        except Exception:
            continue
    return norms


def report_unit_key(row: dict) -> tuple[str, str] | None:
    cells = row.get("cells") or []
    series = s(cells[1]).strip().upper() if len(cells) > 1 else ""
    number = s(cells[2]).strip().upper() if len(cells) > 2 else ""
    if not series or not number:
        return None
    return series, number


def build_rows_by_unit(month_data: dict, table_type: str) -> dict[tuple[str, str], int]:
    rows_by_unit: dict[tuple[str, str], int] = {}
    for idx, row in enumerate(month_data.get(table_type, []) or []):
        if row.get("excluded"):
            continue
        key = report_unit_key(row)
        if key and key not in rows_by_unit:
            rows_by_unit[key] = idx
    return rows_by_unit


def row_cell_is_numeric(row: dict | None, day: int) -> bool:
    if not row or not row.get("cells"):
        return False
    idx = day + 3
    raw = s(row["cells"][idx]) if idx < len(row["cells"]) else ""
    if not raw.strip():
        return False
    try:
        float(raw.replace(",", "."))
    except ValueError:
        return False
    return True


def collect_unplanned_starts_across_months(
    year: int,
    months: list[dict],
    month_index: int,
    table_type: str,
    row_key: tuple[str, str],
) -> list[tuple[int, int]]:
    if not row_key or not months or not (0 <= month_index < len(months)):
        return []
    row_maps = [build_rows_by_unit(month, table_type) for month in months[: month_index + 1]]
    curr_month = months[month_index]
    curr_month_num = int(curr_month.get("month") or month_index + 1)
    prev_month_num = int(months[month_index - 1].get("month") or month_index) if month_index > 0 else None
    window_start = dt.date(year, prev_month_num, 26) if prev_month_num else dt.date(year, curr_month_num, 1)
    window_end = dt.date(year, curr_month_num, MONTH_DAY_LIMIT_FOR_REPORT)
    def row_for_date(date: dt.date) -> dict | None:
        month_idx = date.month - 1
        if month_idx < 0 or month_idx >= len(row_maps):
            return None
        row_idx = row_maps[month_idx].get(row_key)
        rows = months[month_idx].get(table_type, []) or []
        return rows[row_idx] if row_idx is not None and row_idx < len(rows) else None

    def numeric_on(date: dt.date) -> bool:
        return row_cell_is_numeric(row_for_date(date), date.day)

    starts: list[tuple[int, int]] = []
    seen = set()

    def add_start(date: dt.date) -> None:
        key = (date.day, date.month)
        if key in seen:
            return
        seen.add(key)
        starts.append(key)

    if numeric_on(window_start):
        start = window_start
        prev = start - dt.timedelta(days=1)
        while prev.year == year and prev >= dt.date(year, 1, 1) and numeric_on(prev):
            start = prev
            prev -= dt.timedelta(days=1)
        add_start(start)

    prev_is_num = numeric_on(window_start)
    day = window_start + dt.timedelta(days=1)
    while day <= window_end:
        is_num = numeric_on(day)
        if is_num and not prev_is_num:
            add_start(day)
        prev_is_num = is_num
        day += dt.timedelta(days=1)

    return starts


def collect_row_notes(row: dict, unplanned_starts: list[tuple[int, int]], number: str) -> list[str]:
    row_notes: list[str] = []
    cells = row.get("cells") or []
    note = s(cells[-1]).strip() if cells else ""
    has_unplanned = bool(unplanned_starts)
    is_auto_act_note = note.startswith("Акт № ") and "-" in note
    if note and not is_auto_act_note:
        row_notes.append(note)
    if has_unplanned:
        for day, month in unplanned_starts:
            auto_note = f"Акт № {day:02d}-{month:02d}-{number}"
            if auto_note not in row_notes:
                row_notes.append(auto_note)
    if not row_notes and note and is_auto_act_note:
        row_notes.append(note)
    return row_notes


def collect_report_act_notes_for_category(state: dict, month_index: int, table_type: str, category: str) -> list[str]:
    months = state.get("months", []) or []
    if not months or not (0 <= month_index < len(months)):
        return []
    year = int(state.get("year") or dt.date.today().year)
    month = months[month_index]
    notes: list[str] = []
    for row in month.get(table_type, []) or []:
        if row.get("excluded"):
            continue
        cells = row.get("cells") or []
        series = s(cells[1]).strip().upper() if len(cells) > 1 else ""
        number = s(cells[2]).strip().upper() if len(cells) > 2 else ""
        if not series or not number:
            continue
        row_category = "agr" if "ПЭ" in series else "tep"
        if row_category != category:
            continue
        key = report_unit_key(row)
        if not key:
            continue
        for day, month_num in collect_unplanned_starts_across_months(year, months, month_index, table_type, key):
            note = f"Акт № {day:02d}-{month_num:02d}-{number}"
            if note not in notes:
                notes.append(note)
    return notes


def process_report_day(acc, row: dict, day: int, month_num: int, category: str, table_type: str, state: dict) -> None:
    cells = row.get("cells") or []
    idx = day + 3
    val = s(cells[idx]).strip().upper() if idx < len(cells) else ""
    if val:
        try:
            hours = float(val.replace(",", "."))
        except ValueError:
            acc.result[table_type][category][val] = acc.result[table_type][category].get(val, 0) + 1
            state["last_is_num"] = False
        else:
            acc.unplanned_hours[table_type][category] += hours
            if not state["last_is_num"]:
                acc.unplanned_blocks[table_type][category] += 1
                state["unplanned_starts"].append((day, month_num))
            state["last_is_num"] = True
    else:
        state["last_is_num"] = False


def process_report_row(
    acc,
    table_type: str,
    curr_row: dict,
    prev_row: dict | None,
    curr_m: int,
    prev_m: int | None,
    fund_days: int,
    year: int,
    months: list[dict],
    month_index: int,
) -> None:
    if curr_row.get("excluded"):
        return
    cells = curr_row.get("cells") or []
    series = s(cells[1]).strip().upper() if len(cells) > 1 else ""
    number = s(cells[2]).strip().upper() if len(cells) > 2 else ""
    if not series:
        return
    category = "agr" if "ПЭ" in series else "tep"
    acc.units[table_type][category] += 1
    state = {"last_is_num": False, "unplanned_starts": []}
    if prev_row is not None and prev_m is not None:
        for day in range(MONTH_DAY_LIMIT_FOR_REPORT + 1, fund_days + 1):
            process_report_day(acc, prev_row, day, prev_m, category, table_type, state)
    for day in range(1, MONTH_DAY_LIMIT_FOR_REPORT + 1):
        process_report_day(acc, curr_row, day, curr_m, category, table_type, state)
    row_key = report_unit_key(curr_row)
    if table_type == "fact" and row_key:
        auto_starts = collect_unplanned_starts_across_months(year, months, month_index, table_type, row_key)
    else:
        auto_starts = state["unplanned_starts"]
    for note in collect_row_notes(curr_row, auto_starts, number):
        if note not in acc.notes[table_type][category]:
            acc.notes[table_type][category].append(note)


def get_report_period(year: int, month_name: str) -> tuple[int, int, int | None, str | None, int]:
    m_idx = MONTHS_RU.index(month_name)
    curr_m = m_idx + 1
    if m_idx == 0:
        return m_idx, curr_m, None, None, MONTH_DAY_LIMIT_FOR_REPORT
    prev_m_idx = m_idx - 1
    prev_month_name = MONTHS_RU[prev_m_idx]
    fund_days = calendar.monthrange(year, prev_m_idx + 1)[1]
    return m_idx, curr_m, prev_m_idx + 1, prev_month_name, fund_days


def calculate_ok_units(state: dict, year: int, month_name: str, acc, period_hours: int) -> tuple[float, float]:
    _, curr_m, prev_m, _, fund_days = get_report_period(year, month_name)
    _ = curr_m, prev_m, fund_days
    n_tep = read_norm_hours_from_state(state.get("norms", {}).get("h_tep", []))
    n_agr = read_norm_hours_from_state(state.get("norms", {}).get("h_agr", []))
    fact_tep_hours = acc.unplanned_hours["fact"]["tep"] + sum(
        acc.result["fact"]["tep"].get(code, 0) / factor * n_tep.get(code, 0)
        for code, factor in TEP_HOUR_FACTORS.items()
    )
    fact_agr_hours = acc.unplanned_hours["fact"]["agr"] + sum(
        acc.result["fact"]["agr"].get(code, 0) / factor * n_agr.get(code, 0)
        for code, factor in AGR_HOUR_FACTORS.items()
    )
    fact_tep_ok = acc.units["fact"]["tep"] - (fact_tep_hours / period_hours) if period_hours else 0
    fact_agr_ok = acc.units["fact"]["agr"] - (fact_agr_hours / period_hours) if period_hours else 0
    return fact_tep_ok, fact_agr_ok


def calculate_report_data_from_state(state: dict, month_name: str) -> dict:
    m_idx, curr_m, prev_m, prev_month_name, fund_days = get_report_period(int(state.get("year") or dt.date.today().year), month_name)
    period_hours = fund_days * 24
    acc = type("ReportAccumulatorLike", (), {})()
    acc.result = {"plan": {"tep": {}, "agr": {}}, "fact": {"tep": {}, "agr": {}}}
    acc.unplanned_blocks = {"plan": {"tep": 0, "agr": 0}, "fact": {"tep": 0, "agr": 0}}
    acc.unplanned_hours = {"plan": {"tep": 0.0, "agr": 0.0}, "fact": {"tep": 0.0, "agr": 0.0}}
    acc.units = {"plan": {"tep": 0, "agr": 0}, "fact": {"tep": 0, "agr": 0}}
    acc.notes = {"plan": {"tep": [], "agr": []}, "fact": {"tep": [], "agr": []}}
    acc.excluded = {"plan": set(), "fact": set()}

    months = state.get("months", []) or []
    if not (0 <= m_idx < len(months)):
        raise ValueError("Не найден месяц отчета")
    curr_month = months[m_idx]
    prev_month = months[m_idx - 1] if prev_month_name and m_idx - 1 >= 0 else None
    prev_rows_by_unit = build_rows_by_unit(prev_month, "plan") if prev_month else {}
    prev_rows_by_unit_fact = build_rows_by_unit(prev_month, "fact") if prev_month else {}

    for table_type in ["plan", "fact"]:
        curr_table = curr_month.get(table_type, []) or []
        prev_rows = prev_rows_by_unit_fact if table_type == "fact" else prev_rows_by_unit
        for row in curr_table:
            key = report_unit_key(row)
            if row.get("excluded") and key:
                acc.excluded[table_type].add(key)
            prev_row = prev_month.get(table_type, [])[prev_rows[key]] if prev_month and key and key in prev_rows else None
            process_report_row(acc, table_type, row, prev_row, curr_m, prev_m, fund_days, int(state.get("year") or dt.date.today().year), months, m_idx)

    if acc.notes["fact"]["tep"] is not None:
        for note in collect_report_act_notes_for_category(state, m_idx, "fact", "tep"):
            if note not in acc.notes["fact"]["tep"]:
                acc.notes["fact"]["tep"].append(note)
    if acc.notes["fact"]["agr"] is not None:
        for note in collect_report_act_notes_for_category(state, m_idx, "fact", "agr"):
            if note not in acc.notes["fact"]["agr"]:
                acc.notes["fact"]["agr"].append(note)

    fact_tep_ok, fact_agr_ok = calculate_ok_units(state, int(state.get("year") or dt.date.today().year), month_name, acc, period_hours)

    def month_norm_value(key: str, fallback_index: int) -> str:
        for item in state.get("norms", {}).get(key, []) or []:
            item_key = s(item.get("k")).strip()
            item_value = s(item.get("v")).strip()
            if item_key == month_name:
                return item_value or "0"
        items = state.get("norms", {}).get(key, []) or []
        if 0 <= fallback_index < len(items):
            return s(items[fallback_index].get("v")).strip() or "0"
        return "0"

    return {
        "m": month_name,
        "y": int(state.get("year") or dt.date.today().year),
        "tp": month_norm_value("p_tep", m_idx),
        "tf": fact_tep_ok,
        "ap": month_norm_value("p_agr", m_idx),
        "af": fact_agr_ok,
        "res": acc.result,
        "ub": acc.unplanned_blocks,
        "notes": acc.notes,
        "excluded": {
            "plan": sorted(["|".join(key) for key in acc.excluded["plan"]]),
            "fact": sorted(["|".join(key) for key in acc.excluded["fact"]]),
        },
    }


def build_report_excel_tags(month_name: str, data: dict, saved_notes: dict[str, str]) -> dict[str, str]:
    def count(section: str, category: str, code: str, factor: int) -> float:
        return data["res"][section][category].get(code, 0) / factor

    def merge_notes(saved: str, auto: str) -> str:
        saved = s(saved).strip()
        auto = s(auto).strip()
        if saved and auto:
            lines = []
            seen = set()
            for chunk in (saved, auto):
                for line in chunk.splitlines():
                    line = line.strip()
                    if not line or line in seen:
                        continue
                    seen.add(line)
                    lines.append(line)
            return "\n".join(lines)
        return saved or auto

    tep_plan = {code: count("plan", "tep", code, factor) for code, factor in TEP_REPORT_FACTORS.items()}
    tep_fact = {code: count("fact", "tep", code, factor) for code, factor in TEP_REPORT_FACTORS.items()}
    agr_plan = {code: count("plan", "agr", code, factor) for code, factor in AGR_REPORT_FACTORS.items()}
    agr_fact = {code: count("fact", "agr", code, factor) for code, factor in AGR_REPORT_FACTORS.items()}

    p_tep_ub = data["ub"]["plan"]["tep"]
    f_tep_ub = data["ub"]["fact"]["tep"]
    p_agr_ub = data["ub"]["plan"]["agr"]
    f_agr_ub = data["ub"]["fact"]["agr"]

    sum_p = sum(tep_plan.values()) + p_tep_ub + sum(agr_plan.values()) + p_agr_ub
    sum_f = sum(tep_fact.values()) + f_tep_ub + sum(agr_fact.values()) + f_agr_ub

    return {
        "[МЕСЯЦ]": month_name,
        "[ГОД]": str(data["y"]),
        "[ПЛАН_ТЕП_ПАРК]": format_n(data["tp"]),
        "[ФАКТ_ТЕП_ПАРК]": format_n(data["tf"]),
        "[ПРИМ_ТЕП_ПАРК]": saved_notes.get("tep_park", ""),
        "[ПЛАН_ТЕП_ТО2]": format_n(tep_plan["ТО2"]),
        "[ФАКТ_ТЕП_ТО2]": format_n(tep_fact["ТО2"]),
        "[ПРИМ_ТЕП_ТО2]": saved_notes.get("tep_ТО2", ""),
        "[ПЛАН_ТЕП_ТО3]": format_n(tep_plan["ТО3"]),
        "[ФАКТ_ТЕП_ТО3]": format_n(tep_fact["ТО3"]),
        "[ПРИМ_ТЕП_ТО3]": saved_notes.get("tep_ТО3", ""),
        "[ПЛАН_ТЕП_ТР1]": format_n(tep_plan["ТР1"]),
        "[ФАКТ_ТЕП_ТР1]": format_n(tep_fact["ТР1"]),
        "[ПРИМ_ТЕП_ТР1]": saved_notes.get("tep_ТР1", ""),
        "[ПЛАН_ТЕП_ТР2]": format_n(tep_plan["ТР2"]),
        "[ФАКТ_ТЕП_ТР2]": format_n(tep_fact["ТР2"]),
        "[ПРИМ_ТЕП_ТР2]": saved_notes.get("tep_ТР2", ""),
        "[ПЛАН_ТЕП_ТР3]": format_n(tep_plan["ТР3"]),
        "[ФАКТ_ТЕП_ТР3]": format_n(tep_fact["ТР3"]),
        "[ПРИМ_ТЕП_ТР3]": saved_notes.get("tep_ТР3", ""),
        "[ПЛАН_ТЕП_НЕПЛАН]": format_n(p_tep_ub),
        "[ФАКТ_ТЕП_НЕПЛАН]": format_n(f_tep_ub),
        "[ПРИМ_ТЕП_НЕПЛАН]": merge_notes(saved_notes.get("tep_ТР_unplan", ""), "\n".join(data["notes"]["fact"]["tep"])),
        "[ПЛАН_АГР_ПАРК]": format_n(data["ap"]),
        "[ФАКТ_АГР_ПАРК]": format_n(data["af"]),
        "[ПРИМ_АГР_ПАРК]": saved_notes.get("agr_park", ""),
        "[ПЛАН_АГР_ТО]": format_n(agr_plan["ТО"]),
        "[ФАКТ_АГР_ТО]": format_n(agr_fact["ТО"]),
        "[ПРИМ_АГР_ТО]": saved_notes.get("agr_ТО", ""),
        "[ПЛАН_АГР_ТР]": format_n(agr_plan["ТР"]),
        "[ФАКТ_АГР_ТР]": format_n(agr_fact["ТР"]),
        "[ПРИМ_АГР_ТР]": saved_notes.get("agr_ТР", ""),
        "[ПЛАН_АГР_НЕПЛАН]": format_n(p_agr_ub),
        "[ФАКТ_АГР_НЕПЛАН]": format_n(f_agr_ub),
        "[ПРИМ_АГР_НЕПЛАН]": merge_notes(saved_notes.get("agr_ТР_unplan", ""), "\n".join(data["notes"]["fact"]["agr"])),
        "[ПЛАН_СУММА]": format_n(sum_p),
        "[ФАКТ_СУММА]": format_n(sum_f),
    }


def build_report_preview(month_name: str, data: dict, saved_notes: dict[str, str]) -> dict:
    def count(section: str, category: str, code: str, factor: int) -> str:
        return format_n(data["res"][section][category].get(code, 0) / factor)

    def merge_notes(saved: str, auto: str) -> str:
        saved = s(saved).strip()
        auto = s(auto).strip()
        if saved and auto:
            lines = []
            seen = set()
            for chunk in (saved, auto):
                for line in chunk.splitlines():
                    line = line.strip()
                    if not line or line in seen:
                        continue
                    seen.add(line)
                    lines.append(line)
            return "\n".join(lines)
        return saved or auto

    tep_notes = "\n".join(data["notes"]["fact"]["tep"])
    agr_notes = "\n".join(data["notes"]["fact"]["agr"])
    rows = [
        {"kind": "group", "label": "Кол-во тех.испр. локомотивов\nТЕПЛОВОЗЫ МАНЕВРОВЫЕ", "plan": format_n(data["tp"]), "fact": format_n(data["tf"]), "note_key": "tep_park", "note": saved_notes.get("tep_park", "")},
        {"kind": "row", "key": "tep_ТО2", "label": "ТО2", "plan": count("plan", "tep", "ТО2", 1), "fact": count("fact", "tep", "ТО2", 1), "note": saved_notes.get("tep_ТО2", "")},
        {"kind": "row", "key": "tep_ТО3", "label": "ТО3", "plan": count("plan", "tep", "ТО3", 2), "fact": count("fact", "tep", "ТО3", 2), "note": saved_notes.get("tep_ТО3", "")},
        {"kind": "row", "key": "tep_ТР1", "label": "ТР1", "plan": count("plan", "tep", "ТР1", 5), "fact": count("fact", "tep", "ТР1", 5), "note": saved_notes.get("tep_ТР1", "")},
        {"kind": "row", "key": "tep_ТР2", "label": "ТР2", "plan": count("plan", "tep", "ТР2", 10), "fact": count("fact", "tep", "ТР2", 10), "note": saved_notes.get("tep_ТР2", "")},
        {"kind": "row", "key": "tep_ТР3", "label": "ТР3", "plan": count("plan", "tep", "ТР3", 15), "fact": count("fact", "tep", "ТР3", 15), "note": saved_notes.get("tep_ТР3", "")},
        {"kind": "row", "key": "tep_ТР_unplan", "label": "ТР (текущий ремонт)", "plan": format_n(data["ub"]["plan"]["tep"]), "fact": format_n(data["ub"]["fact"]["tep"]), "note": merge_notes(saved_notes.get("tep_ТР_unplan", ""), tep_notes)},
        {"kind": "group", "label": "Кол-во тех.испр. локомотивов\nАГРЕГАТЫ ТЯГОВЫЕ", "plan": format_n(data["ap"]), "fact": format_n(data["af"]), "note_key": "agr_park", "note": saved_notes.get("agr_park", "")},
        {"kind": "row", "key": "agr_ТО", "label": "ТО", "plan": count("plan", "agr", "ТО", 1), "fact": count("fact", "agr", "ТО", 1), "note": saved_notes.get("agr_ТО", "")},
        {"kind": "row", "key": "agr_ТР", "label": "ТР", "plan": count("plan", "agr", "ТР", 5), "fact": count("fact", "agr", "ТР", 5), "note": saved_notes.get("agr_ТР", "")},
        {"kind": "row", "key": "agr_ТР_unplan", "label": "ТР (текущий ремонт)", "plan": format_n(data["ub"]["plan"]["agr"]), "fact": format_n(data["ub"]["fact"]["agr"]), "note": merge_notes(saved_notes.get("agr_ТР_unplan", ""), agr_notes)},
    ]
    return {"month": month_name, "year": data["y"], "rows": rows}


def build_report_workbook(year: int, month_name: str, state: dict | None = None) -> tuple[bytes, str]:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as exc:
        raise RuntimeError("На сервере не установлен openpyxl") from exc

    state = state or load_state(year)
    data = calculate_report_data_from_state(state, month_name)
    saved_notes = state.get("notes", {}).get(month_name, {}) or {}
    tags = build_report_excel_tags(month_name, data, saved_notes)

    template_path = find_report_template_path()
    if template_path:
        wb = load_workbook(template_path)
        replace_tags_in_workbook(wb, tags)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"
        for idx, (k, v) in enumerate(tags.items(), start=1):
            ws[f"A{idx}"] = k
            ws[f"B{idx}"] = v
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="top")

    out = BytesIO()
    wb.save(out)
    return out.getvalue(), f"Отчет_{month_name}_{year}.xlsx"


def get_act_inventory_item(year: int, number: str) -> tuple[str, str]:
    if not SOURCE_DB.exists():
        return "", ""
    try:
        with sqlite3.connect(SOURCE_DB) as db:
            cur = db.cursor()
            row = cur.execute("SELECT ser, inv FROM inventory WHERE y=? AND num=?", (year, number)).fetchone()
        if not row:
            return "", ""
        return s(row[0]), s(row[1])
    except Exception:
        return "", ""


def format_fio_initials(full_name: str) -> str:
    parts = s(full_name).split()
    if len(parts) >= 3:
        return f"{parts[0]} {parts[1][0]}. {parts[2][0]}."
    if len(parts) == 2:
        return f"{parts[0]} {parts[1][0]}."
    return s(full_name).strip()


def get_all_employee_names() -> list[str]:
    try:
        if not SOURCE_DB.exists():
            return []
        with sqlite3.connect(SOURCE_DB) as db:
            cur = db.cursor()
            cur.execute(
                "SELECT DISTINCT CASE WHEN COALESCE(full_name, '') != '' THEN full_name ELSE name END "
                "FROM employees ORDER BY 1"
            )
            return [s(row[0]).strip() for row in cur.fetchall() if s(row[0]).strip()]
    except Exception:
        return []


def _parse_tabel_date(value: str, year: int | None = None) -> dt.date | None:
    text = s(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except Exception:
            pass
    parts = text.split(".")
    if len(parts) == 2 and year:
        try:
            return dt.date(int(year), int(parts[1]), int(parts[0]))
        except Exception:
            return None
    return None


def get_employee_vacations() -> dict[str, list[dict]]:
    if not SOURCE_DB.exists():
        return {}
    try:
        with sqlite3.connect(SOURCE_DB) as db:
            db.row_factory = sqlite3.Row
            rows = db.execute(
                """
                SELECT e.y, e.name, e.full_name, e.tab_num, v.c, v.v
                FROM employees e
                JOIN vacations v ON v.y=e.y AND v.tab_num=e.tab_num
                WHERE TRIM(COALESCE(e.tab_num, '')) <> ''
                  AND TRIM(COALESCE(v.v, '')) <> ''
                ORDER BY e.y, e.rowid, v.c
                """
            ).fetchall()
    except Exception:
        return {}

    grouped: dict[tuple[int, str], dict[int, str]] = {}
    names: dict[tuple[int, str], set[str]] = {}
    for row in rows:
        try:
            year = int(row["y"])
            col = int(row["c"])
        except Exception:
            continue
        tab_num = s(row["tab_num"]).strip()
        if not tab_num:
            continue
        key = (year, tab_num)
        grouped.setdefault(key, {})[col] = s(row["v"]).strip()
        name_values = {s(row["name"]).strip(), s(row["full_name"]).strip()}
        names.setdefault(key, set()).update(item for item in name_values if item)

    result: dict[str, list[dict]] = {}
    for (year, _tab_num), cells in grouped.items():
        for start_col, end_col in ((1, 2), (5, 6), (9, 10)):
            start = _parse_tabel_date(cells.get(start_col, ""), year)
            end = _parse_tabel_date(cells.get(end_col, ""), year)
            if not start or not end:
                continue
            if end < start:
                start, end = end, start
            item = {
                "start": start.isoformat(),
                "end": end.isoformat(),
                "label": f"{start.strftime('%d.%m.%Y')}-{end.strftime('%d.%m.%Y')}",
            }
            for name in names.get((year, _tab_num), set()):
                result.setdefault(name, []).append(item)
    return result


def replace_tags_in_workbook(wb, tags: dict[str, str]) -> None:
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value:
                    value = cell.value
                    for tag, rep in tags.items():
                        if tag in value:
                            value = value.replace(tag, s(rep))
                    cell.value = value


def build_act_workbook(year: int, act: str) -> tuple[bytes, str]:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as exc:
        raise RuntimeError("На сервере не установлен openpyxl") from exc

    clean_act_num = act.replace("Акт № ", "").strip()
    parts = clean_act_num.split("-")
    if len(parts) != 3:
        raise ValueError("Не удалось распознать формат акта")

    d_act, m_act, num_act = parts
    months_ru = {
        "01": "января", "02": "февраля", "03": "марта",
        "04": "апреля", "05": "мая", "06": "июня",
        "07": "июля", "08": "августа", "09": "сентября",
        "10": "октября", "11": "ноября", "12": "декабря",
    }
    date_str = f"{d_act} {months_ru.get(m_act, 'января')} {year} г."
    ser, inv = get_act_inventory_item(year, num_act)
    if "ПЭ" in ser.upper():
        eq_type = "Тяговый агрегат"
    elif ser:
        eq_type = "Тепловоз маневровый"
    else:
        eq_type = ""

    tags = {
        "[АКТ]": clean_act_num, "[Акт]": clean_act_num, "[акт]": clean_act_num,
        "[ДАТА]": date_str, "[Дата]": date_str,
        "[НОМЕР]": num_act, "[Номер]": num_act,
        "[АГРЕГАТ]": eq_type, "[Агрегат]": eq_type,
        "[СЕРИЯ]": ser, "[Серия]": ser,
        "[ИНВ]": inv, "[Инв]": inv,
    }

    template_path = find_act_template_path()
    if template_path:
        wb = load_workbook(template_path)
        replace_tags_in_workbook(wb, tags)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Акт"
        ws["A1"] = "Акт"
        ws["B1"] = clean_act_num
        ws["A2"] = "Дата"
        ws["B2"] = date_str
        ws["A3"] = "Серия"
        ws["B3"] = ser
        ws["A4"] = "Инвентарный номер"
        ws["B4"] = inv
        ws["A5"] = "Тип"
        ws["B5"] = eq_type
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for cell in ws["A1:B5"]:
            for item in cell:
                item.alignment = Alignment(horizontal="left")

    out = BytesIO()
    wb.save(out)
    return out.getvalue(), f"Акт_{clean_act_num}.xlsx"


def build_tu28_workbook(year: int, month_name: str, row_idx: int, staff_list: list[str] | None = None, state: dict | None = None, extra_repairs: list[str] | None = None) -> tuple[bytes, str]:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as exc:
        raise RuntimeError("На сервере не установлен openpyxl") from exc

    state = state or load_state(year)
    month = next((m for m in state.get("months", []) if s(m.get("name")) == month_name), None)
    if not month:
        raise ValueError("Не удалось найти месяц")
    fact_rows = month.get("fact") or []
    if row_idx < 0 or row_idx >= len(fact_rows):
        raise ValueError("Не удалось найти строку ремонта")
    row = fact_rows[row_idx]
    cells = row.get("cells") or []
    series = s(cells[1]).strip()
    number = s(cells[2]).strip()
    if not number:
        raise ValueError("Не удалось определить номер")

    repair_code = ""
    repair_days = []
    for col in range(4, 4 + int(month.get("days") or 0)):
        value = normalize_repair_code(s(cells[col]) if col < len(cells) else "")
        if value in TU28_REPAIR_CODES:
            if not repair_code:
                repair_code = s(cells[col]).strip().upper()
            repair_days.append(col - 3)
    if not repair_days:
        raise ValueError("В выбранной строке не найден ремонт для ТУ-28")

    if "ПЭ" in series.upper():
        eq_type = "Тяговый агрегат"
    elif series:
        eq_type = "Тепловоз маневровый"
    else:
        eq_type = ""

    try:
        month_num = int(month.get("month") or 0)
    except Exception:
        month_num = 0
    repair_day = repair_days[0]
    date_str = f"{repair_day:02d}.{month_num:02d}.{year}"

    start_date_str = f"{year}-{month_num:02d}-{repair_days[0]:02d}"
    end_date_str = f"{year}-{month_num:02d}-{repair_days[-1]:02d}"

    db_path = Path(__file__).resolve().parent.parent / "base" / "common_database.db"
    measurements = {}
    print(f"DEBUG: ZAMER KP: path_exists={db_path.exists()} number={number} start={start_date_str} end={end_date_str}", flush=True)
    if db_path.exists() and number and start_date_str and end_date_str:
        import sqlite3
        try:
            with sqlite3.connect(db_path) as conn:
                cur = conn.cursor()
                db_rows = cur.execute(
                    """
                    SELECT measurement_date, r, c, v 
                    FROM archive_data 
                    WHERE locomotive=? AND measurement_date >= ? AND measurement_date <= ?
                    ORDER BY measurement_date DESC, r, c
                    """,
                    (number, start_date_str, end_date_str)
                ).fetchall()
                print(f"DEBUG: ZAMER KP: Fetched {len(db_rows)} rows from archive for {number} between {start_date_str} and {end_date_str}", flush=True)
                dates_dict = {}
                for d_str, r, c, v in db_rows:
                    if d_str not in dates_dict:
                        dates_dict[d_str] = {}
                    try:
                        r_int = int(r)
                        c_int = int(c)
                    except Exception:
                        continue
                    dates_dict[d_str].setdefault(r_int, {})[c_int] = str(v).strip() if v else ""
                if dates_dict:
                    best_date = sorted(dates_dict.keys(), reverse=True)[0]
                    measurements = dates_dict[best_date]
                    print(f"DEBUG: ZAMER KP: best_date={best_date} measurements dict: {measurements}", flush=True)
                    print(f"DEBUG: ZAMER KP: best_date={best_date} measurements size={len(measurements)}", flush=True)
        except Exception as e:
            print("Error reading Zamer KP archive db:", e, flush=True)
    tags = {
        "[СЕРИЯ]": series,
        "[НОМЕР]": number,
        "[ДАТА]": date_str,
        "[ВИД]": repair_code,
        "[АГРЕГАТ]": eq_type,
        "[ДИЗЕЛЬ]": "",
        "[ЭКИПАЖ 1]": "",
        "[ЭКИПАЖ 2]": "",
        "[АКБ]": "",
        "[ЭЛМАШ]": "",
        "[ЭЛАП]": "",
        "[ТОРМОЗ]": "",
    }

    col_to_tag_prefix = {
        2: "ПР_Л", 3: "ПР_П",
        4: "ТГ_Л", 5: "ТГ_П",
        6: "КР_Л", 7: "КР_П",
        8: "ТБ_Л", 9: "ТБ_П",
        10: "ДБ_Л", 11: "ДБ_П"
    }
    for axle in range(1, 13):
        for c_idx, prefix in col_to_tag_prefix.items():
            # In archive_data, r=2 corresponds to axle 1, r=3 to axle 2, etc.
            measurement_value = s(measurements.get(axle + 1, {}).get(c_idx, "")).strip()
            tags[f"[{prefix}_{axle}]"] = measurement_value.replace(".", ",")
            
    print(f"DEBUG: ZAMER KP TAGS: { {k: tags[k] for k in tags if 'ПР' in k or 'ТГ' in k} }", flush=True)

    components = [
        "ДИЗЕЛЬ",
        "ЭКИПАЖ 1",
        "ЭКИПАЖ 2",
        "АКБ",
        "ЭЛМАШ",
        "ЭЛАП",
        "ТОРМОЗ",
    ]
    for idx, name in enumerate(staff_list or []):
        if idx >= len(components):
            break
        tags[f"[{components[idx]}]"] = format_fio_initials(name)

    extra_repairs = extra_repairs or []
    print(f"DEBUG: build_tu28_workbook received extra_repairs={extra_repairs}", flush=True)
    for i in range(1, 21):
        tags[f"[ДОП_РЕМОНТ_{i}]"] = ""
        tags[f"[ДОП_НОМЕР_{i}]"] = ""
    for i, extra in enumerate(extra_repairs, start=1):
        if i <= 20:
            val = str(extra).strip()
            if val:
                tags[f"[ДОП_РЕМОНТ_{i}]"] = val
                tags[f"[ДОП_НОМЕР_{i}]"] = str(i)
    print(f"DEBUG: tags built: {[k for k in tags.keys() if 'ДОП' in k and tags[k]]}", flush=True)

    template_path = find_tu28_template_path()
    if template_path:
        wb = load_workbook(template_path)
        replace_tags_in_workbook(wb, tags)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "ТУ-28"
        for idx, (k, v) in enumerate(tags.items(), start=1):
            ws[f"A{idx}"] = k
            ws[f"B{idx}"] = v
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(horizontal="left", vertical="top")

    out = BytesIO()
    wb.save(out)
    return out.getvalue(), f"ТУ-28_{month_name}_{year}.xlsx"


def content_disposition_attachment(filename: str) -> str:
    ascii_name = filename.encode("ascii", "ignore").decode("ascii") or "file.xlsx"
    return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(filename)}"


def save_state(state: dict) -> dict:
    year = int(state.get("year") or dt.date.today().year)
    with DB_LOCK, conn() as db:
        cur = db.cursor()
        with db:
            cur.execute("DELETE FROM repairs WHERE y=?", (year,))
            cur.execute("DELETE FROM norms WHERE y=?", (year,))
            cur.execute("DELETE FROM acts_state WHERE y=?", (year,))
            cur.execute("DELETE FROM report_notes WHERE y=?", (year,))
            cur.execute("DELETE FROM tu28_data WHERE y=?", (year,))
            cur.execute("DELETE FROM repair_schedule WHERE y=?", (year,))
            cur.execute("INSERT OR REPLACE INTO repair_settings VALUES ('last_year', ?)", (str(year),))

            for month in state.get("months", []):
                month_name = s(month.get("name"))
                for table_type in ["plan", "fact"]:
                    for r, row in enumerate(month.get(table_type, [])):
                        if row.get("excluded"):
                            cur.execute("INSERT INTO repairs VALUES (?,?,?,?,?,?)", (year, month_name, table_type, r, -1, "EXC"))
                        cells = row.get("cells", [])
                        for c, value in enumerate(cells):
                            value = s(value).strip()
                            if value:
                                if c == 3:
                                    continue
                                db_c = 999 if c == len(cells) - 1 else (c - 1 if c >= 4 else c)
                                cur.execute("INSERT INTO repairs VALUES (?,?,?,?,?,?)", (year, month_name, table_type, r, db_c, value))
                        if table_type == "fact":
                            if "tu28_extra" in row:
                                cur.execute("INSERT INTO tu28_data VALUES (?,?,?,?,?)", (year, month_name, r, "tu28_extra", json.dumps(row["tu28_extra"])))
                            if "tu28_staff" in row:
                                cur.execute("INSERT INTO tu28_data VALUES (?,?,?,?,?)", (year, month_name, r, "tu28_staff", json.dumps(row["tu28_staff"])))
                            if "tu28_locked" in row:
                                cur.execute("INSERT INTO tu28_data VALUES (?,?,?,?,?)", (year, month_name, r, "tu28_locked", json.dumps(row["tu28_locked"])))

            for cat, rows in state.get("norms", {}).items():
                for row in rows:
                    k = s(row.get("k")).strip()
                    v = s(row.get("v")).strip()
                    if k or v:
                        cur.execute("INSERT INTO norms VALUES (?,?,?,?)", (year, cat, k, v))

            for m_name, acts in state.get("acts", {}).items():
                for act_num, flags in acts.items():
                    cur.execute(
                        "INSERT INTO acts_state VALUES (?,?,?,?,?)",
                        (year, m_name, act_num, 1 if flags.get("is_done") else 0, 1 if flags.get("sap_order_done") else 0),
                    )

            for m_name, keys in state.get("notes", {}).items():
                for key, value in keys.items():
                    value = s(value)
                    if value:
                        cur.execute("INSERT INTO report_notes VALUES (?,?,?,?)", (year, m_name, key, value))

            schedule = state.get("repair_schedule", {}) or {}
            if isinstance(schedule, dict):
                schedule = compute_repair_schedule_derived(schedule)
                state["repair_schedule"] = schedule
            columns = schedule.get("columns", []) if isinstance(schedule, dict) else []
            objects = schedule.get("objects", []) if isinstance(schedule, dict) else []
            periodicity = schedule.get("periodicity", {}) if isinstance(schedule, dict) else {}
            for cidx, col in enumerate(columns):
                value = normalize_repair_code(s((col or {}).get("code")).strip())
                if value:
                    cur.execute("INSERT INTO repair_schedule VALUES (?,?,?,?)", (year, -1, f"col_{cidx}", value))
            series_rows = periodicity.get("series", []) if isinstance(periodicity, dict) else []
            for r, value in enumerate(series_rows):
                value = s(value).strip()
                if value:
                    cur.execute("INSERT INTO repair_schedule VALUES (?,?,?,?)", (year, r, f"periodicity_series_{r}", value))
            values_rows = periodicity.get("values", []) if isinstance(periodicity, dict) else []
            for r, row in enumerate(values_rows):
                if not isinstance(row, list):
                    continue
                for cidx, value in enumerate(row):
                    value = s(value).strip()
                    if value:
                        cur.execute("INSERT INTO repair_schedule VALUES (?,?,?,?)", (year, r, f"periodicity_value_{r}_{cidx}", value))
            for r, row in enumerate(objects):
                if not isinstance(row, dict):
                    continue
                series = s(row.get("series")).strip()
                number = s(row.get("number")).strip()
                if series:
                    cur.execute("INSERT INTO repair_schedule VALUES (?,?,?,?)", (year, r, "series", series))
                if number:
                    cur.execute("INSERT INTO repair_schedule VALUES (?,?,?,?)", (year, r, "number", number))
                kr = row.get("kr") or {}
                kr_plan = s(kr.get("plan")).strip()
                kr_fact = s(kr.get("fact")).strip()
                if kr_plan:
                    cur.execute("INSERT INTO repair_schedule VALUES (?,?,?,?)", (year, r, "kr_plan", kr_plan))
                if kr_fact:
                    cur.execute("INSERT INTO repair_schedule VALUES (?,?,?,?)", (year, r, "kr_fact", kr_fact))
                for cidx, value in enumerate(row.get("plan", []) or []):
                    value = s(value).strip()
                    if value:
                        cur.execute("INSERT INTO repair_schedule VALUES (?,?,?,?)", (year, r, f"plan_{cidx}", value))
                for cidx, value in enumerate(row.get("fact", []) or []):
                    value = s(value).strip()
                    if value:
                        cur.execute("INSERT INTO repair_schedule VALUES (?,?,?,?)", (year, r, f"fact_{cidx}", value))

    return load_state(year)


def json_response(handler: BaseHTTPRequestHandler, payload: dict, status: int = 200) -> None:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
    handler.send_header("Pragma", "no-cache")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def _cookie_value(username: str, role: str) -> str:
    payload = f"{username}:{role}:{int(dt.datetime.now().timestamp())}"
    signature = hmac.new(WEB_SECRET.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    return f"{payload}:{signature}"


def _verify_cookie(value: str) -> tuple[str, str, str, str] | None:
    for sep in (":", "|"):
        try:
            parts = value.rsplit(sep, 5)
            if len(parts) == 6:
                user_id, role, modules, safe_name, expiry_text, sig = parts
                payload = f"{user_id}{sep}{role}{sep}{modules}{sep}{safe_name}{sep}{expiry_text}"
            elif len(parts) == 5:
                user_id, role, modules, safe_name, sig = parts
                payload = f"{user_id}{sep}{role}{sep}{modules}{sep}{safe_name}"
                expiry_text = "2000000000"
            else:
                continue
                
            secrets_to_try = [WEB_SECRET]
                
            matched = False
            import hmac
            import hashlib
            for secret in secrets_to_try:
                expected = hmac.new(secret.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
                if hmac.compare_digest(expected, sig):
                    matched = True
                    break
                    
            if not matched:
                continue
            import datetime as dt
            if float(expiry_text) < dt.datetime.now().timestamp():
                return None
                
            import urllib.parse
            return urllib.parse.unquote(user_id), urllib.parse.unquote(role), urllib.parse.unquote(modules), urllib.parse.unquote(safe_name)
        except Exception:
            continue
    return None

def parse_cookie_values(handler, name: str) -> list[str]:
    raw = handler.headers.get("Cookie", "")
    values = []
    for part in raw.split(";"):
        if "=" not in part: continue
        k, v = part.split("=", 1)
        if k.strip() == name: values.append(v.strip())
    return values


def current_session(handler: BaseHTTPRequestHandler) -> tuple[str, str, str, str] | None:
    for token in parse_cookie_values(handler, SESSION_COOKIE):
        session = _verify_cookie(token)
        if session:
            user_id, role, modules, safe_name = session
            SESSIONS[token] = (user_id, role, modules, safe_name, dt.datetime.now().timestamp())
            return session
        else:
            try:
                with open(ROOT.parent / "data" / "grafik_auth.log", "a", encoding="utf-8") as f:
                    f.write(f"Token verification failed for token: {token}\n")
            except Exception:
                pass
    return None


def get_mod_role(session: tuple[str, str, str, str] | None, mod_name: str) -> str | None:
    if not session: return None
    username = session[0]
    role = session[1]
    modules = session[2]

    resolved = resolve_user_access(ROOT.parent / "base" / "web_users.db", username, role, modules)
    if not resolved:
        return None
    role, modules = resolved
    return module_role(role, modules, mod_name)

def require_auth(handler: BaseHTTPRequestHandler, need_edit: bool = False) -> bool:
    if not AUTH_ENABLED:
        return True
    session = current_session(handler)
    mod_role = get_mod_role(session, "grafik_ppr")
    if mod_role and (not need_edit or mod_role in ("edit", "editor", "admin")):
        return True
    handler.send_response(HTTPStatus.UNAUTHORIZED)
    handler.send_header("Content-Type", "text/plain; charset=utf-8")
    handler.send_header("WWW-Authenticate", 'Form realm="Grafik PPR"')
    handler.end_headers()
    handler.wfile.write("Требуется вход".encode("utf-8"))
    return False


def render_page(state: dict, can_edit: bool, username: str | None) -> str:
    state_json = json.dumps(state, ensure_ascii=False).replace("</", "<\\/")
    employees_json = json.dumps(get_all_employee_names(), ensure_ascii=False).replace("</", "<\\/")
    employee_vacations_json = json.dumps(get_employee_vacations(), ensure_ascii=False).replace("</", "<\\/")
    started_at = SERVER_STARTED_AT.strftime("%H:%M:%S %d.%m.%Y") if SERVER_STARTED_AT else "неизвестно"
    toolbar = EDIT_TOOLBAR if can_edit else READONLY_TOOLBAR
    with open(ROOT / "templates" / "index.html", "r", encoding="utf-8") as f:
        html_template = f.read()
    return (
        html_template.replace("{{STATE_JSON}}", state_json)
        .replace("{{EMPLOYEE_NAMES}}", employees_json)
        .replace("{{EMPLOYEE_VACATIONS}}", employee_vacations_json)
        .replace("{{STARTED_AT}}", started_at)
        .replace("{{APP_VERSION}}", APP_VERSION)
        .replace("{{TOOLBAR}}", toolbar)
        .replace("{{CAN_EDIT}}", "true" if can_edit else "false")
        .replace("{{APP_PREFIX}}", APP_PREFIX)
        .replace("{{TEM_NORM_ROWS}}", json.dumps(TEM_NORM_ROWS, ensure_ascii=False))
        .replace("{{AGR_NORM_ROWS}}", json.dumps(AGR_NORM_ROWS, ensure_ascii=False))
    )


def _route_path(path: str) -> str:
    if path == APP_PREFIX:
        return "/grafik-ppr"
    if path.startswith(APP_PREFIX + "/"):
        return path[len(APP_PREFIX):]
    return path


def render_login(extra: str = "") -> str:
    with open(ROOT / "templates" / "login.html", "r", encoding="utf-8") as f:
        login_template = f.read()
    return (
        login_template.replace("{{USER}}", WEB_USER)
        .replace("{{APP_PREFIX}}", APP_PREFIX)
        + extra
    )


EDIT_TOOLBAR = """
      <div class="toolbar">
        <label>Год <select id="yearInput" onchange="loadYearFromInput()"></select></label>
        <button onclick="openReport()">Отчет</button>
        <button id="saveButton" onclick="saveState()">Сохранить</button>
        <div class="json-menu" id="jsonMenuWrap">
          <button type="button" onclick="toggleJsonMenu(event)">JSON</button>
          <div class="json-menu-panel" id="jsonMenuPanel" aria-hidden="true">
            <button type="button" onclick="downloadJson(); closeJsonMenu()">Экспорт JSON</button>
            <button type="button" onclick="triggerImportJson()">Импорт JSON</button>
          </div>
        </div>
        <input id="importFile" type="file" accept=".json,application/json" style="display:none" onchange="importJson(event)">
      </div>
"""

READONLY_TOOLBAR = """
      <div class="toolbar">
        <label>Год <select id="yearInput" onchange="loadYearFromInput()"></select></label>
        <button onclick="openReport()">Отчет</button>
      </div>
"""






from fastapi import FastAPI, Request, Response, Depends, Form, HTTPException, Cookie, status
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import uvicorn

app = FastAPI(title="RTPS Grafik PPR")

@app.middleware("http")
async def strip_prefix(request: Request, call_next):
    if request.scope["path"].startswith(APP_PREFIX + "/"):
        request.scope["path"] = request.scope["path"][len(APP_PREFIX):]
    return await call_next(request)

app.mount("/static", StaticFiles(directory=str(ROOT / "static")), name="static")

def _login_cookie(username: str, role: str) -> str:
    token = _cookie_value(username, role)
    SESSIONS[token] = (username, role, "", username, dt.datetime.now().timestamp())
    return f"{SESSION_COOKIE}={token}; HttpOnly; Path=/; SameSite=Lax"

def json_response(data: dict | list, status_code: int = 200) -> JSONResponse:
    return JSONResponse(content=data, status_code=status_code)

def get_current_session(request: Request):
    cookie = request.cookies.get(SESSION_COOKIE)
    if cookie:
        return _verify_cookie(cookie)
    return None

def require_auth_fastapi(request: Request, need_edit: bool = False):
    if not AUTH_ENABLED:
        return True, None
    session = get_current_session(request)
    role = get_mod_role(session, "grafik_ppr")
    if not role:
        return False, None
    if need_edit and role not in ("edit", "editor", "admin"):
        return False, None
    return True, session

@app.get("/grafik-ppr", response_class=HTMLResponse)
@app.get("/", response_class=HTMLResponse)
async def home_route(request: Request, year: int = None):
    if year is None:
        year = dt.date.today().year
    session = get_current_session(request)
    user = session[0] if session else None
    mod_role = get_mod_role(session, "grafik_ppr")
    
    try:
        raw_cookie = request.cookies.get(SESSION_COOKIE, "")
        with open(ROOT.parent / "data" / "grafik_auth.log", "a", encoding="utf-8") as f:
            f.write(f"Access /grafik-ppr. Session: {session}. Cookie: {raw_cookie}\n")
    except Exception:
        pass
        
    if AUTH_ENABLED and not mod_role:
        return RedirectResponse("/login", status_code=303)
        
    can_edit = mod_role in ("edit", "editor", "admin") if AUTH_ENABLED else True
    
    # We will reuse the original render_page logic, but it returned a string. We return HTMLResponse.
    html_content = render_page(load_state(year), can_edit, user)
    response = HTMLResponse(content=html_content)
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response

@app.get("/login", response_class=HTMLResponse)
async def login_get(request: Request):
    return RedirectResponse("/login", status_code=303)

@app.post("/login", response_class=HTMLResponse)
async def login_post(request: Request, user: str = Form(""), password: str = Form("")):
    return RedirectResponse("/login", status_code=303)

@app.get("/logout")
async def logout_route():
    resp = RedirectResponse("/login", status_code=303)
    resp.delete_cookie(SESSION_COOKIE, path="/", httponly=True, samesite="lax")
    return resp

@app.get("/api/state")
async def get_state(year: int = None):
    if year is None: year = dt.date.today().year
    return json_response(load_state(year))

@app.post("/api/state")
@app.post("/api/import")
async def post_state(request: Request):
    auth_ok, session = require_auth_fastapi(request, need_edit=True)
    if not auth_ok:
        return json_response({"error": "Unauthorized"}, status_code=401)
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    try:
        saved = save_state(payload)
    except Exception as exc:
        return json_response({"error": str(exc)}, status_code=400)
    return json_response(saved)

@app.get("/api/export")
async def export_state(request: Request, year: int = None):
    auth_ok, session = require_auth_fastapi(request)
    if not auth_ok:
        return Response("Unauthorized", status_code=401)
    if year is None: year = dt.date.today().year
    body = json.dumps(load_state(year), ensure_ascii=False, indent=2).encode("utf-8")
    return Response(
        content=body, 
        media_type="application/json", 
        headers={"Content-Disposition": f'attachment; filename="grafik_ppr_{year}.json"'}
    )

@app.get("/api/act-export")
async def export_act(request: Request, year: int = None, month: str = "", act: str = ""):
    auth_ok, session = require_auth_fastapi(request)
    if not auth_ok:
        return Response("Unauthorized", status_code=401)
    if year is None: year = dt.date.today().year
    try:
        body, filename = build_act_workbook(year, act.strip())
    except Exception as exc:
        return json_response({"error": str(exc)}, status_code=400)
    return Response(
        content=body,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": content_disposition_attachment(filename)}
    )

@app.get("/api/report-export")
async def export_report(request: Request, year: int = None, month: str = ""):
    auth_ok, session = require_auth_fastapi(request)
    if not auth_ok:
        return Response("Unauthorized", status_code=401)
    if year is None: year = dt.date.today().year
    if not month.strip(): month = MONTHS_RU[dt.date.today().month - 1]
    try:
        body, filename = build_report_workbook(year, month.strip())
    except Exception as exc:
        return json_response({"error": str(exc)}, status_code=400)
    return Response(
        content=body,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": content_disposition_attachment(filename)}
    )

@app.get("/api/report-preview")
async def preview_report(request: Request, year: int = None, month: str = ""):
    auth_ok, session = require_auth_fastapi(request)
    if not auth_ok:
        return json_response({"error": "Unauthorized"}, status_code=401)
    if year is None: year = dt.date.today().year
    if not month.strip(): month = MONTHS_RU[dt.date.today().month - 1]
    try:
        state = load_state(year)
        data = calculate_report_data_from_state(state, month.strip())
        saved_notes = state.get("notes", {}).get(month.strip(), {}) or {}
        return json_response(build_report_preview(month.strip(), data, saved_notes))
    except Exception as exc:
        return json_response({"error": str(exc)}, status_code=400)

@app.post("/api/tu28_extra")
async def post_tu28_extra(request: Request):
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    year = payload.get("year")
    month_name = payload.get("month_name")
    r = payload.get("r")
    extra = payload.get("extra", [])
    with DB_LOCK, conn() as db:
        cur = db.cursor()
        with db:
            row = cur.execute("SELECT v FROM tu28_data WHERE y=? AND m=? AND r=? AND k='tu28_locked'", (year, month_name, r)).fetchone()
            is_locked = False
            if row and row["v"]:
                is_locked = json.loads(row["v"])
            if not is_locked:
                cur.execute("INSERT OR REPLACE INTO tu28_data VALUES (?,?,?,?,?)", (year, month_name, r, "tu28_extra", json.dumps(extra)))
    return json_response({"status": "ok"})

@app.post("/api/tu28-export")
async def export_tu28(request: Request):
    auth_ok, session = require_auth_fastapi(request)
    if not auth_ok:
        return Response("Unauthorized", status_code=401)
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    year = int(payload.get("year") or dt.date.today().year)
    month = str(payload.get("month", "")).strip() or MONTHS_RU[dt.date.today().month - 1]
    row_raw = payload.get("row", None)
    if row_raw in (None, ""):
        return json_response({"error": "В месяце нет ремонтов для ТУ-28"}, status_code=400)
    try:
        row_idx = int(row_raw)
    except Exception:
        return json_response({"error": "Не удалось определить строку ремонта"}, status_code=400)
    
    staff_list = payload.get("staff") or []
    if not isinstance(staff_list, list): staff_list = []
    extra_repairs = payload.get("extra_repairs") or []
    if not isinstance(extra_repairs, list): extra_repairs = []
    
    try:
        body, filename = build_tu28_workbook(year, month, row_idx, staff_list, extra_repairs=extra_repairs)
    except Exception as exc:
        return json_response({"error": str(exc)}, status_code=400)
        
    return Response(
        content=body,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": content_disposition_attachment(filename)}
    )

def main() -> None:
    global SERVER_STARTED_AT
    SERVER_STARTED_AT = dt.datetime.now()
    ensure_database()
    host = os.environ.get("WEB_HOST", "127.0.0.1")
    port = int(os.environ.get("WEB_PORT", "8000"))
    url = f"http://{host}:{port}/grafik-ppr"
    print(f"График ППР web ready (FastAPI): {url} | started at {SERVER_STARTED_AT:%H:%M:%S %d.%m.%Y}")
    if host in {"127.0.0.1", "localhost", "0.0.0.0"}:
        import threading, webbrowser
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    uvicorn.run("app:app", host=host, port=port, reload=False)

if __name__ == "__main__":
    main()
