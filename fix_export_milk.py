import re

app_py = r'g:\Мой диск\Codex\rtps\web_tabel\app.py'
with open(app_py, 'r', encoding='utf-8') as f:
    text = f.read()

# I will write a script to rewrite export_milk function.
# Let's extract export_milk boundaries.
import ast
class FuncFinder(ast.NodeVisitor):
    def __init__(self):
        self.start = None
        self.end = None
    def visit_AsyncFunctionDef(self, node):
        if node.name == 'export_milk':
            self.start = node.lineno
            self.end = node.end_lineno
        self.generic_visit(node)

tree = ast.parse(text)
finder = FuncFinder()
finder.visit(tree)

if finder.start is None:
    print("export_milk not found!")
else:
    # Read lines
    lines = text.split('\n')
    # The decorator is before start, but let's just replace the whole function.
    # The decorator is @app.get("/api/export-milk")
    
    start_idx = finder.start - 1
    # Check if there is decorator
    if lines[start_idx-1].startswith('@app.get("/api/export-milk")'):
        start_idx -= 1
        
    end_idx = finder.end
    
    new_func = """@app.get("/api/export-milk")
async def export_milk(year: int, month: int, type: str):
    import urllib.parse
    import tempfile
    import copy
    import os
    import openpyxl
    type = urllib.parse.unquote(type)
    
    templates = {
        "компенсация": "Молоко_комп_шаблон.xlsx",
        "план": "Молоко_план_шаблон.xlsx",
        "факт": "Молоко_факт_шаблон.xlsx"
    }
    
    template_name = templates.get(type)
    if not template_name:
        return {"error": "Invalid template type"}
        
    template_path = ROOT / "resources" / template_name
    if not template_path.exists():
        return {"error": f"Template {template_name} not found"}
        
    sys_dates = load_system_dates(year)
    transfer_dates = sys_dates["transfer"]
    holiday_dates = sys_dates["holiday"]
    
    m_str = MONTH_NAMES[month] if 1 <= month <= 12 else str(month)
    days_cnt = calendar.monthrange(year, month)[1]
    
    with DB_LOCK, connect() as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        emp_rows = cur.execute("SELECT pos, name, tab_num, milk, milk_issue, full_name, milk_note FROM employees WHERE y=? AND name != '' ORDER BY rowid", (year,)).fetchall()
        
        m_comp_set = set()
        m_issue_set = set()
        
        for r in emp_rows:
            name = str(r["name"]).upper()
            if r["milk"]: m_comp_set.add(name)
            if r["milk_issue"]: m_issue_set.add(name)
            
        ts_rows = cur.execute("SELECT tab_num, c, v FROM timesheet WHERE y=? AND m=?", (year, m_str)).fetchall()
        ts_data = {}
        for r in ts_rows:
            ts_data.setdefault(str(r["tab_num"]), {})[int(r["c"])] = str(r["v"]).upper()
            
        norm_row = cur.execute("SELECT v FROM ts_norms_data WHERE y=? AND r=? AND c=2", (year, month - 1)).fetchone()
        work_days_norm = str(norm_row["v"]) if norm_row else "0"
            
    def is_workday(y, m, d):
        dt = calendar.datetime.date(y, m, d)
        is_we = dt.weekday() >= 5
        is_hol = (m, d) in holiday_dates
        is_tr = (m, d) in transfer_dates
        if is_tr: return True
        if is_hol: return False
        return not is_we

    final_list = []
    grand_total = 0

    for emp in emp_rows:
        name = str(emp["name"])
        name_up = name.upper()
        tab_num = str(emp["tab_num"])
        pos = str(emp["pos"])
        full_name = str(emp["full_name"])
        milk_note = str(emp["milk_note"])
        
        if type == "компенсация" and name_up not in m_comp_set: continue
        if type in ["план", "факт"] and name_up not in m_issue_set: continue
        
        count = 0
        emp_ts = ts_data.get(tab_num, {})
        for d in range(1, days_cnt + 1):
            val = emp_ts.get(d, "")
            if type == "компенсация":
                try:
                    float(val.replace(',', '.'))
                    count += 1
                except ValueError:
                    if val in ["В", "РВ", "ДО", "О", "У", "Б"]: 
                        if val in ["РВ"]: count += 1 # is_milk_cell equivalent
            else:
                try:
                    float(val.replace(',', '.'))
                    count += 1
                except ValueError:
                    if not val and is_workday(year, month, d): 
                        count += 1
                        
        final_list.append({
            "fio": name,
            "full_name": full_name,
            "pos": pos,
            "tab": tab_num,
            "shifts": count,
            "milk_note": milk_note
        })
        grand_total += count
        
    # Load workbook
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    row_tpl = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell.value = cell.value.replace("[МЕСЯЦ]", m_str).replace("[ГОД]", str(year)).replace("[НОРМА_ДНЕЙ]", work_days_norm).replace("[ИТОГО]", str(grand_total))
                if row_tpl is None and any(tag in cell.value for tag in [
                    "[№]",
                    "[ФИО]",
                    "[ФИО_ПОЛНОЕ]",
                    "[ДОЛЖНОСТЬ]",
                    "[ТАБ]",
                    "[СМЕНЫ]",
                    "[ПРИМЕЧАНИЕ]"
                ]):
                    row_tpl = cell.row

    if row_tpl is not None and final_list:
        tpl_vals = {c: ws.cell(row=row_tpl, column=c).value for c in range(1, ws.max_column + 1)}
        ws.insert_rows(row_tpl + 1, len(final_list) - 1)
        for i, data in enumerate(final_list):
            curr_r = row_tpl + i
            ws.row_dimensions[curr_r].height = None 
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=curr_r, column=c_idx)
                if i > 0: 
                    src = ws.cell(row=row_tpl, column=c_idx)
                    cell.value = src.value
                    if src.has_style:
                        cell.font = copy.copy(src.font)
                        cell.border = copy.copy(src.border)
                        cell.fill = copy.copy(src.fill)
                        cell.alignment = copy.copy(src.alignment)
                v = tpl_vals.get(c_idx)
                if v and isinstance(v, str):
                    v = v.replace("[№]", str(i+1)).replace("[ФИО]", data["fio"]).replace("[ФИО_ПОЛНОЕ]", data["full_name"]).replace("[ДОЛЖНОСТЬ]", data["pos"]).replace("[ТАБ]", data["tab"]).replace("[СМЕНЫ]", str(data["shifts"])).replace("[ПРИМЕЧАНИЕ]", data["milk_note"])
                    cell.value = int(v) if str(v).isdigit() else v

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    wb.save(tmp.name)
    
    return FileResponse(tmp.name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=f"Отчет_Молоко_{type}.xlsx")
"""
    
    new_text = "\n".join(lines[:start_idx]) + "\n" + new_func + "\n" + "\n".join(lines[end_idx:])
    
    # Bump version
    new_text = re.sub(r'APP_VERSION = "web-tabel-1\.\d+"', 'APP_VERSION = "web-tabel-1.39"', new_text)
    
    with open(app_py, 'w', encoding='utf-8') as f:
        f.write(new_text)
    
    print("Replaced export_milk")
